import os
import tempfile
import sys
import uuid
import json
import re
import logging
import urllib.parse
from collections import Counter
from datetime import datetime, timedelta, date
from io import BytesIO

from flask import (
    Flask, request, redirect, url_for, session, render_template,
    render_template_string, flash, Response, jsonify
)

import openpyxl
from openpyxl import load_workbook

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

# =========================
# CONFIG
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Runtime configuration via environment variables (safe defaults preserve current behavior)
APP_TITLE = os.environ.get("VETAPP_TITLE", "Selim Vet Clinic")
ADMIN_USER = os.environ.get("VETAPP_ADMIN_USER", "Admin")
ADMIN_PASS = os.environ.get("VETAPP_ADMIN_PASS", "1234")

# Prefer writable paths on Vercel (serverless filesystem is read-only except temp dir)
IS_VERCEL = (os.environ.get("VERCEL") == "1") or bool(os.environ.get("VERCEL_ENV"))
TMP_ROOT = tempfile.gettempdir()  # /tmp on Linux, temp path on Windows

DEFAULT_DATA_DIR = os.path.join(TMP_ROOT, "vetapp_data") if IS_VERCEL else os.path.join(BASE_DIR, "data")
DEFAULT_TEMPLATES_DIR = os.path.join(TMP_ROOT, "vetapp_templates") if IS_VERCEL else os.path.join(BASE_DIR, "templates")

DATA_DIR = os.environ.get("VETAPP_DATA_DIR", DEFAULT_DATA_DIR)
TEMPLATES_DIR = os.environ.get("VETAPP_TEMPLATES_DIR", DEFAULT_TEMPLATES_DIR)

# Create dirs (fallback to TMP_ROOT if project root is not writable)
try:
    os.makedirs(DATA_DIR, exist_ok=True)
except Exception:
    DATA_DIR = os.path.join(TMP_ROOT, "vetapp_data")
    os.makedirs(DATA_DIR, exist_ok=True)

try:
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
except Exception:
    TEMPLATES_DIR = os.path.join(TMP_ROOT, "vetapp_templates")
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
LOG_LEVEL = os.environ.get("VETAPP_LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
logger = logging.getLogger("vetapp")

OWNERS_XLSX = os.path.join(DATA_DIR, "owners.xlsx")
PETS_XLSX = os.path.join(DATA_DIR, "pets.xlsx")
BOOKINGS_XLSX = os.path.join(DATA_DIR, "bookings.xlsx")
REMINDERS_XLSX = os.path.join(DATA_DIR, "reminders.xlsx")

WHATSAPP_TEMPLATES_XLSX = os.path.join(DATA_DIR, "whatsapp_templates.xlsx")
ROLES_PERMISSIONS_XLSX = os.path.join(DATA_DIR, "roles_permissions.xlsx")

USERS_XLSX = os.path.join(DATA_DIR, "users.xlsx")
VETS_XLSX = os.path.join(DATA_DIR, "vets.xlsx")
ROOMS_XLSX = os.path.join(DATA_DIR, "rooms.xlsx")
SERVICES_XLSX = os.path.join(DATA_DIR, "services.xlsx")

ROLES = ["admin", "vet", "reception", "staff", "user"]

# Security permissions (used for role-based access control)
PERMISSIONS_CATALOG = [
    ("config_manage", "Manage system configuration"),
    ("users_manage", "Manage users"),
    ("services_manage", "Manage services"),
    ("bookings_manage", "Create/Edit bookings"),
    ("reports_view", "View reports"),
    ("invoices_view", "View/Print invoices"),
    ("whatsapp_manage", "Manage WhatsApp templates"),
]

DEFAULT_ROLE_PERMISSIONS = {
    "admin": [p[0] for p in PERMISSIONS_CATALOG],
    "reception": ["bookings_manage", "invoices_view", "reports_view"],
    "vet": ["bookings_manage", "invoices_view"],
    "staff": ["bookings_manage"],
}

ROLES_PERMISSIONS_HEADERS = ["id", "role", "permissions", "created_at", "updated_at"]

SERVICES = [
    {"name": "General Exam", "fee": 150},
    {"name": "Vaccination", "fee": 120},
    {"name": "Deworming", "fee": 80},
    {"name": "Grooming Basic", "fee": 200},
    {"name": "Grooming Full", "fee": 350},
    {"name": "Dental Cleaning", "fee": 600},
    {"name": "X-Ray", "fee": 450},
    {"name": "Ultrasound", "fee": 500},
    {"name": "Blood Test Panel", "fee": 400},
    {"name": "Wound Dressing", "fee": 180},
]

STATUS_FLOW = ["Scheduled", "Checked-in", "In Treatment", "Completed", "Cancelled", "No-Show"]
PRIORITIES = ["Normal", "Urgent", "Emergency"]
APPOINTMENT_TYPES = ["Consultation", "Vaccination", "Surgery", "Grooming", "Follow-up", "Lab Test", "Other"]
PAYMENT_STATUSES = ["Unpaid", "Paid", "Partial", "Insurance"]
PAYMENT_METHODS = ["Cash", "Card", "Transfer", "Insurance", "Other"]
PAYMENT_CHANNELS = ["Cash", "Visa", "Instapay"]
PAYMENT_CHANNEL_ALIASES = {
    "cash": "Cash",
    "visa": "Visa",
    "insta": "Instapay",
    "instapay": "Instapay",
    "insta pay": "Instapay",
}

CHANNELS = ["Walk-in", "Phone", "App", "WhatsApp", "Email", "Other"]
REMINDER_CHANNELS = ["WhatsApp", "SMS", "Email", "Call", "None"]

VET_NAMES = ["Ahmed", "Zaineb", "Hatem", "Hayaa"]
ROOMS = ["Room 1", "Room 2", "Room 3", "Room 4"]


def active_vet_names():
    """Return active vet names from config, with fallback to default list."""
    try:
        rows = get_vets(include_inactive=False)
        names = [r.get("name", "").strip() for r in rows if (r.get("name") or "").strip()]
        return names if names else VET_NAMES
    except Exception:
        return VET_NAMES


def active_room_names():
    """Return active room names from config, with fallback to default list."""
    try:
        rows = get_rooms(include_inactive=False)
        names = [r.get("name", "").strip() for r in rows if (r.get("name") or "").strip()]
        return names if names else ROOMS
    except Exception:
        return ROOMS


# =========================
# SCHEMAS (with migration)
# =========================
OWNERS_HEADERS = [
    "id", "owner_name", "phone", "email", "address", "preferred_contact", "notes",
    "created_at", "updated_at"
]

PETS_HEADERS = [
    "id", "pet_name", "species", "breed", "sex", "dob", "age_years", "weight_kg",
    "color", "microchip_id", "spayed_neutered", "allergies", "chronic_conditions",
    "vaccinations_summary", "owner_id", "notes",
    "created_at", "updated_at"
]

USERS_HEADERS = [
    "id", "username", "password", "role", "active", "created_at", "updated_at"
]

VETS_HEADERS = [
    "id", "name", "active", "created_at", "updated_at"
]

ROOMS_HEADERS = [
    "id", "name", "active", "created_at", "updated_at"
]
SERVICES_HEADERS = ["id", "name", "cost", "fee", "active", "created_at", "updated_at"]

WHATSAPP_TEMPLATES_HEADERS = [
    "id", "name", "scenario", "booking_type",
    "template_text", "active", "is_default",
    "created_at", "updated_at"
]
WHATSAPP_SCENARIOS = ["Appointment", "Service"]

# FUTURISTIC fields added:
# - portal_token (owner portal link)
# - owner_confirmed, owner_update_message, owner_update_datetime
# - ai_last_applied_at (audit)
BOOKINGS_HEADERS = [
    "id",
    "appointment_start", "duration_min", "appointment_end",
    "owner_id", "pet_id",
    "visit_weight_kg", "visit_temp_c",
    "appointment_type", "priority", "status", "channel",
    "reason", "symptoms",
    "vet_name", "room", "services_json",
    "service_name", "service_fee", "discount_type", "discount_value", "discount", "paid_amount", "due_amount",
    "fee_amount", "payment_status", "payment_method", "payment_channel", "invoice_no",
    "diagnosis", "treatment_plan", "prescription", "lab_tests", "vaccines_given",
    "followup_datetime",
    "reminder_channel", "reminder_sent", "reminder_last_opened",
    "portal_token", "owner_confirmed", "owner_update_message", "owner_update_datetime",
    "ai_last_applied_at",
    "notes",
    "created_at", "updated_at"
]

REMINDERS_HEADERS = [
    "id", "booking_id", "owner_id", "pet_id",
    "reminder_type", "service_name",
    "channel", "status", "scheduled_for", "opened_at", "sent_at",
    "message",
    "created_at", "updated_at"
]

# =========================
# FLASK APP
# =========================
app = Flask(__name__, template_folder=TEMPLATES_DIR)
app.secret_key = os.environ.get("VETAPP_SECRET_KEY", "elite-vet-secret-key-change-me")

# Safe cookie defaults (do NOT force Secure=True because many deployments run on HTTP internally)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)


# =========================
# UTILITIES
# =========================
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_get(d, k, default=""):
    try:
        v = d.get(k, default)
        return "" if v is None else v
    except Exception:
        return default


def to_float(v, default=0.0):
    try:
        if v is None: return default
        s = str(v).strip()
        if not s: return default
        return float(s)
    except Exception:
        return default


def normalize_payment_channel(raw: str) -> str:
    """Normalize payment channel input (case-insensitive), keeping safe defaults."""
    v = (raw or "").strip()
    if not v:
        return ""
    key = re.sub(r"\s+", " ", v).strip().lower()
    if key in PAYMENT_CHANNEL_ALIASES:
        return PAYMENT_CHANNEL_ALIASES[key]
    # If it matches one of the configured channels (case-insensitive), return canonical value
    for opt in PAYMENT_CHANNELS:
        if key == opt.lower():
            return opt
    return v


def validated_discount(subtotal: float, raw_discount: str) -> float:
    """Return a safe discount amount between 0 and subtotal."""
    disc = round(to_float(raw_discount, 0.0), 2)
    if disc < 0:
        disc = 0.0
    if subtotal is None:
        return disc
    if disc > float(subtotal):
        disc = float(subtotal)
    return round(disc, 2)


def normalize_discount_type(raw: str) -> str:
    raw = (raw or "").strip().lower()
    if raw in ("percent", "percentage", "%", "pct"):
        return "percent"
    return "value"


def compute_discount_amount(subtotal: float, discount_type: str = "value", discount_value: str = "",
                            legacy_discount: str = "") -> float:
    """Compute discount amount from either:
       - discount_type + discount_value (preferred)
       - legacy_discount (amount) as fallback
    """
    st = float(subtotal or 0.0)
    dtype = normalize_discount_type(discount_type)
    amt = 0.0
    if dtype == "percent":
        pct = to_float(discount_value, 0.0)
        if pct < 0:
            pct = 0.0
        if pct > 100:
            pct = 100.0
        amt = st * (pct / 100.0)
    else:
        src_val = discount_value if str(discount_value).strip() != "" else legacy_discount
        amt = to_float(src_val, 0.0)
        if amt < 0:
            amt = 0.0
    if amt > st:
        amt = st
    if amt < 0:
        amt = 0.0
    return round(amt, 2)


def normalize_dt(s: str) -> str:
    if not s: return ""
    s = str(s).strip().replace("T", " ")
    if not s: return ""
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""


def parse_dt(s: str):
    s = normalize_dt(s)
    if not s: return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M")
    except Exception:
        return None


def dt_to_local(s: str) -> str:
    """Convert normalized 'YYYY-MM-DD HH:MM' to HTML datetime-local value 'YYYY-MM-DDTHH:MM'."""
    s = normalize_dt(s)
    return s.replace(" ", "T") if s else ""


def parse_services_json(raw) -> list:
    """Best-effort parser for services_json into a normalized list of {name, fee, qty, reminder_at}."""
    try:
        if raw is None:
            return []
        s = raw
        if isinstance(raw, str):
            s = raw.strip()
            if not s:
                return []
            data = json.loads(s)
        else:
            data = raw
        if not isinstance(data, list):
            return []
        out = []
        for it in data:
            if not isinstance(it, dict):
                continue
            name = str(it.get("name") or it.get("service_name") or it.get("service") or it.get("title") or "").strip()
            if not name:
                continue
            fee = round(to_float(it.get("fee", it.get("price", it.get("amount", 0.0))), 0.0), 2)
            qty_raw = it.get("qty", it.get("quantity", 1))
            try:
                qty = int(float(qty_raw))
            except Exception:
                qty = 1
            if qty < 1:
                qty = 1
            reminder_at = normalize_dt(
                it.get("reminder_at") or it.get("reminder_date") or it.get("reminder_datetime") or "")
            out.append({
                "name": name,
                "fee": fee,
                "qty": qty,
                "reminder_at": reminder_at
            })
        return out
    except Exception:
        return []


def services_subtotal(services: list) -> float:
    total = 0.0
    for it in (services or []):
        try:
            qty = int(float(it.get("qty", 1)))
        except Exception:
            qty = 1
        if qty < 1:
            qty = 1
        fee = to_float(it.get("fee", 0.0), 0.0)
        total += (qty * fee)
    return round(total, 2)


def services_summary_name(services: list) -> str:
    if not services:
        return ""
    first = str(services[0].get("name", "") or "").strip()
    if len(services) == 1:
        return first
    return f"{first} +{len(services) - 1}"


def update_booking_services_and_amounts(booking_id: str, services: list) -> bool:
    """Persist services_json + recompute amounts (subtotal/discount/vat/total/due) for a booking.

    Notes:
      - fee_amount represents the net total AFTER discount (and BEFORE VAT display; VAT is applied in due/reporting).
      - due_amount is computed using (net_total + VAT) - paid_amount.
    """
    booking_id = str(booking_id or "").strip()
    if not booking_id:
        return False

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = next((x for x in bookings_rows if str(x.get("id", "")).strip() == booking_id), None)
    if not b:
        return False

    subtotal = services_subtotal(services)
    discount = compute_discount_amount(subtotal, b.get("discount_type"), b.get("discount_value"), b.get("discount"))
    net_total = round(subtotal - discount, 2)

    vat_calc = round(net_total * VAT_RATE, 2)
    total_calc = round(net_total + vat_calc, 2)

    paid = round(to_float(b.get("paid_amount"), 0.0), 2)
    due_calc = round(total_calc - paid, 2)
    if due_calc < 0:
        due_calc = 0.0

    # payment status
    payment_status = ""
    if total_calc > 0:
        if paid <= 0:
            payment_status = "Unpaid"
        elif due_calc <= 0:
            payment_status = "Paid"
        else:
            payment_status = "Partial"

    update_row_by_id(
        BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id,
        {
            "services_json": json.dumps(services or [], ensure_ascii=False),
            "service_name": services_summary_name(services or []),
            "service_fee": round(subtotal, 2),
            "discount": round(discount, 2),
            "fee_amount": round(net_total, 2),
            "due_amount": round(due_calc, 2),
            "payment_status": payment_status,
            "updated_at": now_str(),
        }
    )
    return True


def upsert_service_reminders(booking_row: dict, decorated_booking: dict, services: list) -> None:
    """Create/update reminder rows for per-service reminders under the current booking."""
    if not booking_row:
        return
    booking_id = str(booking_row.get("id", "")).strip()
    if not booking_id:
        return

    selected = {(str(s.get("name", "")).strip().lower()) for s in (services or []) if str(s.get("name", "")).strip()}
    rem_rows = read_all(REMINDERS_XLSX)
    now = now_str()

    # cancel reminders for removed services
    for rr in rem_rows:
        if str(rr.get("booking_id", "")).strip() != booking_id:
            continue
        if (rr.get("reminder_type") or "").strip() != "Service":
            continue
        sn = str(rr.get("service_name", "")).strip().lower()
        if sn and sn not in selected and (rr.get("status") or "") not in ("Sent",):
            rid = str(rr.get("id", "")).strip()
            if rid:
                update_row_by_id(REMINDERS_XLSX, REMINDERS_HEADERS, rid, {"status": "Cancelled", "updated_at": now})

    channel = str(booking_row.get("reminder_channel", "") or "WhatsApp").strip() or "WhatsApp"
    owner_name = str(decorated_booking.get("owner_name", "") or "").strip()
    pet_name = str(decorated_booking.get("pet_name", "") or "").strip()
    portal = str(decorated_booking.get("portal_link", "") or "").strip()

    # upsert for each selected service
    for s in (services or []):
        name = str(s.get("name", "") or "").strip()
        if not name:
            continue
        reminder_at = normalize_dt(s.get("reminder_at", ""))
        # locate existing
        existing = next(
            (rr for rr in rem_rows
             if str(rr.get("booking_id", "")).strip() == booking_id
             and (rr.get("reminder_type") or "").strip() == "Service"
             and str(rr.get("service_name", "")).strip().lower() == name.lower()),
            None
        )

        if not reminder_at:
            if existing and (existing.get("status") or "") not in ("Sent",):
                rid = str(existing.get("id", "")).strip()
                if rid:
                    update_row_by_id(REMINDERS_XLSX, REMINDERS_HEADERS, rid,
                                     {"status": "Cancelled", "scheduled_for": "", "updated_at": now})
            continue

        details = format_booking_details(booking_row, services=services)
        msg = service_reminder_message_template(
            owner_name=owner_name,
            pet_name=pet_name,
            service_name=name,
            scheduled_for=reminder_at,
            portal_link=portal,
            booking_details=details,
            booking_type=str(decorated_booking.get("appointment_type", "Any") or "Any"),
        )

        row = {
            "id": str(existing.get("id")) if existing else str(uuid.uuid4()),
            "booking_id": booking_id,
            "owner_id": str(booking_row.get("owner_id", "") or ""),
            "pet_id": str(booking_row.get("pet_id", "") or ""),
            "reminder_type": "Service",
            "service_name": name,
            "channel": channel,
            "status": "Scheduled",
            "scheduled_for": reminder_at,
            "opened_at": existing.get("opened_at", "") if existing else "",
            "sent_at": existing.get("sent_at", "") if existing else "",
            "message": msg,
            "created_at": existing.get("created_at", now) if existing else now,
            "updated_at": now,
        }

        if existing:
            rid = str(existing.get("id", "")).strip()
            if rid:
                update_row_by_id(REMINDERS_XLSX, REMINDERS_HEADERS, rid, row)
        else:
            append_row(REMINDERS_XLSX, REMINDERS_HEADERS, row)


def ensure_workbook(path, headers):
    if os.path.exists(path):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(12, len(h) + 2), 28)
    wb.save(path)


def ensure_headers(path, desired_headers):
    """
    Non-destructive schema migration for existing Excel files:
    - If file doesn't exist -> create.
    - If headers missing -> append missing columns and fill blanks for all rows.
    """
    if not os.path.exists(path):
        ensure_workbook(path, desired_headers)
        return

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if ws.max_row < 1:
        ws.append(desired_headers)
        wb.save(path)
        return

    existing = [c.value for c in ws[1]]
    existing = [str(x).strip() if x is not None else "" for x in existing]
    missing = [h for h in desired_headers if h not in existing]

    if not missing:
        wb.save(path)
        return

    # append missing headers
    start_col = len(existing) + 1
    for j, h in enumerate(missing):
        c = ws.cell(row=1, column=start_col + j)
        c.value = h
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col + j)].width = min(max(12, len(h) + 2), 28)

    # fill blanks for all existing rows
    for r in range(2, ws.max_row + 1):
        for j in range(len(missing)):
            ws.cell(row=r, column=start_col + j).value = ""

    wb.save(path)


def read_all(path):
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None:
            continue
        item = {}
        empty = True
        for i, h in enumerate(headers):
            v = r[i] if i < len(r) else ""
            if v not in (None, ""):
                empty = False
            item[h] = "" if v is None else v
        if not empty:
            out.append(item)
    return out


def append_row(path, *args):
    """Append a row into an XLSX file.

    Supported calls:
      - append_row(path, row_dict)
      - append_row(path, headers, row_dict)

    If headers are not provided, headers will be inferred from row_dict keys.
    """
    if len(args) == 1:
        row_dict = args[0]
        headers = list(row_dict.keys())
    elif len(args) == 2:
        headers, row_dict = args
    else:
        raise TypeError("append_row() expects (path, row_dict) or (path, headers, row_dict)")

    ensure_headers(path, headers)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    file_headers = [cell.value for cell in ws[1]]
    ws.append([row_dict.get(h, "") for h in file_headers])

    wb.save(path)


def update_row_by_id(path, *args):
    """Update a row (by id) in an XLSX.

    Supported calls:
      - update_row_by_id(path, row_id, updates_dict)
      - update_row_by_id(path, headers, row_id, updates_dict)

    The 3-arg form is used by the Config toggles (Users/Vets/Rooms) and will
    auto-infer/extend headers safely.
    """
    if len(args) == 2:
        row_id, updates = args
        if not isinstance(updates, dict):
            raise TypeError("update_row_by_id(path, row_id, updates_dict) requires updates_dict as dict")
        # Infer current headers if file exists, then extend with update keys + id.
        desired_headers = []
        if os.path.exists(path):
            try:
                wb_tmp = openpyxl.load_workbook(path)
                ws_tmp = wb_tmp.active
                desired_headers = [c.value for c in ws_tmp[1] if c.value]
            except Exception:
                desired_headers = []
        # Always ensure id exists
        if "id" not in desired_headers:
            desired_headers = ["id"] + [h for h in desired_headers if h != "id"]
        for k in updates.keys():
            if k not in desired_headers:
                desired_headers.append(k)
        # Keep common timestamps if present / needed
        for k in ("updated_at", "created_at"):
            if k in updates and k not in desired_headers:
                desired_headers.append(k)

        ensure_headers(path, desired_headers)
        headers = desired_headers

    elif len(args) == 3:
        headers, row_id, updates = args
        if not isinstance(headers, (list, tuple)):
            raise TypeError("update_row_by_id(path, headers, row_id, updates_dict): headers must be list/tuple")
        if not isinstance(updates, dict):
            raise TypeError("update_row_by_id(path, headers, row_id, updates_dict): updates_dict must be dict")
        ensure_headers(path, list(headers))
    else:
        raise TypeError(
            "update_row_by_id() expects (path, row_id, updates_dict) or (path, headers, row_id, updates_dict)")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    if "id" not in hdr:
        wb.save(path)
        return False

    id_idx = hdr.index("id") + 1
    updated = False
    for ridx in range(2, ws.max_row + 1):
        cell = ws.cell(row=ridx, column=id_idx).value
        if str(cell) == str(row_id):
            for k, v in updates.items():
                if k in hdr:
                    ws.cell(row=ridx, column=hdr.index(k) + 1).value = v
            updated = True
            break

    wb.save(path)
    return updated


def delete_row_by_id(path, row_id):
    if not os.path.exists(path):
        return False
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdr = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if "id" not in hdr:
        wb.save(path)
        return False
    id_idx = hdr.index("id") + 1
    for ridx in range(2, ws.max_row + 1):
        cell = ws.cell(row=ridx, column=id_idx).value
        if str(cell) == str(row_id):
            ws.delete_rows(ridx, 1)
            wb.save(path)
            return True
    wb.save(path)
    return False


def find_by_id(rows, row_id):
    for r in rows:
        if str(r.get("id", "")) == str(row_id):
            return r
    return None


def norm_username(u: str) -> str:
    return (u or "").strip().lower()


def _boolish(v):
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "on")


def get_users(include_inactive=False):
    rows = read_all(USERS_XLSX)
    out = []
    for r in rows:
        if not include_inactive and not _boolish(r.get("active", "1")):
            continue
        r["username"] = norm_username(r.get("username", ""))
        r["role"] = (r.get("role") or "user").strip().lower()
        out.append(r)
    return out


def get_user_by_username(username):
    un = norm_username(username)
    for r in get_users(include_inactive=True):
        if norm_username(r.get("username", "")) == un:
            return r
    return None


def ensure_admin_user_present():
    # Ensure admin exists in users file for visibility in Config page
    users = read_all(USERS_XLSX)
    for u in users:
        if norm_username(u.get("username", "")) == norm_username(ADMIN_USER):
            return
    append_row(USERS_XLSX, {
        "id": str(uuid.uuid4()),
        "username": norm_username(ADMIN_USER),
        "password": ADMIN_PASS,
        "role": "admin",
        "active": "1",
        "created_at": now_str(),
        "updated_at": now_str()
    })


def get_vets(include_inactive=False):
    rows = read_all(VETS_XLSX)
    out = []
    for r in rows:
        if not include_inactive and not _boolish(r.get("active", "1")):
            continue
        name = (r.get("name") or "").strip()
        if name:
            out.append({**r, "name": name})
    # sort by name
    out.sort(key=lambda x: x.get("name", "").lower())
    return out


def get_rooms(include_inactive=False):
    rows = read_all(ROOMS_XLSX)
    out = []
    for r in rows:
        if not include_inactive and not _boolish(r.get("active", "1")):
            continue
        name = (r.get("name") or "").strip()
        if name:
            out.append({**r, "name": name})
    out.sort(key=lambda x: x.get("name", "").lower())
    return out


def get_services(include_inactive=False):
    rows = read_all(SERVICES_XLSX)
    out = []
    for r in rows:
        if not include_inactive and not _boolish(r.get("active", "1")):
            continue
        name = (r.get("name") or "").strip()
        if not name:
            continue

        # Customer selling price
        try:
            fee = float(r.get("fee", 0) or 0)
        except Exception:
            fee = 0.0

        # Internal cost
        try:
            cost = float(r.get("cost", 0) or 0)
        except Exception:
            cost = 0.0

        fee = round(fee, 2)
        cost = round(cost, 2)
        margin = round(fee - cost, 2)
        margin_pct = round((margin / fee * 100.0), 2) if fee else 0.0

        out.append({**r, "name": name, "fee": fee, "cost": cost, "margin": margin, "margin_pct": margin_pct})
    out.sort(key=lambda x: x.get("name", "").lower())
    return out


def active_services():
    # returns list of dicts for templates
    # margin is derived (fee - cost)
    services = get_services(include_inactive=False)
    return [{"name": s["name"], "fee": s.get("fee", 0.0), "cost": s.get("cost", 0.0),
             "margin": round(float(s.get("fee", 0.0)) - float(s.get("cost", 0.0)), 2)} for s in services]


def seed_config_defaults():
    # Users
    ensure_admin_user_present()


# Seed default WhatsApp templates (if empty)
try:
    wa_rows = read_all(WHATSAPP_TEMPLATES_XLSX)
except Exception:
    wa_rows = []
if not wa_rows:
    nowv = now_str()
    defaults = [
        {
            "id": str(uuid.uuid4()),
            "name": "Default - Appointment Reminder",
            "scenario": "Appointment",
            "booking_type": "Any",
            "template_text": (
                "Hello {owner_name}, this is a reminder for your booking.\n"
                "Service: {service_name}\n"
                "Date/Time: {appointment_start}\n"
                "Pet: {pet_name}\n"
                "Notes: {reason}\n"
                "Thank you."
            ),
            "is_default": True,
            "active": True,
            "created_at": nowv,
            "updated_at": nowv,
        },
        {
            "id": str(uuid.uuid4()),
            "name": "Default - Service Reminder",
            "scenario": "Service",
            "booking_type": "Any",
            "template_text": (
                "Hello {owner_name},\n"
                "Service reminder: {service_name}\n"
                "{booking_details}\n"
                "Thank you."
            ),
            "is_default": False,
            "active": True,
            "created_at": nowv,
            "updated_at": nowv,
        },
    ]
    for row in defaults:
        append_row(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS, row)

# Seed role permissions (if empty)
try:
    rp_rows = read_all(ROLES_PERMISSIONS_XLSX)
except Exception:
    rp_rows = []
if not rp_rows:
    nowv = now_str()
    for r in ROLES:
        perms = DEFAULT_ROLE_PERMISSIONS.get(r, [])
        append_row(ROLES_PERMISSIONS_XLSX, ROLES_PERMISSIONS_HEADERS, {
            "id": str(uuid.uuid4()),
            "role": r,
            "permissions": ",".join(perms),
            "created_at": nowv,
            "updated_at": nowv,
        })

    # Vets
    if len(get_vets(include_inactive=False)) == 0:
        for name in ["ahmed", "zaineb", "hatem", "hayaa"]:
            append_row(VETS_XLSX, {
                "id": str(uuid.uuid4()),
                "name": name,
                "active": "1",
                "created_at": now_str(),
                "updated_at": now_str()
            })

    # Rooms
    if len(get_rooms(include_inactive=False)) == 0:
        for name in ["Room 1", "Room 2", "Room 3", "Room 4"]:
            append_row(ROOMS_XLSX, {
                "id": str(uuid.uuid4()),
                "name": name,
                "active": "1",
                "created_at": now_str(),
                "updated_at": now_str()
            })

    # Services
    if len(get_services(include_inactive=False)) == 0:
        for s in SERVICES:
            try:
                fee = float(s.get("fee", 0))
            except Exception:
                fee = 0.0
            append_row(SERVICES_XLSX, {
                "id": str(uuid.uuid4()),
                "name": (s.get("name", "") or "").strip(),
                "fee": round(fee, 2),
                "active": "1",
                "created_at": now_str(),
                "updated_at": now_str()
            })


def get_whatsapp_templates(include_inactive: bool = False) -> list:
    rows = read_all(WHATSAPP_TEMPLATES_XLSX)
    if include_inactive:
        return rows
    return [r for r in rows if _boolish(r.get("active", "1"))]


def select_whatsapp_template(scenario: str, booking_type: str = "Any") -> dict:
    scenario = (scenario or "Appointment").strip()
    booking_type = (booking_type or "Any").strip() or "Any"
    rows = get_whatsapp_templates(include_inactive=False)

    def _bt(r):
        return str(r.get("booking_type", "") or "Any").strip() or "Any"

    def _is_default(r):
        return _boolish(r.get("is_default", ""))

    # 1) default exact booking type
    for r in rows:
        if str(r.get("scenario", "") or "").strip() == scenario and _bt(r) == booking_type and _is_default(r):
            return r
    # 2) default Any
    for r in rows:
        if str(r.get("scenario", "") or "").strip() == scenario and _bt(r) in ("Any", "") and _is_default(r):
            return r
    # 3) any active exact booking type
    for r in rows:
        if str(r.get("scenario", "") or "").strip() == scenario and _bt(r) == booking_type:
            return r
    # 4) any active Any
    for r in rows:
        if str(r.get("scenario", "") or "").strip() == scenario and _bt(r) in ("Any", ""):
            return r
    return {}


def format_booking_details(b: dict, services: list = None) -> str:
    b = b or {}
    parts = []
    for label, key in [
        ("Type", "appointment_type"),
        ("Priority", "priority"),
        ("Channel", "channel"),
        ("Status", "status"),
    ]:
        v = str(b.get(key, "") or "").strip()
        if v:
            parts.append(f"{label}: {v}")

    reason = str(b.get("reason", "") or "").strip()
    if reason:
        parts.append(f"Reason: {reason}")

    symptoms = str(b.get("symptoms", "") or "").strip()
    if symptoms:
        parts.append(f"Symptoms: {symptoms}")

    if services:
        items = []
        for it in services:
            if not isinstance(it, dict):
                continue
            name = str(it.get("name", "") or "").strip()
            if not name:
                continue
            qty = it.get("qty", 1) or 1
            try:
                qty = max(1, int(float(qty)))
            except Exception:
                qty = 1
            items.append(f"{name} x{qty}" if qty > 1 else name)
        if items:
            parts.append("Services: " + ", ".join(items))

    fee = str(b.get("fee_amount", "") or "").strip()
    if fee:
        parts.append(f"Total: {fee}")

    return "\n".join(parts).strip()


def build_whatsapp_message_text(
        scenario: str,
        booking_type: str,
        owner_name: str,
        pet_name: str,
        service_name: str = "",
        appointment_start: str = "",
        scheduled_for: str = "",
        booking_details: str = "",
        portal_link: str = "",
) -> str:
    tpl = select_whatsapp_template(scenario=scenario, booking_type=booking_type)
    template_text = (tpl.get("template_text") if isinstance(tpl, dict) else "") or ""

    if not template_text:
        if str(scenario).strip() == "Service":
            template_text = (
                "Hello {owner_name},\n"
                "This is {clinic_name}.\n"
                "Reminder: {pet_name} has {service_name} on {scheduled_for}.\n"
                "{booking_details}\n"
                "{portal_link_line}\n"
                "Thank you."
            )
        else:
            template_text = (
                "Hello {owner_name},\n"
                "This is {clinic_name}.\n"
                "Reminder: {pet_name} appointment on {appointment_start}.\n"
                "Service: {service_name}\n"
                "{booking_details}\n"
                "{portal_link_line}\n"
                "Thank you."
            )

    portal_link_line = f"Portal: {portal_link}" if portal_link else ""
    ctx = {
        "clinic_name": APP_TITLE,
        "owner_name": owner_name or "",
        "pet_name": pet_name or "",
        "service_name": service_name or "",
        "appointment_start": appointment_start or "",
        "scheduled_for": scheduled_for or "",
        "booking_details": booking_details or "",
        "portal_link": portal_link or "",
        "portal_link_line": portal_link_line,
    }
    msg = _safe_format_template(template_text, ctx).strip()

    # Enforce service + booking details + portal line
    if (service_name or "").strip() and ("{service_name}" not in template_text) and ("Service:" not in msg):
        msg = (msg + "\n" if msg else "") + f"Service: {service_name}"
    if (booking_details or "").strip() and ("{booking_details}" not in template_text) and (booking_details not in msg):
        msg = (msg + "\n" if msg else "") + booking_details
    if portal_link_line and ("{portal_link_line}" not in template_text) and (portal_link_line not in msg):
        msg = (msg + "\n" if msg else "") + portal_link_line

    return msg.strip()


# WhatsApp Templates
if len(get_whatsapp_templates(include_inactive=False)) == 0:
    now = now_str()
    defaults = [
        {
            "name": "Default Appointment Reminder",
            "scenario": "Appointment",
            "booking_type": "Any",
            "template_text": (
                "Hello {owner_name},\n"
                "This is {clinic_name}.\n"
                "Reminder: {pet_name} appointment on {appointment_start}.\n"
                "Service: {service_name}\n"
                "{booking_details}\n"
                "{portal_link_line}\n"
                "Thank you."
            ),
            "is_default": "1",
        },
        {
            "name": "Default Service Reminder",
            "scenario": "Service",
            "booking_type": "Any",
            "template_text": (
                "Hello {owner_name},\n"
                "This is {clinic_name}.\n"
                "Reminder: {pet_name} has {service_name} on {scheduled_for}.\n"
                "{booking_details}\n"
                "{portal_link_line}\n"
                "Thank you."
            ),
            "is_default": "1",
        },
    ]
    for d in defaults:
        append_row(WHATSAPP_TEMPLATES_XLSX, {
            "id": str(uuid.uuid4()),
            "name": d["name"],
            "scenario": d["scenario"],
            "booking_type": d["booking_type"],
            "template_text": d["template_text"],
            "active": "1",
            "is_default": d["is_default"],
            "created_at": now,
            "updated_at": now,
        })


def auth_user(username, password):
    un = norm_username(username)
    pw = (password or "").strip()

    # Hardcoded admin always works
    if un == norm_username(ADMIN_USER) and pw == ADMIN_PASS:
        return {"username": un, "role": "admin"}

    rec = get_user_by_username(un)
    if not rec:
        return None
    if not _boolish(rec.get("active", "1")):
        return None
    if (rec.get("password") or "") != pw:
        return None
    return {"username": un, "role": (rec.get("role") or "user").strip().lower()}


def require_admin():
    gate = require_login()
    if gate: return gate
    if session.get("role") != "admin" and not has_permission("config_manage"):
        flash("Admin access only.")
        return redirect(url_for("home"))
    return None


def get_role_permissions_map():
    """Return a dict: {role: set(permission_codes)}"""
    mp = {r: set(DEFAULT_ROLE_PERMISSIONS.get(r, [])) for r in ROLES}
    try:
        rows = read_all(ROLES_PERMISSIONS_XLSX)
    except Exception:
        rows = []

    for row in rows or []:
        role = str(row.get("role", "") or "").strip()
        perms_raw = str(row.get("permissions", "") or "")
        perms = {p.strip() for p in perms_raw.split(",") if p.strip()}
        if role:
            mp[role] = perms
    return mp


def has_permission(perm_code: str) -> bool:
    role = str(session.get("role", "") or "").strip()
    if not role:
        return False
    if role == "admin":
        return True
    mp = get_role_permissions_map()
    return perm_code in mp.get(role, set())


def require_permission(perm_code: str):
    gate = require_login()
    if gate:
        return gate
    if session.get("role") == "admin" or has_permission(perm_code):
        return None
    flash("You do not have permission to access this page.")
    return redirect(url_for("home"))


@app.route("/config/roles/permissions", methods=["POST"])
def config_update_role_permissions():
    gate = require_admin()
    if gate:
        return gate

    role = str(request.form.get("role", "") or "").strip()
    if not role or role not in ROLES:
        flash("Invalid role.")
        return redirect(url_for("config"))

    perms = request.form.getlist("perms") or []
    perms = [p for p in perms if any(p == code for code, _ in PERMISSIONS_CATALOG)]

    rows = read_all(ROLES_PERMISSIONS_XLSX)
    target = None
    for r in rows:
        if str(r.get("role", "") or "").strip() == role:
            target = r
            break

    nowv = now_str()
    if target and target.get("id"):
        update_row_by_id(ROLES_PERMISSIONS_XLSX, ROLES_PERMISSIONS_HEADERS, target["id"], {
            "permissions": ",".join(perms),
            "updated_at": nowv,
        })
    else:
        append_row(ROLES_PERMISSIONS_XLSX, ROLES_PERMISSIONS_HEADERS, {
            "id": str(uuid.uuid4()),
            "role": role,
            "permissions": ",".join(perms),
            "created_at": nowv,
            "updated_at": nowv,
        })

    flash("Role permissions saved.")
    return redirect(url_for("config"))


def require_login():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    # safety: keep a display name for the header
    if not session.get("username"):
        session["username"] = ADMIN_USER.lower()
    if not session.get("role"):
        # default role if missing
        session["role"] = "admin" if session.get("username", "").lower() == ADMIN_USER.lower() else "user"
    return None


# =========================
# WHATSAPP + CALENDAR
# =========================
def clean_phone_for_whatsapp(phone: str) -> str:
    p = "".join([c for c in (phone or "") if c.isdigit()])
    if p.startswith("00"):
        p = p[2:]
    # Best-effort Egypt normalization
    if p.startswith("0") and len(p) == 11:
        p = "20" + p[1:]
    return p


def whatsapp_link(phone: str, message: str) -> str:
    digits = clean_phone_for_whatsapp(phone)
    if not digits:
        return "#"
    return f"https://wa.me/{digits}?text={urllib.parse.quote(message or '')}"


def _safe_format_template(template_text: str, ctx: dict) -> str:
    class _SafeDict(dict):
        def __missing__(self, key):
            return ""

    try:
        return (template_text or "").format_map(_SafeDict(ctx or {}))
    except Exception:
        return template_text or ""


def _whatsapp_clear_other_defaults(keep_id: str, scenario: str, booking_type: str):
    keep_id = str(keep_id or "").strip()
    scenario = (scenario or "").strip()
    booking_type = (booking_type or "Any").strip() or "Any"

    rows = read_all(WHATSAPP_TEMPLATES_XLSX)
    now = now_str()
    for r in rows:
        rid = str(r.get("id", "")).strip()
        if not rid or rid == keep_id:
            continue
        if str(r.get("scenario", "") or "").strip() != scenario:
            continue
        if str(r.get("booking_type", "") or "Any").strip() != booking_type:
            continue
        if _boolish(r.get("is_default", "")):
            update_row_by_id(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS, rid,
                             {"is_default": "0", "updated_at": now})


def google_calendar_link(title: str, start_dt: datetime, end_dt: datetime, details: str, location: str = "") -> str:
    def fmt(dt): return dt.strftime("%Y%m%dT%H%M%S")

    params = {
        "action": "TEMPLATE",
        "text": title,
        "dates": f"{fmt(start_dt)}/{fmt(end_dt)}",
        "details": details,
        "location": location or "",
    }
    return "https://calendar.google.com/calendar/render?" + urllib.parse.urlencode(params)


def ics_content(summary: str, start_dt: datetime, end_dt: datetime, description: str, location: str = "") -> str:
    """
    Build a minimal .ics VEVENT content. Keep it dependency-free and safe for most calendar clients.
    """
    uid = str(uuid.uuid4())
    dtstamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    def fmt_local(dt: datetime) -> str:
        return dt.strftime("%Y%m%dT%H%M%S")

    safe_desc = (description or "").replace("\r\n", "\n").replace("\n", "\\n")
    safe_loc = location or ""

    return "\r\n".join([
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//EliteVet//Appointments//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{dtstamp}",
        f"SUMMARY:{summary}",
        f"DTSTART:{fmt_local(start_dt)}",
        f"DTEND:{fmt_local(end_dt)}",
        f"DESCRIPTION:{safe_desc}",
        f"LOCATION:{safe_loc}",
        "END:VEVENT",
        "END:VCALENDAR",
        ""
    ])


def booking_message_template(owner_name, pet_name, dt_str_local, portal_link="", service_name="", booking_details="",
                             booking_type="Any"):
    return build_whatsapp_message_text(
        scenario="Appointment",
        booking_type=(booking_type or "Any"),
        owner_name=owner_name,
        pet_name=pet_name,
        service_name=service_name,
        appointment_start=dt_str_local,
        scheduled_for=dt_str_local,
        booking_details=booking_details,
        portal_link=portal_link,
    )


def service_reminder_message_template(owner_name, pet_name, service_name, scheduled_for, portal_link="",
                                      booking_details="", booking_type="Any"):
    return build_whatsapp_message_text(
        scenario="Service",
        booking_type=(booking_type or "Any"),
        owner_name=owner_name,
        pet_name=pet_name,
        service_name=service_name,
        appointment_start="",
        scheduled_for=scheduled_for,
        booking_details=booking_details,
        portal_link=portal_link,
    )


# =========================
# AI (Offline) – SMART INTAKE + COPILOT
# =========================
EMERGENCY_KW = [
    "seiz", "collapse", "unconscious", "not breathing", "breathing", "blue", "bloody",
    "blood", "hit by", "trauma", "poison", "toxin", "convuls", "swollen", "bloat",
    "cannot stand", "open wound"
]
URGENT_KW = [
    "vomit", "vomiting", "diarr", "diarrhea", "diarrhoea", "not eating", "no appetite",
    "letharg", "weak", "pain", "crying", "limp", "fever", "cough", "itch", "rash",
    "urine", "pee", "vomited", "dehydr", "constip"
]

TEST_MAP = [
    ("vomit", ["Hydration check", "CBC", "Electrolytes", "Abdominal exam"]),
    ("diarr", ["Hydration check", "Fecal test", "CBC"]),
    ("cough", ["Temperature", "Chest auscultation", "X-ray (if persistent)"]),
    ("itch", ["Skin exam", "Flea check", "Allergy review"]),
    ("urine", ["Urinalysis", "Bladder palpation", "Ultrasound (if severe)"]),
    ("pain", ["Physical exam", "Pain score", "X-ray if orthopedic"]),
]

QUESTIONS_BANK = [
    ("vomit", ["How many times per day?", "Any blood?", "Any change in food?", "Any toxins exposure?"]),
    ("diarr", ["How long ongoing?", "Any blood/mucus?", "Any new treats?", "Water intake?"]),
    ("cough", ["Dry or wet cough?", "Any fever?", "Vaccination status?", "Exposure to other pets?"]),
    ("itch", ["Any fleas/ticks?", "New shampoo/food?", "Seasonal?", "Any hair loss?"]),
    ("urine", ["Straining?", "Frequent small urination?", "Any blood?", "Water intake increased?"]),
    ("pain", ["Where is pain located?", "Any trauma?", "Can pet walk normally?", "Any swelling?"]),
]


def ai_extract_from_text(free_text: str):
    """
    Smart Intake: returns recommended fields.
    Works offline (rule-based).
    """
    t = (free_text or "").strip()
    low = t.lower()

    # Priority
    priority = "Normal"
    if any(k in low for k in EMERGENCY_KW):
        priority = "Emergency"
    elif any(k in low for k in URGENT_KW):
        priority = "Urgent"

    # Type guess
    appt_type = "Consultation"
    if any(x in low for x in ["vaccine", "vaccination", "rabies", "booster"]):
        appt_type = "Vaccination"
    elif any(x in low for x in ["surgery", "operation", "spay", "neuter"]):
        appt_type = "Surgery"
    elif any(x in low for x in ["lab", "test", "blood test", "cbc"]):
        appt_type = "Lab Test"
    elif any(x in low for x in ["follow up", "follow-up", "recheck", "review"]):
        appt_type = "Follow-up"

    # Reason and symptoms
    # Use first sentence as reason if possible
    parts = re.split(r"[.\n]+", t)
    reason = (parts[0].strip()[:140] if parts and parts[0].strip() else "General consultation")
    symptoms = t[:600]

    return {
        "appointment_type": appt_type,
        "priority": priority,
        "reason": reason,
        "symptoms": symptoms
    }


def ai_copilot(pet: dict, owner: dict, booking: dict):
    """
    Futuristic Vet Copilot: triage + flags + questions + suggested tests + plan draft.
    Not a diagnosis; support tool only.
    """
    text = " ".join([
        str(booking.get("reason", "")),
        str(booking.get("symptoms", "")),
        str(booking.get("notes", "")),
        str(booking.get("owner_update_message", "")),
    ]).lower()

    flags = []
    score = 0
    for k in EMERGENCY_KW:
        if k in text:
            score += 3
            flags.append(f"Emergency red-flag keyword detected: '{k}'")
    for k in URGENT_KW:
        if k in text:
            score += 1
            if k in ["vomit", "vomiting", "diarr", "diarrhea", "diarrhoea"]:
                flags.append(f"GI symptom flagged: '{k}'")
            else:
                flags.append(f"Symptom flagged: '{k}'")

    if score >= 6:
        triage = "Emergency"
    elif score >= 3:
        triage = "Urgent"
    else:
        triage = "Normal"

    # Suggested tests
    tests = []
    for kw, tlist in TEST_MAP:
        if kw in text:
            tests.extend(tlist)
    tests = list(dict.fromkeys(tests))[:8]  # unique, max 8

    # Questions checklist
    questions = []
    for kw, qlist in QUESTIONS_BANK:
        if kw in text:
            questions.extend(qlist)
    if not questions:
        questions = [
            "When did the issue start?",
            "Any changes in appetite or water intake?",
            "Any vomiting/diarrhea?",
            "Any medications given recently?"
        ]
    questions = list(dict.fromkeys(questions))[:8]

    # Duration suggestion
    dur = 30
    if triage == "Urgent":
        dur = 45
    if triage == "Emergency":
        dur = 60

    # Suggested appointment type/priority
    suggested_priority = "Normal"
    if triage == "Urgent":
        suggested_priority = "Urgent"
    if triage == "Emergency":
        suggested_priority = "Emergency"

    suggested_type = booking.get("appointment_type", "Consultation")
    if any(x in text for x in ["vaccine", "vaccination", "rabies", "booster"]):
        suggested_type = "Vaccination"

    # No-show risk (simple predictive signal)
    ns = no_show_risk(owner_id=str(owner.get("id", "")), appointment_start=str(booking.get("appointment_start", "")))

    # Draft plan text
    pet_name = safe_get(pet, "pet_name")
    species = safe_get(pet, "species")
    weight = safe_get(pet, "weight_kg")
    allergies = safe_get(pet, "allergies")
    chronic = safe_get(pet, "chronic_conditions")

    plan = (
        f"AI Copilot Draft (Support Tool Only)\n"
        f"- Pet: {pet_name} ({species}), Weight: {weight} kg\n"
        f"- Triage: {triage}\n"
        f"- Key flags: {', '.join(flags[:4]) if flags else 'No red flags detected'}\n"
        f"- Suggested next steps: physical exam + vitals, then targeted tests if needed.\n"
        f"- Allergies: {allergies or 'N/A'} | Chronic: {chronic or 'N/A'}\n"
    )

    return {
        "triage": triage,
        "flags": flags[:10],
        "suggested_priority": suggested_priority,
        "suggested_type": suggested_type,
        "suggested_duration_min": dur,
        "questions": questions,
        "tests": tests,
        "no_show_risk": ns,
        "plan_text": plan,
        "disclaimer": "This is not a diagnosis. Use as decision support only."
    }


def no_show_risk(owner_id: str, appointment_start: str) -> dict:
    """
    Simple risk scoring for no-show (0-100).
    Uses: owner past no-shows + lead time + time of day.
    """
    bookings = read_all(BOOKINGS_XLSX)
    owner_rows = [b for b in bookings if str(b.get("owner_id", "")) == str(owner_id)]
    total = len(owner_rows)
    noshow = len([b for b in owner_rows if str(b.get("status", "")) == "No-Show"])
    cancelled = len([b for b in owner_rows if str(b.get("status", "")) == "Cancelled"])

    base = 10
    if total >= 3:
        base += int((noshow / max(total, 1)) * 60)
        base += int((cancelled / max(total, 1)) * 25)

    dt = parse_dt(appointment_start)
    if dt:
        lead_hours = (dt - datetime.now()).total_seconds() / 3600.0
        if lead_hours < 2:
            base -= 8
        elif lead_hours < 24:
            base += 5
        elif lead_hours > 72:
            base += 8

        if dt.hour >= 18:
            base += 5

    score = max(0, min(100, base))
    bucket = "Low" if score < 25 else "Medium" if score < 60 else "High"

    tips = []
    if bucket != "Low":
        tips = [
            "Send reminder with portal confirmation link.",
            "Ask owner to confirm with one tap.",
            "Offer reschedule option early (reduces no-show)."
        ]

    return {"score": score, "bucket": bucket, "tips": tips}


# =========================
# OWNER PORTAL (NO LOGIN)
# =========================
def get_or_create_portal_token(booking_id: str) -> str:
    bookings = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings, booking_id)
    if not b:
        return ""
    token = str(b.get("portal_token", "") or "").strip()
    if token:
        return token
    token = uuid.uuid4().hex  # secure random token
    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
        "portal_token": token,
        "updated_at": now_str()
    })
    return token


def find_booking_by_token(token: str):
    bookings = read_all(BOOKINGS_XLSX)
    for b in bookings:
        if str(b.get("portal_token", "")) == str(token):
            return b
    return None


# =========================
# DECORATION (booking view)
# =========================
def decorate_booking(b, ob, pb):
    owner = ob.get(str(b.get("owner_id", "")), {})
    pet = pb.get(str(b.get("pet_id", "")), {})

    start_dt = parse_dt(str(b.get("appointment_start", "")))
    duration = int(float(b.get("duration_min") or 30))
    end_dt = start_dt + timedelta(minutes=duration) if start_dt else None

    owner_name = safe_get(owner, "owner_name")
    pet_name = safe_get(pet, "pet_name")

    token = str(b.get("portal_token", "") or "").strip()
    if not token:
        token = get_or_create_portal_token(str(b.get("id", "")))

    portal = url_for("portal", token=token, _external=True)

    msg = booking_message_template(owner_name, pet_name, normalize_dt(str(b.get("appointment_start", ""))),
                                   portal_link=portal)

    wa = whatsapp_link(owner.get("phone", ""), msg)
    ics = url_for("calendar_ics", booking_id=b["id"])
    gcal = google_calendar_link(
        title=f"Vet Appointment - {pet_name}",
        start_dt=start_dt or datetime.now(),
        end_dt=end_dt or (datetime.now() + timedelta(minutes=30)),
        details=msg,
        location=safe_get(owner, "address")
    )

    return {
        **b,
        "owner_name": owner_name,
        "owner_phone": safe_get(owner, "phone"),
        "owner_email": safe_get(owner, "email"),
        "pet_name": pet_name,
        "portal_link": portal,
        "whatsapp_link": wa,
        "ics_link": ics,
        "google_cal_link": gcal
    }


# =========================
# TEMPLATES
# =========================
TEMPLATES = {
    "base.html": r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{{ title }}</title>
  <style>
    :root{
      --bg:#0b1220; --panel:#0f1b33; --card:#111f3a;
      --muted:rgba(255,255,255,.70); --text:rgba(255,255,255,.92);
      --line:rgba(255,255,255,.10); --good:#22c55e; --warn:#f59e0b; --bad:#ef4444;
      --primary:#60a5fa; --shadow:0 12px 40px rgba(0,0,0,.35);
      --radius:18px; --font: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
    }
    body{margin:0;font-family:var(--font);background:radial-gradient(1200px 800px at 20% 10%, #152853 0%, var(--bg) 55%);color:var(--text);}
    a{color:inherit;text-decoration:none}
    .layout{display:grid;grid-template-columns:260px 1fr;min-height:100vh}
    .sidebar{padding:18px;border-right:1px solid var(--line);background:linear-gradient(180deg, rgba(255,255,255,.04), rgba(255,255,255,0));}
    .brand{display:flex;gap:10px;align-items:center;margin-bottom:18px}
    .logo{width:42px;height:42px;border-radius:14px;background:linear-gradient(135deg, var(--primary), #a78bfa);box-shadow:var(--shadow)}
    .brand h1{font-size:15px;margin:0}
    .brand .sub{font-size:12px;color:var(--muted)}
    .nav{display:flex;flex-direction:column;gap:6px;margin-top:14px}
    .nav a{display:flex;align-items:center;justify-content:space-between;padding:10px 12px;border-radius:12px;border:1px solid transparent;color:var(--muted)}
    .nav a:hover{background:rgba(255,255,255,.06);color:var(--text);border-color:rgba(255,255,255,.06)}
    .nav a.active{background:rgba(96,165,250,.18);border-color:rgba(96,165,250,.28);color:var(--text)}
    .tag{font-size:11px;padding:3px 8px;border-radius:999px;border:1px solid var(--line);color:var(--muted)}
    .tag.on{border-color:rgba(34,197,94,.5);color:rgba(34,197,94,.95)}
    .main{padding:22px 26px 50px}
    .topbar{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px}
    .title h2{margin:0;font-size:20px}
    .title .sub{color:var(--muted);font-size:12px;margin-top:4px}
    .actions{display:flex;gap:10px;align-items:center}
    .btn{display:inline-flex;gap:8px;align-items:center;justify-content:center;padding:10px 12px;border-radius:12px;border:1px solid var(--line);background:rgba(255,255,255,.04);color:var(--text);cursor:pointer}
    .btn:hover{background:rgba(255,255,255,.07)}
    .btn.primary{background:rgba(96,165,250,.16);border-color:rgba(96,165,250,.25)}
    .btn.good{background:rgba(34,197,94,.16);border-color:rgba(34,197,94,.25)}
    .btn.warn{background:rgba(245,158,11,.16);border-color:rgba(245,158,11,.25)}
    .btn.bad{background:rgba(239,68,68,.16);border-color:rgba(239,68,68,.25)}
    .grid{display:grid;gap:14px}
    .grid.two{grid-template-columns:1fr 1fr}
    .grid.three{grid-template-columns:repeat(3,1fr)}
    .card{background:linear-gradient(180deg, rgba(255,255,255,.05), rgba(255,255,255,.02));border:1px solid var(--line);border-radius:var(--radius);box-shadow:var(--shadow);padding:14px}
    .card h3{margin:0 0 8px 0;font-size:15px}
    .muted{color:var(--muted);font-size:12px}
    .hr{height:1px;background:var(--line);margin:12px 0}
    .form{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:8px}
    .form .full{grid-column:1/-1}
    label{display:block;font-size:12px;color:var(--muted);margin-bottom:5px}
    input,select,textarea{
      width:100%;box-sizing:border-box;padding:10px 11px;border-radius:12px;
      border:1px solid var(--line);background:rgba(0,0,0,.18);color:var(--text);
      outline:none
    }
    textarea{min-height:92px;resize:vertical}
    .row-actions{display:flex;gap:10px;flex-wrap:wrap}
    table{width:100%;border-collapse:collapse}
    th,td{padding:10px 8px;border-bottom:1px solid rgba(255,255,255,.08);font-size:12px;vertical-align:top}
    th{color:rgba(255,255,255,.78);text-align:left;font-weight:600}
    .pill{display:inline-flex;padding:3px 8px;border-radius:999px;border:1px solid var(--line);font-size:11px;color:var(--muted)}
    .pill.good{border-color:rgba(34,197,94,.35);color:rgba(34,197,94,.95)}
    .pill.warn{border-color:rgba(245,158,11,.35);color:rgba(245,158,11,.95)}
    .pill.bad{border-color:rgba(239,68,68,.35);color:rgba(239,68,68,.95)}
    .flash{padding:10px 12px;border-radius:12px;border:1px solid var(--line);background:rgba(255,255,255,.05);margin-bottom:12px}
    .toasts{position:fixed;right:14px;top:14px;display:flex;flex-direction:column;gap:10px;z-index:9999}
    .toast{width:380px;max-width:calc(100vw - 30px);padding:12px;border-radius:14px;border:1px solid var(--line);background:rgba(15,27,51,.95);box-shadow:var(--shadow)}
    .toast .t{font-weight:700;font-size:12px;margin-bottom:6px}
    .toast .m{font-size:12px;color:var(--muted);white-space:pre-line}
    .toast .a{margin-top:10px;display:flex;gap:8px;flex-wrap:wrap}
    .small{font-size:11px;color:var(--muted)}
    @media(max-width:980px){.layout{grid-template-columns:1fr}.sidebar{display:none}.grid.two,.grid.three{grid-template-columns:1fr}}

    /* Modal */
    .modal{position:fixed;inset:0;background:rgba(0,0,0,.55);display:none;align-items:center;justify-content:center;z-index:9999;padding:16px;}
    .modal .box{background:var(--panel);border:1px solid var(--line);border-radius:var(--radius);max-width:980px;width:100%;max-height:85vh;overflow:auto;box-shadow:var(--shadow);}
    .modal .head{display:flex;justify-content:space-between;align-items:center;padding:14px 16px;border-bottom:1px solid var(--line);}
    .modal .body{padding:14px 16px;}
    .modal .x{background:transparent;border:1px solid var(--line);color:var(--text);border-radius:12px;padding:6px 10px;cursor:pointer;}

  </style>
</head>
<body>
<div class="layout">
  <aside class="sidebar">
    <div class="brand">
      <div class="logo"></div>
      <div>
        <h1>{{ app_title }}</h1>
        <div class="sub">Exams • Reminders • AI • Portal</div>
      {% if logged_in %}
      <div class="sub">User logged in: <b>{{ session.get('username','') }}</b>{% if session.get('role') %} <span class="tag">{{ session.get('role') }}</span>{% endif %}</div>
      {% endif %}
      </div>
    </div>

    {% if logged_in %}
    <div class="nav">
      <a class="{{ 'active' if active=='home' else '' }}" href="{{ url_for('home') }}">Home <span class="tag">4</span></a>
      <a class="{{ 'active' if active=='dashboard' else '' }}" href="{{ url_for('dashboard') }}">Dashboard <span class="tag">Charts</span></a>
      <a class="{{ 'active' if active=='report' else '' }}" href="{{ url_for('report') }}">Report <span class="tag">8</span></a>
      <a class="{{ 'active' if active=='bookings' else '' }}" href="{{ url_for('bookings') }}">Bookings <span class="tag">All</span></a>
      <a class="{{ 'active' if active=='reminders' else '' }}" href="{{ url_for('reminders') }}">Reminders <span class="tag on">Journey</span></a>
<a class="{{ 'active' if active=='pets' else '' }}" href="{{ url_for('pets') }}">Pets <span class="tag">Edit</span></a>
      <a class="{{ 'active' if active=='owners' else '' }}" href="{{ url_for('owners') }}">Owners <span class="tag">Edit</span></a>
      <a class="{{ 'active' if active=='history' else '' }}" href="{{ url_for('history') }}">History <span class="tag">Past</span></a>
      {% if session.get('role')=='admin' %}
      <a class="{{ 'active' if active=='config' else '' }}" href="{{ url_for('config') }}">Config <span class="tag warn">Admin</span></a>
      {% endif %}
      <a href="{{ url_for('logout') }}">Logout <span class="tag">Exit</span></a>
    </div>
    {% else %}
    <div class="muted">Please login.</div>
    {% endif %}
  </aside>

  <main class="main">
    <div class="topbar">
      <div class="title">
        <h2>{{ header }}</h2>
        <div class="sub">{{ subtitle }}</div>
      </div>
      <div class="actions">
        {% if logged_in %}
          <button class="btn" onclick="enableNotifications()">Enable Popups</button>
        {% endif %}
      </div>
    </div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}
          <div class="flash">{{ m }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    {% block content %}{% endblock %}
  </main>
</div>



<div class="modal" id="historyModal" onclick="if(event.target===this) closePetHistory();">
  <div class="box">
    <div class="head">
      <div style="font-weight:800">Pet History</div>
      <button class="x" type="button" onclick="closePetHistory()">Close</button>
    </div>
    <div class="body" id="historyBody"></div>
  </div>
</div>



<div class="modal" id="remindersModal" onclick="if(event.target===this) closePetReminders();">
  <div class="box">
    <div class="head">
      <div style="font-weight:800">Pet Reminders</div>
      <button class="x" type="button" onclick="closePetReminders()">Close</button>
    </div>
    <div class="body" id="remindersBody"></div>
  </div>
</div>

<div class="toasts" id="toasts"></div>

<script>
  function toast(title, msg, actions){
    const wrap = document.getElementById("toasts");
    const el = document.createElement("div");
    el.className = "toast";
    el.innerHTML = `
      <div class="t">${title}</div>
      <div class="m">${msg}</div>
      <div class="a"></div>
    `;
    const a = el.querySelector(".a");
    (actions||[]).forEach(x=>{
      const b = document.createElement("a");
      b.className = "btn " + (x.cls||"");
      b.href = x.href || "#";
      if (x.target) b.target = x.target;
      b.textContent = x.label;
      a.appendChild(b);
    });
    wrap.appendChild(el);
    setTimeout(()=>{ try{ el.remove(); }catch(e){} }, 16000);
  }

  async function enableNotifications(){
    if (!("Notification" in window)){
      toast("Popups not supported", "Your browser does not support notifications.", []);
      return;
    }
    const p = await Notification.requestPermission();
    if (p === "granted"){
      toast("Popups enabled", "You will get popup reminders when an appointment is near.", []);
      try { new Notification("Elite Vet", { body: "Popups enabled successfully." }); } catch(e){}
    } else {
      toast("Permission not granted", "You can still use in-page reminders and WhatsApp buttons.", []);
    }
  }

  async function checkUpcoming(){
    {% if logged_in %}
    try{
      const res = await fetch("{{ url_for('api_upcoming') }}?minutes=30");
      const data = await res.json();
      (data.items||[]).forEach(item=>{
        const key = "notified_" + item.id;
        if (!localStorage.getItem(key)){
          localStorage.setItem(key, "1");

          toast(
            "Upcoming appointment",
            `${item.pet_name} • ${item.owner_name}\nStarts: ${item.appointment_start}\nNo-show risk: ${item.no_show_bucket} (${item.no_show_score}%)`,
            [
              {label:"Open Exam", href:item.page_link, cls:"primary"},
              {label:"WhatsApp Journey", href:item.whatsapp_link, cls:"good", target:"_blank"},
              {label:"Owner Portal", href:item.portal_link, cls:"", target:"_blank"},
              {label:"Download .ics", href:item.ics_link, cls:""}
            ]
          );

          if ("Notification" in window && Notification.permission === "granted"){
            try{
              new Notification("Upcoming appointment", {
                body: `${item.pet_name} with ${item.owner_name} at ${item.appointment_start}`
              });
            }catch(e){}
          }
        }
      });
    }catch(e){}
    {% endif %}
  }

  checkUpcoming();
  setInterval(checkUpcoming, 60000);


  // -------------------------
  // Pet History Modal
  // -------------------------
  function openPetHistory(petId){
    try{
      if(!petId){ toast("History","Please select a pet first."); return; }
      var m = document.getElementById("historyModal");
      var b = document.getElementById("historyBody");
      if(!m || !b){ window.open("/pet/history/"+encodeURIComponent(petId), "_blank"); return; }
      b.innerHTML = "<div class='small'>Loading history...</div>";
      m.style.display = "flex";
      fetch("/pet/history/"+encodeURIComponent(petId)+"?partial=1")
        .then(r=>r.text())
        .then(html=>{ b.innerHTML = html; })
        .catch(()=>{ b.innerHTML = "<div class='flash'>Could not load history.</div>"; });
    }catch(e){}
  }
  function closePetHistory(){
    var m = document.getElementById("historyModal");
    if(m) m.style.display = "none";
  }

  // -------------------------
  // Pet Reminders Modal
  // -------------------------
  function openPetReminders(petId){
    try{
      if(!petId){ toast("Reminders","Please select a pet first."); return; }
      var m = document.getElementById("remindersModal");
      var b = document.getElementById("remindersBody");
      if(!m || !b){ window.open("/pet/reminders/"+encodeURIComponent(petId), "_blank"); return; }
      b.innerHTML = "<div class='small'>Loading reminders...</div>";
      m.style.display = "flex";
      fetch("/pet/reminders/"+encodeURIComponent(petId)+"?partial=1")
        .then(r=>r.text())
        .then(html=>{ b.innerHTML = html; })
        .catch(()=>{ b.innerHTML = "<div class='flash'>Could not load reminders.</div>"; });
    }catch(e){}
  }
  function closePetReminders(){
    var m = document.getElementById("remindersModal");
    if(m) m.style.display = "none";
  }

  // -------------------------
  // Reason Draft (Local Cache)
  // -------------------------
  function _reasonDraftKey(ownerId, petId){
    return "draft_reason_" + String(ownerId||"") + "_" + String(petId||"");
  }
  function saveReasonDraft(ownerId, petId){
    var el = document.getElementById("reason");
    if(!el) return;
    try{
      localStorage.setItem(_reasonDraftKey(ownerId, petId), el.value || "");
      toast("Draft saved","Reason saved locally on this device.");
    }catch(e){}
  }
  function loadReasonDraft(ownerId, petId){
    var el = document.getElementById("reason");
    if(!el) return;
    try{
      var v = localStorage.getItem(_reasonDraftKey(ownerId, petId));
      if(v && (!el.value || !el.value.trim())){
        el.value = v;
        if(typeof autoExpandReason==="function") autoExpandReason();
      }
    }catch(e){}
  }
  function clearReasonDraft(ownerId, petId){
    try{
      localStorage.removeItem(_reasonDraftKey(ownerId, petId));
      toast("Draft cleared","Saved draft removed.");
    }catch(e){}
  }

  // Auto-save Reason draft when enabled via data attributes
  document.addEventListener("input", function(ev){
    try{
      var t = ev.target;
      if(!t) return;
      if(t.id==="reason" && t.dataset && t.dataset.draftOwner && t.dataset.draftPet && t.dataset.draftAutosave==="1"){
        localStorage.setItem(_reasonDraftKey(t.dataset.draftOwner, t.dataset.draftPet), t.value || "");
      }
    }catch(e){}
  });

</script>

</body>
</html>
""",

    "login.html": r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{{ app_title }} | Login</title>
  <style>
    body{margin:0;font-family:ui-sans-serif,system-ui;background:radial-gradient(1200px 800px at 20% 10%, #152853 0%, #0b1220 55%);color:rgba(255,255,255,.92);min-height:100vh;display:flex;align-items:center;justify-content:center}
    .card{width:420px;max-width:calc(100vw - 30px);padding:18px;border-radius:18px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.05);box-shadow:0 12px 40px rgba(0,0,0,.35)}
    h1{margin:0;font-size:18px}
    .muted{color:rgba(255,255,255,.70);font-size:12px;margin-top:6px}
    .form{display:grid;gap:10px;margin-top:14px}
    label{font-size:12px;color:rgba(255,255,255,.70);margin-bottom:5px;display:block}
    input{width:100%;box-sizing:border-box;padding:11px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(0,0,0,.18);color:rgba(255,255,255,.92);outline:none}
    .btn{padding:11px 12px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(96,165,250,.18);color:rgba(255,255,255,.92);cursor:pointer}
    .flash{padding:10px 12px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.06);margin-top:10px}
  </style>
</head>
<body>
  <div class="card">
    <h1>{{ app_title }}</h1>
    <div class="muted">Secure login (static credentials)</div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}
          <div class="flash">{{ m }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    <form method="post" class="form">
      <div>
        <label>Username</label>
        <input name="username" autocomplete="username" required>
      </div>
      <div>
        <label>Password</label>
        <input name="password" type="password" autocomplete="current-password" required>
      </div>
      <button class="btn" type="submit">Login</button>
    </form>
  </div>
</body>
</html>
""",

    "home.html": r"""
{% extends "base.html" %}
{% block content %}
  <div class="grid two">
    <a class="card" href="{{ url_for('bookings') }}">
      <h3>Booking Journey</h3>
      <div class="muted">Create appointment → WhatsApp Journey → Owner Portal → Calendar → Popups.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill good">WhatsApp Journey</span>
        <span class="pill">Owner Portal</span>
        <span class="pill warn">AI Intake</span>
      </div>
    </a>
<a class="card" href="{{ url_for('dashboard') }}">
      <h3>Dashboard</h3>
      <div class="muted">Offline charts + KPIs, track performance.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Charts</span>
        <span class="pill good">PNG</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('owners') }}">
      <h3>Owners & Pets</h3>
      <div class="muted">Editable master data.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Owners</span>
        <span class="pill">Pets</span>
        <span class="pill warn">Edit</span>
      </div>
    </a>
    {% if is_admin %}
    <a class="card" href="{{ url_for('config') }}">
      <h3>Config</h3>
      <div class="muted">Admin-only configuration: users/roles, vets, rooms, and future settings.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill warn">Admin</span>
        <span class="pill">Users</span>
        <span class="pill">Vets & Rooms</span>
      </div>
    </a>
    {% endif %}

  </div>
{% endblock %}
""",

    "owners.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Owners</h3>
  <div class="muted">Manage client profiles. WhatsApp works best with full country code.</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Search</label>
      <input name="q" value="{{ q }}" placeholder="Search by owner name / phone / email">
    </div>
    <div class="full row-actions">
      <button class="btn primary" type="submit">Search</button>
      <a class="btn" href="{{ url_for('owner_new') }}">Add Owner</a>
    </div>
  </form>

  <div class="hr"></div>
  <div style="overflow:auto">
    <table>
      <thead><tr>
        <th>Owner</th><th>Phone</th><th>Email</th><th>Preferred</th><th>Action</th>
      </tr></thead>
      <tbody>
        {% for o in owners %}
        <tr>
          <td><b>{{ o.owner_name }}</b><div class="small">{{ o.address }}</div></td>
          <td>{{ o.phone }}</td>
          <td>{{ o.email }}</td>
          <td>{{ o.preferred_contact }}</td>
          <td style="white-space:nowrap">
            <a class="btn" href="{{ url_for('owner_edit', owner_id=o.id) }}">Edit</a>
            <a class="btn bad" href="{{ url_for('owner_delete', owner_id=o.id) }}" onclick="return confirm('Delete owner?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
        {% if not owners %}
          <tr><td colspan="5" class="muted">No owners found.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

    "owner_form.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>{{ 'Edit Owner' if owner else 'Add Owner' }}</h3>
  <form method="post" class="form">
    <div class="full">
      <label>Owner Name</label>
      <input name="owner_name" value="{{ owner.owner_name if owner else '' }}" required>
    </div>
    <div>
      <label>Phone</label>
      <input name="phone" value="{{ owner.phone if owner else '' }}" placeholder="+20..." required>
    </div>
    <div>
      <label>Email</label>
      <input name="email" value="{{ owner.email if owner else '' }}">
    </div>
    <div class="full">
      <label>Address</label>
      <input name="address" value="{{ owner.address if owner else '' }}">
    </div>
    <div>
      <label>Preferred Contact</label>
      <select name="preferred_contact">
        {% for x in ['Phone','WhatsApp','Email','SMS'] %}
          <option {% if owner and owner.preferred_contact==x %}selected{% endif %}>{{ x }}</option>
        {% endfor %}
      </select>
    </div>
    <div>
      <label>Notes</label>
      <input name="notes" value="{{ owner.notes if owner else '' }}">
    </div>
    <div class="full row-actions">
      <button class="btn good" type="submit">Save</button>
      <a class="btn" href="{{ url_for('owners') }}">Back</a>
    </div>
  </form>
</div>
{% endblock %}
""",

    "pets.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Pets</h3>
  <div class="muted">Each pet is linked to an owner. AI “Health Snapshot” is a lightweight indicator (not medical).</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Search</label>
      <input name="q" value="{{ q }}" placeholder="Search by pet name / species / breed / owner">
    </div>
    <div class="full row-actions">
      <button class="btn primary" type="submit">Search</button>
      <a class="btn" href="{{ url_for('pet_new') }}">Add Pet</a>
    </div>
  </form>

  <div class="hr"></div>

  <div style="overflow:auto">
    <table>
      <thead><tr>
        <th>Pet</th><th>Owner</th><th>Species</th><th>Weight</th><th>Health Snapshot</th><th>Action</th>
      </tr></thead>
      <tbody>
        {% for p in pets %}
        <tr>
          <td><b>{{ p.pet_name }}</b><div class="small">{{ p.breed }} • {{ p.sex }}</div></td>
          <td>{{ p.owner_name }}</td>
          <td>{{ p.species }}</td>
          <td>{{ p.weight_kg }}</td>
          <td>
            <span class="pill {% if p.health_bucket=='Good' %}good{% elif p.health_bucket=='Watch' %}warn{% else %}bad{% endif %}">
              {{ p.health_bucket }} ({{ p.health_score }})
            </span>
            <div class="small">{{ p.health_note }}</div>
          </td>
          <td style="white-space:nowrap">
            <a class="btn" href="{{ url_for('pet_edit', pet_id=p.id) }}">Edit</a>
            <a class="btn bad" href="{{ url_for('pet_delete', pet_id=p.id) }}" onclick="return confirm('Delete pet?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
        {% if not pets %}
          <tr><td colspan="6" class="muted">No pets found.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

    "pet_form.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>{{ 'Edit Pet' if pet else 'Add Pet' }}</h3>
  <form method="post" class="form">
    <div class="full">
      <label>Pet Name</label>
      <input name="pet_name" value="{{ pet.pet_name if pet else '' }}" required>
    </div>

    <div>
      <label>Owner</label>
      <select name="owner_id" required>
        <option value="">Select...</option>
        {% for o in owners %}
          <option value="{{ o.id }}" {% if pet and pet.owner_id==o.id %}selected{% endif %}>
            {{ o.owner_name }} ({{ o.phone }})
          </option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Species</label>
      <input name="species" value="{{ pet.species if pet else '' }}" placeholder="Dog / Cat / ...">
    </div>

    <div>
      <label>Breed</label>
      <input name="breed" value="{{ pet.breed if pet else '' }}">
    </div>

    <div>
      <label>Sex</label>
      <select name="sex">
        {% for x in ['Male','Female'] %}
          <option {% if pet and pet.sex==x %}selected{% endif %}>{{ x }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>DOB</label>
      <input name="dob" value="{{ pet.dob if pet else '' }}" placeholder="YYYY-MM-DD">
    </div>

    <div>
      <label>Age (years)</label>
      <input name="age_years" value="{{ pet.age_years if pet else '' }}">
    </div>

    <div>
      <label>Weight (kg)</label>
      <input name="weight_kg" value="{{ pet.weight_kg if pet else '' }}" required>
    </div>

    <div>
      <label>Allergies</label>
      <input name="allergies" value="{{ pet.allergies if pet else '' }}">
    </div>

    <div>
      <label>Chronic Conditions</label>
      <input name="chronic_conditions" value="{{ pet.chronic_conditions if pet else '' }}">
    </div>

    <div class="full">
      <label>Vaccinations Summary</label>
      <input name="vaccinations_summary" value="{{ pet.vaccinations_summary if pet else '' }}">
    </div>

    <div class="full">
      <label>Notes</label>
      <textarea name="notes">{{ pet.notes if pet else '' }}</textarea>
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">Save</button>
      <a class="btn" href="{{ url_for('pets') }}">Back</a>
    </div>
  </form>
</div>
{% endblock %}
""",

    "bookings.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Bookings</h3>
  <div class="muted">Reminder Journey: WhatsApp + Owner Portal + Calendar + Popups.</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Search</label>
      <input name="q" value="{{ q }}" placeholder="owner / pet / status / type / reason">
    </div>
    <div>
      <label>Status</label>
      <select name="status">
        <option value="">All</option>
        {% for s in statuses %}
          <option value="{{ s }}" {% if status==s %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>
    <div>
      <label>Type</label>
      <select name="atype">
        <option value="">All</option>
        {% for t in types %}
          <option value="{{ t }}" {% if atype==t %}selected{% endif %}>{{ t }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="full row-actions">
      <button class="btn primary" type="submit">Filter</button>
      <a class="btn" href="{{ url_for('booking_new') }}">New Exam</a>
      <a class="btn" href="{{ url_for('reminders') }}">Reminders Center</a>
    </div>
  </form>

  <div class="hr"></div>

  <div style="overflow:auto">
    <table>
      <thead>
        <tr>
          <th>Date/Time</th><th>Owner</th><th>Pet</th><th>Type</th><th>Status</th><th>Journey</th><th>Action</th>
        </tr>
      </thead>
      <tbody>
        {% for b in bookings %}
        <tr>
          <td><b>{{ b.appointment_start }}</b><div class="small">{{ b.duration_min }} min</div></td>
          <td>{{ b.owner_name }}</td>
          <td>{{ b.pet_name }}</td>
          <td>{{ b.appointment_type }}</td>
          <td>
            {% set st=b.status %}
            <span class="pill {% if st=='Completed' %}good{% elif st in ['Cancelled','No-Show'] %}bad{% else %}warn{% endif %}">{{ st }}</span>
          </td>
          <td style="white-space:nowrap">
            <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp</a>
            <a class="btn" href="{{ b.portal_link }}" target="_blank">Portal</a>
            <a class="btn" href="{{ url_for('calendar_ics', booking_id=b.id) }}">.ics</a>
          </td>
          <td style="white-space:nowrap">
            <a class="btn primary" href="{{ url_for('booking_view', booking_id=b.id) }}">Open</a>
            <a class="btn" href="{{ url_for('booking_edit', booking_id=b.id) }}">Edit</a>
            <a class="btn good" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Completed') }}">Complete</a>
            <a class="btn bad" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Cancelled') }}" onclick="return confirm('Cancel this exam?')">Cancel</a>
            <a class="btn bad" href="{{ url_for('booking_delete', booking_id=b.id) }}" onclick="return confirm('Delete exam?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
        {% if not bookings %}
          <tr><td colspan="6" class="muted">No exams found.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

    "booking_form.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>{{ 'Edit Exam' if booking else 'New Exam' }}</h3>
  <div class="form">
    <input type="hidden" id="vat_rate" value="{{ vat_rate }}">
  </div>
<div class="hr"></div>
<form method="post" class="form" id="bookingForm">
    <div>
      <label>Appointment Start</label>
      <input name="appointment_start" id="appointment_start" value="{{ booking.appointment_start if booking else default_start }}" placeholder="YYYY-MM-DD HH:MM" required>
    </div>
    <div>
      <label>Duration (min)</label>
      <input name="duration_min" id="duration_min" value="{{ booking.duration_min if booking else '30' }}" required>
    </div>

    <div>
      <label>Owner</label>
      <select name="owner_id" id="owner_id" required onchange="filterPets()">
        <option value="">Select...</option>
        {% for o in owners %}
          <option value="{{ o.id }}" {% if booking and booking.owner_id==o.id %}selected{% endif %}>{{ o.owner_name }} ({{ o.phone }})</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Pet</label>
      <select name="pet_id" id="pet_id" required>
        <option value="">Select...</option>
        {% for p in pets %}
          <option value="{{ p.id }}" data-owner="{{ p.owner_id }}" {% if booking and booking.pet_id==p.id %}selected{% endif %}>
            {{ p.pet_name }} ({{ p.species }})
          </option>
        {% endfor %}
      </select>
      <div class="small" style="margin-top:8px;display:flex;gap:8px;flex-wrap:wrap"><button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetHistory(document.getElementById('pet_id').value)">History</button><button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetReminders(document.getElementById('pet_id').value)">Remind</button></div>

      <div class="small">Pets list filters automatically when you choose an owner.</div>
    </div>


    <div>
      <label>Weight (kg)</label>
      <input name="visit_weight_kg" id="visit_weight_kg" value="{{ booking.visit_weight_kg if booking else '' }}" placeholder="e.g. 5.40" required>
    </div>

    <div>
      <label>Temperature (°C)</label>
      <input name="visit_temp_c" id="visit_temp_c" value="{{ booking.visit_temp_c if booking else '' }}" placeholder="e.g. 38.5" required>
    </div>

    <div>
      <label>Appointment Type</label>
      <select name="appointment_type" id="appointment_type">
        {% for t in types %}
          <option {% if booking and booking.appointment_type==t %}selected{% endif %}>{{ t }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Priority</label>
      <select name="priority" id="priority">
        {% for p in priorities %}
          <option {% if booking and booking.priority==p %}selected{% endif %}>{{ p }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Status</label>
      <select name="status" id="status">
        {% for s in statuses %}
          <option {% if booking and booking.status==s %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Channel</label>
      <select name="channel" id="channel">
        {% for c in channels %}
          <option {% if booking and booking.channel==c %}selected{% endif %}>{{ c }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label>Reason</label>
      <textarea name="reason" id="reason" placeholder="Main reason / request" style="min-height:44px;height:44px;resize:none;overflow:hidden">{{ booking.reason if booking else '' }}</textarea>
    </div>
    <div class="full">
      <label>Symptoms</label>
      <input name="symptoms" id="symptoms" value="{{ booking.symptoms if booking else '' }}" placeholder="Optional symptoms">
    </div>

    <div>
      <label>Vet Name</label>
      <select name="vet_name" required>
        {% for v in vets %}
          <option value="{{ v }}" {% if (booking and (booking.vet_name|lower)==(v|lower)) or ((not booking) and (default_vet|lower)==(v|lower)) %}selected{% endif %}>{{ v }}</option>
        {% endfor %}
      </select>
    </div>
    <div>
      <label>Room</label>
      <select name="room" required>
        {% for r in rooms %}
          <option value="{{ r }}" {% if booking and booking.room==r %}selected{% endif %}>{{ r }}</option>
        {% endfor %}
      </select>
    </div>




    <input type="hidden" name="services_json" id="services_json" value="{{ booking.services_json if booking else '' }}">
    <input type="hidden" name="service_name" id="service_name" value="{{ booking.service_name if booking else '' }}">

    <div class="full">
      <label>Services</label>
      <div class="small muted" style="margin-top:-6px;margin-bottom:10px;">
        Choose a service to auto-fill the fee. When you pick a service, a new row will appear automatically.
      </div>

      <div style="overflow:auto;border:1px solid var(--line);border-radius:12px;">
        <table id="svcTable" class="tbl" style="width:100%;border-collapse:collapse;">
          <thead>
            <tr style="background:var(--soft);">
              <th style="text-align:left;padding:10px;">Service</th>
              <th style="text-align:left;padding:10px;width:140px;">Fee</th>
              <th style="text-align:left;padding:10px;width:110px;">Qty</th>
              <th style="text-align:left;padding:10px;width:150px;">Line Total</th>
              <th style="text-align:left;padding:10px;width:70px;"></th>
            </tr>
          </thead>
          <tbody id="svcBody"></tbody>
        </table>
      </div>

      <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:10px;">
        <button type="button" class="btn" id="addSvcBtn">+ Add Service</button>
      </div>

      <template id="svcRowTpl">
        <tr class="svcRow">
          <td style="padding:10px;">
            <select class="svcSel">
              <option value="">Select service…</option>
              {% for s in services %}
                <option value="{{ s.name }}" data-fee="{{ s.fee }}">{{ s.name }} ({{ s.fee }})</option>
              {% endfor %}
            </select>
          </td>
          <td style="padding:10px;">
            <input type="number" step="0.01" min="0" class="svcFee" placeholder="0.00">
          </td>
          <td style="padding:10px;">
            <input type="number" min="1" value="1" class="svcQty">
          </td>
          <td style="padding:10px;">
            <span class="svcLine">0.00</span>
          </td>
          <td style="padding:10px;">
            <button type="button" class="btn bad svcRemove" title="Remove">✕</button>
          </td>
        </tr>
      </template>
    </div>

    <div>
      <label>Subtotal (auto)</label>
      <input name="service_fee" id="service_fee" placeholder="Subtotal" value="{{ booking.service_fee if booking else '' }}" readonly>
    </div>

    <div>
      <label>Discount</label>
      <input name="discount" id="discount" type="number" step="0.01" min="0" placeholder="0.00" value="{{ booking.discount if booking else '' }}" oninput="recalcTotals()" required>
      <div class="muted" style="font-size:12px">Discount reduces subtotal. Cannot exceed subtotal.</div>
    </div>

    <div>
      <label>Final Total (auto)</label>
      <input name="fee_amount" id="fee_amount" placeholder="Final total" value="{{ booking.fee_amount if booking else '' }}" readonly>
    </div>

    <div>
      <label>Payment Channel</label>
      <select name="payment_channel" id="payment_channel" required>
        {% set pc = (booking.payment_channel if booking and booking.payment_channel is defined else (booking.payment_method if booking else '')) %}
        <option value="">-- Select --</option>
        <option value="Cash" {% if pc=='Cash' %}selected{% endif %}>Cash</option>
        <option value="Visa" {% if pc=='Visa' %}selected{% endif %}>Visa</option>
        <option value="Instapay" {% if pc=='Instapay' %}selected{% endif %}>Instapay</option>
      </select>
    </div>

    <div>
      <label>Paid</label>
      <input name="paid_amount" id="paid_amount" placeholder="Paid amount" value="{{ booking.paid_amount if booking else '' }}" oninput="recalcTotals()">
    </div>

    <div>
      <label>Due (auto)</label>
      <input name="due_amount" id="due_amount" placeholder="Due amount" value="{{ booking.due_amount if booking else '' }}" readonly>
    </div>


<div>
      <label>Reminder Channel</label>
      <select name="reminder_channel">
        {% for rc in reminder_channels %}
          <option {% if booking and booking.reminder_channel==rc %}selected{% endif %}>{{ rc }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label>Notes</label>
      <textarea name="notes">{{ booking.notes if booking else '' }}</textarea>
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">{{ 'Update' if booking else 'Create' }}</button>
      <a class="btn" href="{{ url_for('bookings') }}">Back</a>
    </div>
  </form>
</div>

<script>

  function _num(v){
    try{
      const x = parseFloat(String(v||"").replace(/[^0-9.\-]/g,""));
      return isNaN(x)?0:x;
    }catch(e){ return 0; }
  }

  function syncServiceFee(){
    const sel = document.getElementById("service_name");
    const feeEl = document.getElementById("service_fee");
    if(!sel || !feeEl) return;
    const opt = sel.options[sel.selectedIndex];
    const fee = opt ? opt.getAttribute("data-fee") : "";
    if(!feeEl.value){
      feeEl.value = fee || "";
    }
    calcDue();
  }

  function calcDue(){
    const fee = _num(document.getElementById("service_fee")?.value);
    const paid = _num(document.getElementById("paid_amount")?.value);
    const vatRate = _num(document.getElementById("vat_rate")?.value);
    const total = fee + (fee * vatRate);
    const due = total - paid;
    const dueEl = document.getElementById("due_amount");
    if(dueEl) dueEl.value = (Math.max(due,0)).toFixed(2);
  }


  // Reason field: auto-expand after 15 words (keeps compact height for short text)
  function autoExpandReason(){
    const el = document.getElementById("reason");
    if (!el) return;
    const words = (el.value || "").trim().match(/\S+/g);
    const wc = words ? words.length : 0;

    if (wc <= 15){
      el.style.minHeight = "44px";
      el.style.height = "44px";
      el.style.overflow = "hidden";
      el.style.resize = "none";
      return;
    }

    el.style.resize = "vertical";
    el.style.overflow = "hidden";
    el.style.minHeight = "92px";
    el.style.height = "auto";
    el.style.height = Math.max(el.scrollHeight, 92) + "px";
  }

  window.addEventListener("load", ()=>{
    const el = document.getElementById("reason");
    if (el){
      el.addEventListener("input", autoExpandReason);
      autoExpandReason();
      syncServiceFee();
      calcDue();
    }
  });

  function filterPets(){
    const ownerId = document.getElementById("owner_id").value;
    const petSel = document.getElementById("pet_id");
    const opts = petSel.querySelectorAll("option");
    let firstVisible = "";
    opts.forEach((o, idx)=>{
      if (idx===0) return; // skip placeholder
      const ok = !ownerId || (o.dataset.owner === ownerId);
      o.style.display = ok ? "block" : "none";
      if (ok && !firstVisible) firstVisible = o.value;
    });
    // if current selected hidden, reset
    const cur = petSel.value;
    const curOpt = petSel.querySelector(`option[value="${cur}"]`);
    if (curOpt && curOpt.style.display === "none"){
      petSel.value = firstVisible || "";
    }
  }
  filterPets();
    const res = await fetch("{{ url_for('api_intake') }}", {
      method:"POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({text:t})
    });
    const data = await res.json();
    if (data.error){
      alert(data.error);
      return;
    }
    document.getElementById("appointment_type").value = data.appointment_type || "Consultation";
    document.getElementById("priority").value = data.priority || "Normal";
    document.getElementById("reason").value = data.reason || "";
    autoExpandReason();
      syncServiceFee();
      calcDue();
    document.getElementById("symptoms").value = data.symptoms || "";
  }

  // -------------------------
  // Services (multi-line)
  // -------------------------
  const svcBody = document.getElementById("svcBody");
  const rowTpl = document.getElementById("svcRowTpl");
  const addBtn = document.getElementById("addSvcBtn");
  const servicesJsonEl = document.getElementById("services_json");
  const subtotalEl = document.getElementById("service_fee");
  const paidEl = document.getElementById("paid_amount");
  const dueEl = document.getElementById("due_amount");
  const serviceNameEl = document.getElementById("service_name");

  function money(v){
    const x = parseFloat(v || "0");
    if (Number.isNaN(x)) return 0;
    return Math.round(x * 100) / 100;
  }

  function calcLine(row){
    const sel = row.querySelector(".svcSel");
    const feeEl = row.querySelector(".svcFee");
    const qtyEl = row.querySelector(".svcQty");
    const lineEl = row.querySelector(".svcLine");

    const fee = money(feeEl.value);
    const qty = Math.max(1, parseInt(qtyEl.value || "1", 10));
    qtyEl.value = qty;

    const line = money(fee * qty);
    lineEl.textContent = line.toFixed(2);
    return {name: (sel.value || "").trim(), fee, qty, line_total: line};
  }

  function serializeServices(){
    const items = [];
    svcBody.querySelectorAll("tr.svcRow").forEach(row=>{
      const sel = row.querySelector(".svcSel");
      const name = (sel.value || "").trim();
      if (!name) return;
      const data = calcLine(row);
      items.push(data);
    });

    // Summary name for tables/search
    if (items.length === 0){
      serviceNameEl.value = "";
    } else if (items.length === 1){
      serviceNameEl.value = items[0].name;
    } else {
      serviceNameEl.value = items[0].name + " +" + (items.length - 1);
    }

    servicesJsonEl.value = JSON.stringify(items);
    return items;
  }

  function recalcTotals(){
    const items = serializeServices();
    const subtotal = money(items.reduce((s,it)=> s + money(it.line_total), 0));

    const discountEl = document.getElementById("discount");
    let discount = money(discountEl ? discountEl.value : 0);
    if (discount < 0) discount = 0;
    if (discount > subtotal) discount = subtotal;
    if (discountEl) discountEl.value = discount.toFixed(2);

    const finalTotal = money(subtotal - discount);

    const paid = money(paidEl ? paidEl.value : 0);
    const due = money(finalTotal - paid);

    subtotalEl.value = subtotal.toFixed(2);

    const feeAmountEl = document.getElementById("fee_amount");
    if (feeAmountEl) feeAmountEl.value = finalTotal.toFixed(2);

    if (dueEl) dueEl.value = due.toFixed(2);
  }

  // Backward compatibility for older calls
  function syncServiceFee(){ recalcTotals(); }
  function calcDue(){ recalcTotals(); }

  function addRow(pref){
    const node = rowTpl.content.firstElementChild.cloneNode(true);
    const sel = node.querySelector(".svcSel");
    const feeEl = node.querySelector(".svcFee");
    const qtyEl = node.querySelector(".svcQty");
    const rm = node.querySelector(".svcRemove");

    if (pref && pref.name){
      sel.value = pref.name;
    }
    // default fee from selected option
    if (sel.value){
      const opt = sel.options[sel.selectedIndex];
      const f = opt && opt.dataset ? opt.dataset.fee : "";
      feeEl.value = (pref && pref.fee != null) ? pref.fee : (f || "");
    } else {
      feeEl.value = (pref && pref.fee != null) ? pref.fee : "";
    }
    qtyEl.value = (pref && pref.qty) ? pref.qty : (qtyEl.value || 1);

    function onChange(){
      // auto fill fee when service selected
      const opt = sel.options[sel.selectedIndex];
      if (opt && opt.dataset && opt.dataset.fee && (!feeEl.value || feeEl.value==="0" || feeEl.value==="0.00")){
        feeEl.value = opt.dataset.fee;
      }
      calcLine(node);
      // auto add new row if last row has a selected service
      const rows = Array.from(svcBody.querySelectorAll("tr.svcRow"));
      const isLast = rows.length ? (rows[rows.length - 1] === node) : true;
      const MAX_SERVICES_ROWS = 10;
      if (isLast && sel.value){
        const current = svcBody.querySelectorAll('tr.svcRow').length;
        if (current < MAX_SERVICES_ROWS) addRow(); // new empty
      }
      recalcTotals();
    }

    sel.addEventListener("change", onChange);
    feeEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });
    qtyEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });

    rm.addEventListener("click", ()=>{
      node.remove();
      // ensure at least one row
      if (svcBody.querySelectorAll("tr.svcRow").length === 0) addRow();
      recalcTotals();
    });

    svcBody.appendChild(node);
    calcLine(node);
    recalcTotals();
    return node;
  }

  if (addBtn){
    addBtn.addEventListener("click", ()=> { const current = svcBody.querySelectorAll("tr.svcRow").length; if (current < 10) addRow(); });
  }

  // Load existing services from DB
  (function initServices(){
    let existing = [];
    try{
      const raw = (servicesJsonEl && servicesJsonEl.value) ? servicesJsonEl.value : "";
      if (raw) existing = JSON.parse(raw);
    } catch(e){ existing = []; }

    if (existing && existing.length){
      existing.forEach(it=> addRow(it));
      addRow(); // trailing empty row
    } else {
      addRow(); // initial empty row
    }
    recalcTotals();
  })();

  // bind submit
  const bf = document.getElementById("bookingForm");
  if (bf){
    bf.addEventListener("submit", function(){
      recalcTotals();
    });
  }

</script>
{% endblock %}
""",

    "booking_view.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="grid two">
  <div class="card">
    <h3>Exam Details</h3>
    <div class="muted">
      <b>{{ b.pet_name }}</b> • Owner: <b>{{ b.owner_name }}</b><br>
      Starts: <b>{{ b.appointment_start }}</b> • Duration: {{ b.duration_min }} min<br>
      Type: {{ b.appointment_type }} • Status: {{ b.status }} • Priority: {{ b.priority }}<br>
      Vitals: Weight <b>{{ b.visit_weight_kg or "—" }}</b> kg • Temp <b>{{ b.visit_temp_c or "—" }}</b> °C
    </div>

    <div class="hr"></div>

    <div class="row-actions">
      <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp Journey</a>
      <a class="btn" href="{{ b.portal_link }}" target="_blank">Owner Portal</a>
      <a class="btn" href="{{ b.ics_link }}">Download .ics</a>
            <a class="btn primary" href="{{ b.google_cal_link }}" target="_blank">Google Calendar</a>
      <a class="btn good" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Completed') }}">Mark Completed</a>
      <a class="btn bad" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Cancelled') }}" onclick="return confirm('Cancel this exam?')">Cancel</a>
    </div>

    <div class="hr"></div>

    <div class="muted">
      Phone: {{ b.owner_phone }}<br>
      Email: {{ b.owner_email }}<br>
      <div class="hr"></div>
      <b>Financials</b><br>
      Subtotal (services): <b>{{ b.service_fee or '0.00' }}</b><br>
      Discount: <b>{{ b.discount or '0.00' }}</b><br>
      Final Total: <b>{{ b.fee_amount or b.service_fee or '0.00' }}</b><br>
      Paid: <b>{{ b.paid_amount or '0.00' }}</b><br>
      Due: <b>{{ b.due_amount or '0.00' }}</b><br>
      Payment Channel: <b>{{ b.payment_channel or b.payment_method or '—' }}</b><br>
      Owner Confirmed: <b>{{ b.owner_confirmed or '—' }}</b><br>
      Owner Update: <span class="small">{{ b.owner_update_message or '—' }}</span><br>
      Owner Update Time: <span class="small">{{ b.owner_update_datetime or '—' }}</span>
    </div>

    <div class="hr"></div>

    <div class="row-actions">
      <a class="btn" href="{{ url_for('booking_edit', booking_id=b.id) }}">Edit Exam</a>
      <a class="btn bad" href="{{ url_for('booking_delete', booking_id=b.id) }}" onclick="return confirm('Delete exam?')">Delete</a>
    </div>
  </div>

  <div class="card">
    <h3>AI Copilot (Futuristic)</h3>
    <div class="muted">Generates triage + risk flags + questions + tests. Decision support only.</div>

    <div class="hr"></div>

    <div id="aiBox" class="muted">Loading AI insights...</div>

    <div class="hr"></div>

    <div class="row-actions">
      <button class="btn warn" type="button" onclick="reloadAI()">Regenerate</button>
      <a class="btn good" id="aiApplyBtn" href="{{ url_for('booking_apply_ai', booking_id=b.id) }}">Apply AI Suggestions</a>
    </div>
  </div>
</div>

<script>
  async function reloadAI(){
    const box = document.getElementById("aiBox");
    box.textContent = "Loading AI insights...";
    const res = await fetch("{{ url_for('api_copilot', booking_id=b.id) }}");
    const data = await res.json();

    if (data.error){
      box.textContent = data.error;
      return;
    }

    const flags = (data.flags||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";
    const qs = (data.questions||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";
    const ts = (data.tests||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";

    box.textContent =
      "Triage: " + data.triage + "\n" +
      "Suggested Priority: " + data.suggested_priority + "\n" +
      "Suggested Type: " + data.suggested_type + "\n" +
      "Suggested Duration: " + data.suggested_duration_min + " min\n\n" +
      "No-show risk: " + data.no_show_risk.bucket + " (" + data.no_show_risk.score + "%)\n" +
      (data.no_show_risk.tips && data.no_show_risk.tips.length ? ("Tips:\n" + data.no_show_risk.tips.map(x=>"- "+x).join("\n") + "\n\n") : "\n") +
      "Risk Flags:\n" + flags + "\n\n" +
      "Questions Checklist:\n" + qs + "\n\n" +
      "Suggested Tests:\n" + ts + "\n\n" +
      "Draft:\n" + (data.plan_text || "") + "\n\n" +
      "Note: " + (data.disclaimer || "");
  }

  reloadAI();
</script>
{% endblock %}
""",

    "reminders.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Reminders Center</h3>
  <div class="muted">Upcoming + saved reminders. Use Journey buttons (WA + Portal + .ics).</div>

  <div class="hr"></div>


  <div class="card">
    <h3>Pet 360 (Search)</h3>
    <div class="muted">Search by Owner Name / Phone / Pet Name, then choose the pet from the dropdown to view: full medical history, upcoming & all reminders, and financial status.</div>

    <form method="get" class="form" style="margin-top:10px;">
      <div class="full">
        <label>Search</label>
        <input name="q" value="{{ q or '' }}" placeholder="Owner / Phone / Pet">
      </div>
      <div class="full row-actions">
        <button class="btn primary" type="submit">Search</button>
        <a class="btn" href="{{ url_for('reminders') }}">Clear</a>
      </div>
    </form>

    {% if q and not pet_matches %}
      <div class="hr"></div>
      <div class="muted">No matching pets found for: <b>{{ q }}</b></div>
    {% endif %}

    {% if pet_matches %}
      <div class="hr"></div>
      <form method="get" class="form">
        <input type="hidden" name="q" value="{{ q or '' }}">
        <div class="full">
          <label>Select Pet</label>
          <select name="pet_id" onchange="this.form.submit()">
            <option value="">Choose...</option>
            {% for p in pet_matches %}
              <option value="{{ p.pet_id }}" {% if selected_pet_id==p.pet_id %}selected{% endif %}>{{ p.label }}</option>
            {% endfor %}
          </select>
        </div>
      </form>
    {% endif %}

    {% if pet360 %}
      <div class="hr"></div>

      <div class="grid two">
        <div class="card">
          <h3>Profile</h3>
          <div class="muted">
            <div><b>Pet:</b> {{ pet360.pet_name }} • {{ pet360.species or '—' }} • {{ pet360.breed or '—' }}</div>
            <div><b>Sex:</b> {{ pet360.sex or '—' }} • <b>Age:</b> {{ pet360.age or '—' }} • <b>Weight:</b> {{ pet360.weight or '—' }}</div>
            <div class="hr"></div>
            <div><b>Owner:</b> {{ pet360.owner_name }} ({{ pet360.owner_phone or '—' }})</div>
            <div><b>Email:</b> {{ pet360.owner_email or '—' }}</div>
            <div><b>Address:</b> {{ pet360.owner_address or '—' }}</div>
          </div>
        </div>

        <div class="card">
          <h3>Financial Status</h3>
          <div class="muted">
            <div><b>Total (incl. VAT):</b> {{ pet_fin.total }}</div>
            <div><b>Paid:</b> {{ pet_fin.paid }}</div>
            <div><b>Due:</b> {{ pet_fin.due }}</div>
            <div class="hr"></div>
            <div><b>Bookings:</b> {{ pet_fin.bookings_count }} • <b>Unpaid/Partial:</b> {{ pet_fin.open_count }}</div>
          </div>
          {% if pet_fin.open_items %}
            <div class="hr"></div>
            <div style="overflow:auto">
              <table>
                <thead><tr><th>Date</th><th>Invoice</th><th>Status</th><th style="text-align:right">Due</th><th>Action</th></tr></thead>
                <tbody>
                  {% for x in pet_fin.open_items %}
                    <tr>
                      <td><b>{{ x.appointment_start }}</b></td>
                      <td>{{ x.invoice_no or '—' }}</td>
                      <td>{{ x.payment_status or '—' }}</td>
                      <td style="text-align:right"><b>{{ x.due }}</b></td>
                      <td><a class="btn primary" href="{{ url_for('booking_view', booking_id=x.booking_id) }}">Open</a></td>
                    </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
          {% endif %}
        </div>
      </div>

      <div class="hr"></div>

      <details open>
        <summary><b>Upcoming Reminders</b> ({{ pet_reminders_upcoming|length }})</summary>
        <div style="overflow:auto;margin-top:10px">
          <table>
            <thead><tr><th>Scheduled</th><th>Service</th><th>Type</th><th>Status</th><th>Action</th></tr></thead>
            <tbody>
              {% for r in pet_reminders_upcoming %}
                <tr>
                  <td><b>{{ r.scheduled_for }}</b></td>
                  <td>{{ r.service_name }}</td>
                  <td>{{ r.reminder_type }}</td>
                  <td><span class="pill {% if r.status=='Sent' %}good{% elif r.status=='Opened' %}warn{% endif %}">{{ r.status }}</span></td>
                  <td style="white-space:nowrap">
                    <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                    <a class="btn good" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a>
                  </td>
                </tr>
              {% endfor %}
              {% if not pet_reminders_upcoming %}
                <tr><td colspan="5" class="muted">No upcoming reminders.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </details>

      <div class="hr"></div>

      <details>
        <summary><b>All Reminders</b> ({{ pet_reminders_all|length }})</summary>
        <div style="overflow:auto;margin-top:10px">
          <table>
            <thead><tr><th>Scheduled</th><th>Service</th><th>Status</th><th>Channel</th><th>Action</th></tr></thead>
            <tbody>
              {% for r in pet_reminders_all %}
                <tr>
                  <td><b>{{ r.scheduled_for }}</b></td>
                  <td>{{ r.service_name }}</td>
                  <td><span class="pill {% if r.status=='Sent' %}good{% elif r.status=='Opened' %}warn{% endif %}">{{ r.status }}</span></td>
                  <td>{{ r.channel }}</td>
                  <td style="white-space:nowrap">
                    <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                    <a class="btn good" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a>
                    <a class="btn" href="{{ url_for('reminder_mark_sent', reminder_id=r.id) }}">Mark Sent</a>
                  </td>
                </tr>
              {% endfor %}
              {% if not pet_reminders_all %}
                <tr><td colspan="5" class="muted">No reminders found for this pet.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </details>

      <div class="hr"></div>

      <details>
        <summary><b>Medical History</b> ({{ pet_history|length }} visits)</summary>
        {% for h in pet_history %}
          <div class="hr"></div>
          <details>
            <summary>
              <b>{{ h.appointment_start }}</b> — {{ h.appointment_type }} • {{ h.status }}
              &nbsp;|&nbsp; Vet: {{ h.vet_name or '—' }}
              &nbsp;|&nbsp; Total: {{ h.total }} • Due: {{ h.due }}
            </summary>
            <div style="margin-top:10px" class="muted">
              <div><b>Services:</b> {{ h.services_summary or '—' }}</div>
              <div><b>Reason:</b> {{ h.reason or '—' }}</div>
              <div><b>Symptoms:</b> {{ h.symptoms or '—' }}</div>
              <div class="hr"></div>
              <div><b>Diagnosis:</b> {{ h.diagnosis or '—' }}</div>
              <div><b>Treatment Plan:</b> {{ h.treatment_plan or '—' }}</div>
              <div><b>Prescription:</b> {{ h.prescription or '—' }}</div>
              <div><b>Lab Tests:</b> {{ h.lab_tests or '—' }}</div>
              <div><b>Vaccines Given:</b> {{ h.vaccines_given or '—' }}</div>
              <div><b>Follow-up:</b> {{ h.followup_datetime or '—' }}</div>
              <div class="hr"></div>
              <div><b>Invoice:</b> {{ h.invoice_no or '—' }} • <b>Payment:</b> {{ h.payment_status or '—' }} • <b>Paid:</b> {{ h.paid }}</div>
              <div><b>Notes:</b> {{ h.notes or '—' }}</div>
              <div class="row-actions" style="margin-top:10px">
                <a class="btn primary" href="{{ url_for('booking_view', booking_id=h.booking_id) }}">Open Visit</a>
              </div>
            </div>
          </details>
        {% endfor %}
        {% if not pet_history %}
          <div class="hr"></div>
          <div class="muted">No visits found for this pet yet.</div>
        {% endif %}
      </details>

    {% endif %}
  </div>

  <div class="grid two">
    <div class="card">
      <h3>Upcoming in next 24 hours</h3>
      <div style="overflow:auto;margin-top:10px">
        <table>
          <thead><tr><th>Time</th><th>Pet</th><th>Owner</th><th>Action</th></tr></thead>
          <tbody>
            {% for b in upcoming %}
              <tr>
                <td><b>{{ b.appointment_start }}</b></td>
                <td>{{ b.pet_name }}</td>
                <td>{{ b.owner_name }}</td>
                <td style="white-space:nowrap">
                  <a class="btn primary" href="{{ url_for('booking_view', booking_id=b.id) }}">Open</a>
                  <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp</a>
                  <a class="btn" href="{{ b.portal_link }}" target="_blank">Portal</a>
                  <a class="btn" href="{{ b.ics_link }}">.ics</a>
                </td>
              </tr>
            {% endfor %}
            {% if not upcoming %}
              <tr><td colspan="4" class="muted">No upcoming appointments.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <h3>Reminder Records</h3>
      <div style="overflow:auto;margin-top:10px">
        <table>
          <thead><tr><th>Scheduled</th><th>Pet</th><th>Status</th><th>Action</th></tr></thead>
          <tbody>
            {% for r in reminder_rows %}
              <tr>
                <td><b>{{ r.scheduled_for }}</b></td>
                <td>{{ r.pet_name }}</td>
                <td>
                  <span class="pill {% if r.status=='Sent' %}good{% elif r.status=='Opened' %}warn{% endif %}">{{ r.status }}</span>
                </td>
                <td style="white-space:nowrap">
                  <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                  <a class="btn good" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a>
                  <a class="btn" href="{{ url_for('reminder_mark_sent', reminder_id=r.id) }}">Mark Sent</a>
                </td>
              </tr>
            {% endfor %}
            {% if not reminder_rows %}
              <tr><td colspan="4" class="muted">No reminder records yet.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

</div>
{% endblock %}
""",

    "dashboard.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="grid three">
  <div class="card">
    <h3>Total Bookings</h3>
    <div style="font-size:26px;font-weight:800;margin-top:4px">{{ kpi.total }}</div>
    <div class="muted">All time</div>
  </div>
  <div class="card">
    <h3>Upcoming (24h)</h3>
    <div style="font-size:26px;font-weight:800;margin-top:4px">{{ kpi.upcoming }}</div>
    <div class="muted">Next 24 hours</div>
  </div>
  <div class="card">
    <h3>Paid Revenue (6m)</h3>
    <div style="font-size:26px;font-weight:800;margin-top:4px">{{ kpi.revenue }}</div>
    <div class="muted">Paid bookings only</div>
  </div>
</div>

<div class="grid two" style="margin-top:14px;">
  <div class="card">
    <h3>Bookings by Status</h3>
    <img src="{{ url_for('chart_status_png') }}" style="width:100%;border-radius:14px;border:1px solid rgba(255,255,255,.10);">
  </div>
  <div class="card">
    <h3>Appointments Trend (14 days)</h3>
    <img src="{{ url_for('chart_trend_png') }}" style="width:100%;border-radius:14px;border:1px solid rgba(255,255,255,.10);">
  </div>
</div>

<div class="grid two" style="margin-top:14px;">
  <div class="card">
    <h3>Pets by Species</h3>
    <img src="{{ url_for('chart_species_png') }}" style="width:100%;border-radius:14px;border:1px solid rgba(255,255,255,.10);">
  </div>
  <div class="card">
    <h3>Paid Revenue (6 months)</h3>
    <img src="{{ url_for('chart_revenue_png') }}" style="width:100%;border-radius:14px;border:1px solid rgba(255,255,255,.10);">
  </div>
</div>
{% endblock %}
""",

    "history.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>History</h3>
  <div class="muted">Past bookings (audit view).</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Search</label>
      <input name="q" value="{{ q }}" placeholder="owner / pet / type / status / reason">
    </div>
    <div class="full row-actions">
      <button class="btn primary" type="submit">Search</button>
      <a class="btn" href="{{ url_for('bookings') }}">Back to Bookings</a>
    </div>
  </form>

  <div class="hr"></div>

  <div style="overflow:auto">
    <table>
      <thead><tr>
        <th>Date</th><th>Owner</th><th>Pet</th><th>Type</th><th>Status</th>
      </tr></thead>
      <tbody>
        {% for b in rows %}
          <tr>
            <td><b>{{ b.appointment_start }}</b></td>
            <td>{{ b.owner_name }}</td>
            <td>{{ b.pet_name }}</td>
            <td>{{ b.appointment_type }}</td>
            <td>{{ b.status }}</td>
          </tr>
        {% endfor %}
        {% if not rows %}
          <tr><td colspan="6" class="muted">No history items.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

    "copilot.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>AI Copilot Center</h3>
  <div class="muted">Select a booking to view AI output. You can also use Booking page AI panel.</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Select Booking</label>
      <select name="booking_id" onchange="this.form.submit()">
        <option value="">Select...</option>
        {% for x in items %}
          <option value="{{ x.id }}" {% if selected and selected.id==x.id %}selected{% endif %}>
            {{ x.appointment_start }} — {{ x.pet_name }} / {{ x.owner_name }}
          </option>
        {% endfor %}
      </select>
    </div>
  </form>

  {% if selected %}
    <div class="hr"></div>
    <div class="row-actions">
      <a class="btn primary" href="{{ url_for('booking_view', booking_id=selected.id) }}">Open Exam</a>
      <a class="btn good" href="{{ url_for('booking_apply_ai', booking_id=selected.id) }}">Apply AI Suggestions</a>
    </div>
    <div class="hr"></div>
    <pre style="white-space:pre-wrap;color:rgba(255,255,255,.85);font-size:12px">{{ copilot_text }}</pre>
  {% endif %}
</div>
{% endblock %}
""",

    "portal.html": r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Owner Portal | {{ app_title }}</title>
  <style>
    body{margin:0;font-family:ui-sans-serif,system-ui;background:radial-gradient(1200px 800px at 20% 10%, #152853 0%, #0b1220 55%);color:rgba(255,255,255,.92);min-height:100vh;display:flex;align-items:center;justify-content:center}
    .card{width:760px;max-width:calc(100vw - 30px);padding:18px;border-radius:18px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.05);box-shadow:0 12px 40px rgba(0,0,0,.35)}
    h1{margin:0;font-size:18px}
    .muted{color:rgba(255,255,255,.70);font-size:12px;margin-top:6px}
    label{font-size:12px;color:rgba(255,255,255,.70);margin-bottom:5px;display:block}
    input,select,textarea{width:100%;box-sizing:border-box;padding:11px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(0,0,0,.18);color:rgba(255,255,255,.92);outline:none}
    textarea{min-height:100px;resize:vertical}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:12px}
    .full{grid-column:1/-1}
    .btn{display:inline-flex;align-items:center;justify-content:center;padding:11px 12px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(96,165,250,.18);color:rgba(255,255,255,.92);cursor:pointer;text-decoration:none}
    .btn.good{background:rgba(34,197,94,.18)}
    .btn.warn{background:rgba(245,158,11,.18)}
    .row{display:flex;gap:10px;flex-wrap:wrap;margin-top:12px}
    .flash{padding:10px 12px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.06);margin-top:10px}
  </style>
</head>
<body>
  <div class="card">
    <h1>{{ app_title }} — Owner Portal</h1>
    <div class="muted">Confirm appointment, request reschedule, and add symptoms. This is a secure link.</div>

    {% if msg %}
      <div class="flash">{{ msg }}</div>
    {% endif %}

    <div class="grid">
      <div class="full">
        <div class="muted">
          Pet: <b>{{ pet_name }}</b><br>
          Appointment: <b>{{ appointment_start }}</b><br>
          Clinic: {{ app_title }}
        </div>
      </div>
    </div>

    <form method="post" class="grid">
      <div>
        <label>Action</label>
        <select name="action">
          <option value="confirm">Confirm</option>
          <option value="reschedule">Request Reschedule</option>
          <option value="update">Update Symptoms / Notes</option>
        </select>
      </div>
      <div>
        <label>Preferred Contact</label>
        <select name="contact">
          <option>WhatsApp</option>
          <option>Phone</option>
          <option>Email</option>
        </select>
      </div>
      <div class="full">
        <label>Message (optional)</label>
        <textarea name="message" placeholder="Write your note to the clinic..."></textarea>
      </div>

      <div class="full row">
        <button class="btn good" type="submit">Submit</button>
        <a class="btn" href="{{ wa_link }}" target="_blank">Open WhatsApp</a>
      </div>
    </form>
  </div>
</body>
</html>
"""
}

TEMPLATES["config.html"] = r"""
{% extends "base.html" %}
{% block content %}
<style>
  .grid.cols-2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
  .grid.cols-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
  .check{display:flex;gap:10px;align-items:flex-start}
  .check input{margin-top:3px}
  details{border:1px solid rgba(255,255,255,.10);border-radius:12px;padding:10px;background:rgba(255,255,255,.02)}
  summary{cursor:pointer}
</style>
  <div class="grid two">

    <div class="card">
      <h3>Users & Roles</h3>
      <div class="muted">Create users and assign roles. Only <b>admin</b> can access this page.</div>
      <div class="hr"></div>

      <form method="post" action="{{ url_for('config_add_user') }}" class="grid two">
        <div>
          <label>Username</label>
          <input name="username" placeholder="e.g., mohamed" required>
        </div>
        <div>
          <label>Password</label>
          <input name="password" placeholder="Set password" required>
        </div>
        <div>
          <label>Role</label>
          <select name="role" required>
            {% for r in roles %}
              <option value="{{ r }}">{{ r }}</option>
            {% endfor %}
          </select>
        </div>
        <div style="display:flex;align-items:end">
          <button class="btn primary" type="submit">Add User</button>
        </div>
      </form>

      <div class="hr"></div>
      <div style="overflow:auto">
        <table>
          <thead>
            <tr>
              <th>Username</th><th>Role</th><th>Status</th><th style="width:140px">Action</th>
            </tr>
          </thead>
          <tbody>
            {% for u in users %}
              <tr>
                <td><b>{{ u.username }}</b></td>
                <td>{{ u.role }}</td>
                <td>{% if u.active in ['1',1,True,'true','yes'] %}<span class="pill good">Active</span>{% else %}<span class="pill warn">Inactive</span>{% endif %}</td>
                <td>
                  {% if u.username != 'admin' %}
                  <form method="post" action="{{ url_for('config_toggle_user', user_id=u.id) }}">
                    <button class="btn" type="submit">{% if u.active in ['1',1,True,'true','yes'] %}Disable{% else %}Enable{% endif %}</button>
                  </form>
                  {% else %}
                    <span class="muted">Protected</span>
                  {% endif %}
                </td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <h2>Roles, Privileges & Rules</h2>
      <p class="hint">Define what each role can access (RBAC). Admin always has full access. Use <code>config_manage</code> to allow non-admin roles to access the Config page.</p>

      {% for role in roles %}
        <details class="mt">
          <summary><b>{{ role }}</b></summary>
          <form method="post" action="{{ url_for('config_update_role_permissions') }}" class="mt">
            <input type="hidden" name="role" value="{{ role }}">
            <div class="grid cols-2">
              {% for code,label in perms_catalog %}
              <label class="check">
                <input type="checkbox" name="perms" value="{{ code }}" {% if code in role_perms.get(role, []) %}checked{% endif %}>
                <span>{{ label }} <span class="hint">({{ code }})</span></span>
              </label>
              {% endfor %}
            </div>
            <div class="row-actions">
              <button class="btn" type="submit">Save permissions</button>
            </div>
          </form>
        </details>
      {% endfor %}
    </div>

    <div class="card">
      <h3>Vets</h3>
      <div class="muted">These names appear in the Exam forms (Vet Name dropdown).</div>
      <div class="hr"></div>

      <form method="post" action="{{ url_for('config_add_vet') }}" class="row-actions">
        <input name="name" placeholder="Add vet name (e.g., Dr. Ahmed)" required style="flex:1">
        <button class="btn primary" type="submit">Add Vet</button>
      </form>

      <div class="hr"></div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Name</th><th>Status</th><th style="width:140px">Action</th></tr></thead>
          <tbody>
            {% for v in vets_rows %}
              <tr>
                <td><b>{{ v.name }}</b></td>
                <td>{% if v.active in ['1',1,True,'true','yes'] %}<span class="pill good">Active</span>{% else %}<span class="pill warn">Inactive</span>{% endif %}</td>
                <td>
                  <form method="post" action="{{ url_for('config_toggle_vet', vet_id=v.id) }}">
                    <button class="btn" type="submit">{% if v.active in ['1',1,True,'true','yes'] %}Disable{% else %}Enable{% endif %}</button>
                  </form>
                </td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

      <div class="hr"></div>

      <h3>Rooms</h3>
      <div class="muted">These values appear in the Exam forms (Room dropdown).</div>
      <div class="hr"></div>

      <form method="post" action="{{ url_for('config_add_room') }}" class="row-actions">
        <input name="name" placeholder="Add room (e.g., Room 5)" required style="flex:1">
        <button class="btn primary" type="submit">Add Room</button>
      </form>

      <div class="hr"></div>
      <div style="overflow:auto">
        <table>
          <thead><tr><th>Name</th><th>Status</th><th style="width:140px">Action</th></tr></thead>
          <tbody>
            {% for r in rooms_rows %}
              <tr>
                <td><b>{{ r.name }}</b></td>
                <td>{% if r.active in ['1',1,True,'true','yes'] %}<span class="pill good">Active</span>{% else %}<span class="pill warn">Inactive</span>{% endif %}</td>
                <td>
                  <form method="post" action="{{ url_for('config_toggle_room', room_id=r.id) }}">
                    <button class="btn" type="submit">{% if r.active in ['1',1,True,'true','yes'] %}Disable{% else %}Enable{% endif %}</button>
                  </form>
                </td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

      <div class="hr"></div>
      <div class="muted">Future configuration items can be added here (e.g., services, pricing, checklists, templates).</div>
    </div>

  </div>


<div class="card">
  <h3>Services</h3>
  <div class="muted">Add/update services used in Exam pages. Active services will appear in dropdowns.</div>
  <div class="hr"></div>

  <form method="post" action="{{ url_for('config_add_service') }}" class="form" style="margin-bottom:14px;">
    <div>
      <label>Service Name</label>
      <input name="service_name" placeholder="e.g., X-Ray">
    </div>
    <div>
      <label>Cost</label>
      <input name="service_cost" type="number" step="0.01" min="0" placeholder="0.00">
    </div>
    <div>
      <label>Fee</label>
      <input name="service_fee" type="number" step="0.01" min="0" placeholder="0.00">
    </div>
    <div class="full row-actions">
      <button class="btn good" type="submit">Add Service</button>
    </div>
  </form>

  <div style="overflow:auto">
    <table>
      <thead><tr><th>Name</th><th class="right">Cost</th><th class="right">Fee</th><th class="right">Margin</th><th class="right">Margin %</th><th>Status</th><th>Actions</th></tr></thead>
      <tbody>
        {% for s in services %}
          <tr>
            <td><b>{{ s.name }}</b></td>
            <td class="right">{{ s.cost if s.cost is defined else '' }}</td>
            <td class="right">{{ s.fee }}</td>
            <td class="right">{{ s.margin if s.margin is defined else '' }}</td>
            <td class="right">{{ s.margin_pct if s.margin_pct is defined else '' }}</td>
            <td>
              {% if s.active in ['1',1,True,'true','True'] %}
                <span class="pill green">Active</span>
              {% else %}
                <span class="pill red">Disabled</span>
              {% endif %}
            </td>
            <td>
  <form method="post" action="{{ url_for('config_update_service', service_id=s.id) }}" style="display:inline">
    <input name="service_name" value="{{ s.name }}" style="width:180px;">
    <input name="service_cost" type="number" step="0.01" min="0" value="{{ s.cost if s.cost is defined else '' }}" style="width:110px;" placeholder="Cost">
    <input name="service_fee" type="number" step="0.01" min="0" value="{{ s.fee }}" style="width:110px;" placeholder="Fee">
    <button class="btn" type="submit">Save</button>
  </form>
  <form method="post" action="{{ url_for('config_toggle_service', service_id=s.id) }}" style="display:inline;margin-left:6px;">
    <button class="btn {% if s.active in ['1',1,True,'true','True'] %}bad{% else %}good{% endif %}" type="submit">
      {% if s.active in ['1',1,True,'true','True'] %}Disable{% else %}Enable{% endif %}
    </button>
  </form>
</td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>


<div class="card">
  <h2>WhatsApp Templates</h2>
  <p class="hint">
    Create multiple WhatsApp templates and choose the best one per scenario. When you click <b>WhatsApp</b> / <b>Remind</b>, the message will include the <b>service name</b> and booking details.
  </p>
  <div class="hint">
    Placeholders:
    <code>{owner_name}</code>, <code>{pet_name}</code>, <code>{service_name}</code>, <code>{appointment_start}</code>,
    <code>{booking_details}</code>, <code>{reason}</code>, <code>{notes}</code>, <code>{clinic_name}</code>
  </div>

  <h3 class="mt">Add Template</h3>
  <form method="post" action="{{ url_for('config_add_whatsapp_template') }}">
    <div class="grid cols-3">
      <div>
        <label>Name</label>
        <input name="wa_name" required placeholder="e.g., Reminder - Vaccination">
      </div>
      <div>
        <label>Scenario</label>
        <select name="wa_scenario">
          {% for s in wa_scenarios %}
            <option value="{{ s }}">{{ s }}</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Booking Type</label>
        <select name="wa_booking_type">
          <option value="Any" selected>Any</option>
          {% for bt in wa_booking_types %}
            <option value="{{ bt }}">{{ bt }}</option>
          {% endfor %}
        </select>
      </div>
    </div>

    <div class="grid cols-2 mt">
      <div>
        <label>Template Text</label>
        <textarea name="wa_template_text" rows="4" required placeholder="Write your WhatsApp message..."></textarea>
      </div>
      <div>
        <label>Options</label>
        <div class="grid cols-2">
          <label class="check"><input type="checkbox" name="wa_is_default" value="true"> <span>Set as default</span></label>
          <label class="check"><input type="checkbox" name="wa_active" value="true" checked> <span>Active</span></label>
        </div>
        <p class="hint mt">Tip: Use separate templates for Appointment vs Service reminders.</p>
      </div>
    </div>

    <div class="row-actions">
      <button class="btn" type="submit">Create</button>
    </div>
  </form>

  <h3 class="mt">Existing Templates</h3>
  <table class="table mt">
    <thead>
      <tr>
        <th>Name</th>
        <th>Scenario</th>
        <th>Booking Type</th>
        <th>Default</th>
        <th>Active</th>
        <th>Template</th>
        <th>Actions</th>
      </tr>
    </thead>
    <tbody>
      {% for t in wa_templates %}
      <tr>
        <td><b>{{ t.name }}</b></td>
        <td>{{ t.scenario }}</td>
        <td>{{ t.booking_type }}</td>
        <td>{% if t.is_default %}<span class="pill green">Yes</span>{% else %}<span class="pill">No</span>{% endif %}</td>
        <td>{% if t.active %}<span class="pill green">Yes</span>{% else %}<span class="pill">No</span>{% endif %}</td>
        <td style="min-width:320px">
          <form method="post" action="{{ url_for('config_update_whatsapp_template', template_id=t.id) }}">
            <div class="grid cols-3">
              <div>
                <label class="hint">Name</label>
                <input name="wa_name" value="{{ t.name }}" required>
              </div>
              <div>
                <label class="hint">Scenario</label>
                <select name="wa_scenario">
                  {% for s in wa_scenarios %}
                    <option value="{{ s }}" {% if s==t.scenario %}selected{% endif %}>{{ s }}</option>
                  {% endfor %}
                </select>
              </div>
              <div>
                <label class="hint">Booking Type</label>
                <select name="wa_booking_type">
                  <option value="Any" {% if t.booking_type=='Any' %}selected{% endif %}>Any</option>
                  {% for bt in wa_booking_types %}
                    <option value="{{ bt }}" {% if bt==t.booking_type %}selected{% endif %}>{{ bt }}</option>
                  {% endfor %}
                </select>
              </div>
            </div>

            <div class="mt">
              <label class="hint">Template Text</label>
              <textarea name="wa_template_text" rows="3" required>{{ t.template_text }}</textarea>
            </div>

            <div class="grid cols-3 mt">
              <label class="check"><input type="checkbox" name="wa_is_default" value="true" {% if t.is_default %}checked{% endif %}> <span>Default</span></label>
              <label class="check"><input type="checkbox" name="wa_active" value="true" {% if t.active %}checked{% endif %}> <span>Active</span></label>
              <div class="row-actions" style="justify-content:flex-end">
                <button class="btn" type="submit">Save</button>
              </div>
            </div>
          </form>
        </td>
        <td>
          <div class="row-actions">
            <form method="post" action="{{ url_for('config_set_default_whatsapp_template', template_id=t.id) }}">
              <button class="btn" type="submit">Set Default</button>
            </form>
            <form method="post" action="{{ url_for('config_toggle_whatsapp_template', template_id=t.id) }}">
              <button class="btn" type="submit">{{ 'Disable' if t.active else 'Enable' }}</button>
            </form>
          </div>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

</div>


{% endblock %}
"""

TEMPLATES["report.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <div style="display:flex;align-items:center;justify-content:space-between;gap:10px">
    <div>
      <div class="h">Reports</div>
      <div class="muted">Choose criteria and a report view; every chart/table below is recalculated based on your filters.</div>
    </div>
    <form method="get" action="{{ url_for('report') }}" style="display:flex;gap:8px;flex-wrap:wrap;align-items:end;justify-content:flex-end">
      <div>
        <div class="muted" style="font-size:12px">From</div>
        <input type="date" name="start" value="{{ start }}" style="height:36px">
      </div>
      <div>
        <div class="muted" style="font-size:12px">To</div>
        <input type="date" name="end" value="{{ end }}" style="height:36px">
      </div>
      <div>
        <div class="muted" style="font-size:12px">Report</div>
        <select name="view" style="height:36px;min-width:220px">
          <option value="all" {% if view=='all' %}selected{% endif %}>All (8 reports + Futuristic)</option>
          <option value="summary" {% if view=='summary' %}selected{% endif %}>1) Executive Summary</option>
          <option value="trend" {% if view=='trend' %}selected{% endif %}>2) Workload Trend</option>
          <option value="status" {% if view=='status' %}selected{% endif %}>3) Status Distribution</option>
          <option value="financial" {% if view=='financial' %}selected{% endif %}>4) Financials & Collections</option>
          <option value="service" {% if view=='service' %}selected{% endif %}>5) Service Mix</option>
          <option value="vet" {% if view=='vet' %}selected{% endif %}>6) Vet Performance</option>
          <option value="room" {% if view=='room' %}selected{% endif %}>7) Room Utilization</option>
          <option value="closure" {% if view=='closure' %}selected{% endif %}>Daily Closure</option>
          <option value="due" {% if view=='due' %}selected{% endif %}>8) Outstanding Due</option>
          <option value="future" {% if view=='future' %}selected{% endif %}>Futuristic Insights (Forecast + Alerts)</option>
        </select>
      </div>

      <div style="min-width:240px">
        <div class="muted" style="font-size:12px">Owner name contains</div>
        <input name="owner_q" value="{{ owner_q }}" placeholder="e.g., Ahmed" style="height:36px;width:100%">
      </div>
      <div style="min-width:200px">
        <div class="muted" style="font-size:12px">Phone contains</div>
        <input name="phone_q" value="{{ phone_q }}" placeholder="digits" style="height:36px;width:100%">
      </div>
      <div style="min-width:200px">
        <div class="muted" style="font-size:12px">Pet name contains</div>
        <input name="pet_q" value="{{ pet_q }}" placeholder="e.g., Luna" style="height:36px;width:100%">
      </div>

      <div>
        <div class="muted" style="font-size:12px">Vet</div>
        <select name="vet" style="height:36px;min-width:170px">
          <option value="">All</option>
          {% for v in vets %}
            <option value="{{ v }}" {% if vet_f==v %}selected{% endif %}>{{ v }}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <div class="muted" style="font-size:12px">Room</div>
        <select name="room" style="height:36px;min-width:140px">
          <option value="">All</option>
          {% for r in rooms %}
            <option value="{{ r }}" {% if room_f==r %}selected{% endif %}>{{ r }}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <div class="muted" style="font-size:12px">Status</div>
        <select name="status" style="height:36px;min-width:170px">
          <option value="">All</option>
          {% for s in statuses %}
            <option value="{{ s }}" {% if status_f==s %}selected{% endif %}>{{ s }}</option>
          {% endfor %}
        </select>
      </div>

      <div>
        <div class="muted" style="font-size:12px">Service</div>
        <select name="service" style="height:36px;min-width:210px">
          <option value="">All</option>
          {% for s in services %}
            <option value="{{ s }}" {% if service_f==s %}selected{% endif %}>{{ s }}</option>
          {% endfor %}
        </select>
      </div>

      <div style="width:140px">
        <div class="muted" style="font-size:12px">Min Due</div>
        <input type="number" step="0.01" min="0" name="min_due" value="{{ min_due }}" placeholder="0" style="height:36px;width:100%">
      </div>

      <div style="display:flex;gap:8px">
        <button class="btn good" type="submit" style="height:36px">Apply</button>
        <a class="btn" href="{{ url_for('report') }}" style="height:36px;display:inline-flex;align-items:center">Reset</a>
      </div>
    </form>
  </div>

  <div class="muted" style="margin-top:10px">
    Window: <b>{{ start }}</b> → <b>{{ end }}</b>.
    Active filters:
    {% if vet_f %}<span class="pill">Vet: {{ vet_f }}</span>{% endif %}
    {% if room_f %}<span class="pill">Room: {{ room_f }}</span>{% endif %}
    {% if status_f %}<span class="pill">Status: {{ status_f }}</span>{% endif %}
    {% if service_f %}<span class="pill">Service: {{ service_f }}</span>{% endif %}
    {% if owner_q %}<span class="pill">Owner: {{ owner_q }}</span>{% endif %}
    {% if phone_q %}<span class="pill">Phone: {{ phone_q }}</span>{% endif %}
    {% if pet_q %}<span class="pill">Pet: {{ pet_q }}</span>{% endif %}
    {% if min_due %}<span class="pill">Min Due: {{ min_due }}</span>{% endif %}
  </div>
</div>

<!-- 1) Executive Summary -->
{% if view in ['all','summary'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">1) Executive Summary</div>
    <div class="grid2" style="margin-top:10px">
      <div class="kpi"><div class="muted">Exams</div><div class="v">{{ total_exams }}</div></div>
      <div class="kpi"><div class="muted">Unique Owners</div><div class="v">{{ unique_owners }}</div></div>
      <div class="kpi"><div class="muted">Unique Pets</div><div class="v">{{ unique_pets }}</div></div>
      <div class="kpi"><div class="muted">Completion Rate</div><div class="v">{{ "%.1f"|format(completion_rate) }}%</div></div>
      <div class="kpi"><div class="muted">Revenue (after discount)</div><div class="v">{{ "%.2f"|format(fin.revenue_total) }}</div></div>
      <div class="kpi"><div class="muted">Discounts</div><div class="v">{{ "%.2f"|format(fin.discount_total) }}</div></div>
      <div class="kpi"><div class="muted">Collected</div><div class="v">{{ "%.2f"|format(fin.collected_total) }}</div></div>
      <div class="kpi"><div class="muted">Due</div><div class="v">{{ "%.2f"|format(fin.due_total) }}</div></div>
      <div class="kpi"><div class="muted">Collection Rate</div><div class="v">{{ "%.1f"|format(fin.collection_rate) }}%</div></div>
    </div>
  </div>
</div>
{% endif %}


<!-- 1.5) Daily Closure -->
{% if view in ['all','closure'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">1.5) Daily Closure</div>
    <div class="muted">End-of-day summary (bookings, revenue after discount, discounts, and payment channel totals).</div>
    <table class="table" style="margin-top:10px">
      <thead>
        <tr>
          <th>Date</th><th>Bookings</th><th>Revenue</th><th>Discounts</th>
          <th>Cash</th><th>Visa</th><th>Instapay</th><th>Other</th>
        </tr>
      </thead>
      <tbody>
        {% for r in closure_rows %}
          <tr>
            <td>{{ r.date }}</td>
            <td>{{ r.count }}</td>
            <td>{{ "%.2f"|format(r.revenue) }}</td>
            <td>{{ "%.2f"|format(r.discounts) }}</td>
            <td>{{ "%.2f"|format(r.cash) }}</td>
            <td>{{ "%.2f"|format(r.visa) }}</td>
            <td>{{ "%.2f"|format(r.instapay) }}</td>
            <td>{{ "%.2f"|format(r.other) }}</td>
          </tr>
        {% endfor %}
        {% if not closure_rows %}
          <tr><td colspan="8" class="muted">No bookings found in this range.</td></tr>
        {% endif %}
      </tbody>
    </table>

    <div class="hr"></div>
    <div class="h">Payment Breakdown (Totals)</div>
    <div class="muted">Totals are calculated from the same filtered dataset.</div>
    <div class="kpis" style="margin-top:10px">
      <div class="kpi"><div class="muted">Cash</div><div class="v">{{ "%.2f"|format(fin.channel_totals.get('Cash', 0.0) + fin.channel_totals.get('cash', 0.0)) }}</div></div>
      <div class="kpi"><div class="muted">Visa</div><div class="v">{{ "%.2f"|format(fin.channel_totals.get('Visa', 0.0) + fin.channel_totals.get('visa', 0.0)) }}</div></div>
      <div class="kpi"><div class="muted">Instapay</div><div class="v">{{ "%.2f"|format(fin.channel_totals.get('Instapay', 0.0) + fin.channel_totals.get('instapay', 0.0)) }}</div></div>
    </div>
  </div>
</div>
{% endif %}

<!-- 2) Workload Trend -->
{% if view in ['all','trend'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">2) Workload Trend</div>
    <div class="muted">Daily exams and revenue in the selected window.</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Date</th><th>Exams</th><th>Revenue</th></tr></thead>
      <tbody>
        {% for r in trend_rows %}
          <tr><td>{{ r.date }}</td><td>{{ r.count }}</td><td>{{ "%.2f"|format(r.revenue) }}</td></tr>
        {% endfor %}
        {% if trend_rows|length == 0 %}
          <tr><td colspan="3" class="muted">No data for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 3) Status Distribution -->
{% if view in ['all','status'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">3) Status Distribution</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Status</th><th>Count</th><th>%</th></tr></thead>
      <tbody>
        {% for s in status_rows %}
          <tr><td>{{ s.status }}</td><td>{{ s.count }}</td><td>{{ "%.1f"|format(s.pct) }}%</td></tr>
        {% endfor %}
        {% if status_rows|length == 0 %}
          <tr><td colspan="3" class="muted">No data for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 4) Financials & Collections -->
{% if view in ['all','financial'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">4) Financials & Collections</div>
    <table class="table" style="margin-top:10px">
      <tbody>
        <tr><th>Total Revenue</th><td>{{ "%.2f"|format(fin.revenue_total) }}</td></tr>
        <tr><th>Total Collected</th><td>{{ "%.2f"|format(fin.collected_total) }}</td></tr>
        <tr><th>Total Due</th><td>{{ "%.2f"|format(fin.due_total) }}</td></tr>
        <tr><th>Collection Rate</th><td>{{ "%.1f"|format(fin.collection_rate) }}%</td></tr>
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 5) Service Mix -->
{% if view in ['all','service'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">5) Service Mix (Top 10)</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Service</th><th>Count</th><th>Revenue</th></tr></thead>
      <tbody>
        {% for s in top_services %}
          <tr><td>{{ s.name }}</td><td>{{ s.count }}</td><td>{{ "%.2f"|format(s.revenue) }}</td></tr>
        {% endfor %}
        {% if top_services|length == 0 %}
          <tr><td colspan="3" class="muted">No data for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 6) Vet Performance -->
{% if view in ['all','vet'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">6) Vet Performance</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Vet</th><th>Exams</th><th>Completed %</th><th>Revenue</th></tr></thead>
      <tbody>
        {% for v in vet_rows %}
          <tr><td>{{ v.vet }}</td><td>{{ v.count }}</td><td>{{ "%.1f"|format(v.completed_pct) }}%</td><td>{{ "%.2f"|format(v.revenue) }}</td></tr>
        {% endfor %}
        {% if vet_rows|length == 0 %}
          <tr><td colspan="4" class="muted">No data for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 7) Room Utilization -->
{% if view in ['all','room'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">7) Room Utilization</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Room</th><th>Exams</th><th>Revenue</th></tr></thead>
      <tbody>
        {% for r in room_rows %}
          <tr><td>{{ r.room }}</td><td>{{ r.count }}</td><td>{{ "%.2f"|format(r.revenue) }}</td></tr>
        {% endfor %}
        {% if room_rows|length == 0 %}
          <tr><td colspan="3" class="muted">No data for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- 8) Outstanding Due -->
{% if view in ['all','due'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">8) Outstanding Due (Top 20)</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Date</th><th>Owner</th><th>Phone</th><th>Pet</th><th>Vet</th><th>Status</th><th>Due</th></tr></thead>
      <tbody>
        {% for r in due_rows %}
          <tr>
            <td>{{ r.date }}</td>
            <td>{{ r.owner }}</td>
            <td>{{ r.phone }}</td>
            <td>{{ r.pet }}</td>
            <td>{{ r.vet }}</td>
            <td>{{ r.status }}</td>
            <td>{{ "%.2f"|format(r.due) }}</td>
          </tr>
        {% endfor %}
        {% if due_rows|length == 0 %}
          <tr><td colspan="7" class="muted">No outstanding due for this criteria.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<!-- Futuristic Insights -->
{% if view in ['all','future'] %}
<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">Futuristic Insight #1: 7-Day Forecast</div>
    <div class="muted">A lightweight forecast based on the moving average of the most recent activity in your selected criteria.</div>
    <table class="table" style="margin-top:10px">
      <thead><tr><th>Forecast Date</th><th>Predicted Exams</th><th>Predicted Revenue</th></tr></thead>
      <tbody>
        {% for f in forecast %}
          <tr><td>{{ f.date }}</td><td>{{ f.exams }}</td><td>{{ "%.2f"|format(f.revenue) }}</td></tr>
        {% endfor %}
      </tbody>
      <tfoot>
        <tr><th>7-day Total</th><th>{{ forecast_totals.exams }}</th><th>{{ "%.2f"|format(forecast_totals.revenue) }}</th></tr>
      </tfoot>
    </table>
  </div>
</div>

<div class="grid" style="margin-top:14px">
  <div class="card">
    <div class="h">Futuristic Insight #2: Early-Warning Signals</div>
    <div class="muted">Automated week-over-week signals to catch operational or cashflow issues early.</div>
    <ul style="margin-top:10px">
      {% for a in alerts %}
        <li>{{ a }}</li>
      {% endfor %}
    </ul>
  </div>
</div>
{% endif %}

{% endblock %}

"""

# -------------------------
# Templates: Pet History
# -------------------------
TEMPLATES["pet_history.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap">
    <div>
      <h3 style="margin:0">Pet History</h3>
      <div class="small">
        <b>Owner:</b> {{ owner.owner_name }} ({{ owner.phone }}) &nbsp; | &nbsp;
        <b>Pet:</b> {{ pet.pet_name }}{% if pet.species %} ({{ pet.species }}){% endif %}
      </div>
    </div>
    <div>
      <a class="btn" href="{{ back_url or url_for('home') }}">Back</a>
    </div>
  </div>

  <div class="hr"></div>

  {% if items and items|length > 0 %}
    <table class="table">
      <thead>
        <tr>
          <th>Date/Time</th>
          <th>Status</th>
          <th>Vet</th>
          <th>Room</th>
          <th>Services</th>
          <th>Total</th>
          <th>Paid</th>
          <th>Due</th>
        </tr>
      </thead>
      <tbody>
        {% for it in items %}
          <tr>
            <td>{{ it.dt }}</td>
            <td>{{ it.status }}</td>
            <td>{{ it.vet }}</td>
            <td>{{ it.room }}</td>
            <td>
              {% if it.services and it.services|length>0 %}
                <ul style="margin:0;padding-left:18px">
                  {% for s in it.services %}
                    <li>{{ s }}</li>
                  {% endfor %}
                </ul>
              {% else %}
                <span class="small">—</span>
              {% endif %}
              {% if it.reason %}
                <div class="small" style="margin-top:6px"><b>Reason:</b> {{ it.reason }}</div>
              {% endif %}

            {% if it.weight or it.temp %}
              <div class="small" style="margin-top:6px"><b>Vitals:</b>
                {% if it.weight %}Weight {{ it.weight }} kg{% endif %}
                {% if it.weight and it.temp %} • {% endif %}
                {% if it.temp %}Temp {{ it.temp }} °C{% endif %}
              </div>
            {% endif %}

            </td>
            <td>{{ it.total }}</td>
            <td>{{ it.paid }}</td>
            <td>{{ it.due }}</td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  {% else %}
    <div class="small">No exam history found for this pet yet.</div>
  {% endif %}
</div>
{% endblock %}
"""

TEMPLATES["pet_history_partial.html"] = r"""
<div class="small" style="margin-bottom:10px">
  <b>Owner:</b> {{ owner.owner_name }} ({{ owner.phone }}) &nbsp; | &nbsp;
  <b>Pet:</b> {{ pet.pet_name }}{% if pet.species %} ({{ pet.species }}){% endif %}
</div>

{% if items and items|length > 0 %}
  <table class="table">
    <thead>
      <tr>
        <th>Date/Time</th>
        <th>Status</th>
        <th>Vet</th>
        <th>Room</th>
        <th>Services / Reason</th>
        <th>Total</th>
        <th>Paid</th>
        <th>Due</th>
      </tr>
    </thead>
    <tbody>
      {% for it in items %}
        <tr>
          <td>{{ it.dt }}</td>
          <td>{{ it.status }}</td>
          <td>{{ it.vet }}</td>
          <td>{{ it.room }}</td>
          <td>
            {% if it.services and it.services|length>0 %}
              <ul style="margin:0;padding-left:18px">
                {% for s in it.services %}
                  <li>{{ s }}</li>
                {% endfor %}
              </ul>
            {% else %}
              <span class="small">—</span>
            {% endif %}
            {% if it.reason %}
              <div class="small" style="margin-top:6px"><b>Reason:</b> {{ it.reason }}</div>

            {% if it.weight or it.temp %}
              <div class="small" style="margin-top:6px"><b>Vitals:</b>
                {% if it.weight %}Weight {{ it.weight }} kg{% endif %}
                {% if it.weight and it.temp %} • {% endif %}
                {% if it.temp %}Temp {{ it.temp }} °C{% endif %}
              </div>
            {% endif %}

{% endif %}
          </td>
          <td>{{ it.total }}</td>
          <td>{{ it.paid }}</td>
          <td>{{ it.due }}</td>
        </tr>
      {% endfor %}
    </tbody>
  </table>
{% else %}
  <div class="small">No exam history found for this pet yet.</div>
{% endif %}
"""

TEMPLATES["pet_reminders_partial.html"] = r"""
{% if show_header %}
<div class="small" style="margin-bottom:10px">
  <b>Owner:</b> {{ owner.owner_name }} ({{ owner.phone }}) &nbsp; | &nbsp;
  <b>Pet:</b> {{ pet.pet_name }}{% if pet.species %} ({{ pet.species }}){% endif %}
</div>
{% endif %}

{% if upcoming and upcoming|length>0 %}
  <div class="small" style="margin-bottom:8px"><b>Upcoming Appointments</b></div>
  <table class="table">
    <thead>
      <tr><th>Date/Time</th><th>Type</th><th>Status</th><th></th></tr>
    </thead>
    <tbody>
      {% for u in upcoming %}
        <tr>
          <td>{{ u.appointment_start }}</td>
          <td>{{ u.appointment_type }}</td>
          <td>{{ u.status }}</td>
          <td><a class="btn" style="padding:6px 8px;font-size:12px" href="{{ url_for('booking_view', booking_id=u.id) }}" target="_blank">Open</a></td>
        </tr>
      {% endfor %}
    </tbody>
  </table>
  <div class="hr"></div>
{% endif %}

<div class="row-actions" style="margin-bottom:10px;flex-wrap:wrap">
  <a class="btn" href="{{ url_for('reminders', pet_id=pet.id) }}" target="_blank">Open Reminders Center</a>
</div>

{% if reminders and reminders|length>0 %}
  <table class="table">
    <thead>
      <tr>
        <th>Scheduled</th>
        <th>Type</th>
        <th>Channel</th>
        <th>Status</th>
        <th>Message</th>
        <th></th>
      </tr>
    </thead>
    <tbody>
      {% for r in reminders %}
        <tr>
          <td>{{ r.scheduled_for or r.created_at }}</td>
          <td>{{ r.reminder_type }}</td>
          <td>{{ r.channel }}</td>
          <td>{{ r.status }}</td>
          <td class="small">{{ r.message }}</td>
          <td><a class="btn" style="padding:6px 8px;font-size:12px" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a></td>
        </tr>
      {% endfor %}
    </tbody>
  </table>
{% else %}
  <div class="small">No reminders found for this pet yet.</div>
{% endif %}
"""

TEMPLATES["pet_reminders.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap">
    <div>
      <h3 style="margin:0">Pet Reminders</h3>
      <div class="small">
        <b>Owner:</b> {{ owner.owner_name }} ({{ owner.phone }}) &nbsp; | &nbsp;
        <b>Pet:</b> {{ pet.pet_name }}{% if pet.species %} ({{ pet.species }}){% endif %}
      </div>
    </div>
    <div>
      <a class="btn" href="{{ back_url or url_for('home') }}">Back</a>
    </div>
  </div>

  <div class="hr"></div>

  {% set show_header = false %}
  {% include "pet_reminders_partial.html" %}
</div>
{% endblock %}
"""


def write_templates():
    for name, content in TEMPLATES.items():
        with open(os.path.join(TEMPLATES_DIR, name), "w", encoding="utf-8") as f:
            f.write(content)


# =========================
# INIT / SEED
# =========================
def init_storage():
    ensure_headers(OWNERS_XLSX, OWNERS_HEADERS)
    ensure_headers(PETS_XLSX, PETS_HEADERS)
    ensure_headers(BOOKINGS_XLSX, BOOKINGS_HEADERS)
    ensure_headers(REMINDERS_XLSX, REMINDERS_HEADERS)
    ensure_headers(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS)
    ensure_headers(ROLES_PERMISSIONS_XLSX, ROLES_PERMISSIONS_HEADERS)
    ensure_headers(USERS_XLSX, USERS_HEADERS)
    ensure_headers(VETS_XLSX, VETS_HEADERS)
    ensure_headers(ROOMS_XLSX, ROOMS_HEADERS)
    ensure_headers(SERVICES_XLSX, SERVICES_HEADERS)
    seed_config_defaults()


def seed_demo_data(n=10):
    owners = read_all(OWNERS_XLSX)
    pets = read_all(PETS_XLSX)
    bookings = read_all(BOOKINGS_XLSX)
    if owners or pets or bookings:
        return

    owner_names = ["Ahmed Hassan", "Mona Saad", "Karim Ali", "Sara Mostafa", "Omar Fathy", "Nour ElDin",
                   "Laila Mahmoud", "Youssef Adel", "Hany Ibrahim", "Fatma Nasser"]
    for i in range(n):
        oid = str(uuid.uuid4())
        append_row(OWNERS_XLSX, OWNERS_HEADERS, {
            "id": oid,
            "owner_name": owner_names[i % len(owner_names)],
            "phone": "+20" + "10" + str(10000000 + i),
            "email": f"client{i + 1}@mail.com",
            "address": "Cairo, Egypt",
            "preferred_contact": "WhatsApp",
            "notes": "Demo client",
            "created_at": now_str(),
            "updated_at": now_str()
        })

    owners = read_all(OWNERS_XLSX)
    species = ["Dog", "Cat", "Dog", "Cat", "Dog", "Cat", "Rabbit", "Bird", "Dog", "Cat"]
    pet_names = ["Luna", "Max", "Bella", "Charlie", "Milo", "Coco", "Rocky", "Nala", "Simba", "Oreo"]
    for i in range(n):
        pid = str(uuid.uuid4())
        append_row(PETS_XLSX, PETS_HEADERS, {
            "id": pid,
            "pet_name": pet_names[i],
            "species": species[i],
            "breed": "Mixed",
            "sex": "Male" if i % 2 == 0 else "Female",
            "dob": "2020-01-01",
            "age_years": str(2 + (i % 6)),
            "weight_kg": str(3 + i),
            "color": "Brown",
            "microchip_id": f"MC-{100000 + i}",
            "spayed_neutered": "Yes" if i % 2 == 0 else "No",
            "allergies": "",
            "chronic_conditions": "",
            "vaccinations_summary": "Rabies up to date",
            "owner_id": owners[i]["id"],
            "notes": "",
            "created_at": now_str(),
            "updated_at": now_str()
        })

    owners = read_all(OWNERS_XLSX)
    pets = read_all(PETS_XLSX)
    base = datetime.now()

    sample_msgs = [
        "vomiting and not eating",
        "diarrhea since yesterday",
        "itching and rash",
        "coughing and sneezing",
        "limping and pain",
        "vaccination booster needed",
        "urinating frequently",
        "lethargic and weak",
        "possible toxin exposure",
        "follow-up check required"
    ]

    for i in range(n):
        start = (base + timedelta(days=(i - 3))).replace(hour=11 + (i % 4), minute=0)
        dur = 30 if i % 3 else 45
        end = start + timedelta(minutes=dur)
        bid = str(uuid.uuid4())
        token = uuid.uuid4().hex
        intake = ai_extract_from_text(sample_msgs[i])
        append_row(BOOKINGS_XLSX, BOOKINGS_HEADERS, {
            "id": bid,
            "appointment_start": start.strftime("%Y-%m-%d %H:%M"),
            "duration_min": str(dur),
            "appointment_end": end.strftime("%Y-%m-%d %H:%M"),
            "owner_id": owners[i]["id"],
            "pet_id": pets[i]["id"],
            "appointment_type": intake["appointment_type"],
            "priority": intake["priority"],
            "status": "Completed" if i < 4 else "Scheduled",
            "channel": "WhatsApp",
            "reason": intake["reason"],
            "symptoms": intake["symptoms"],
            "vet_name": "Dr. Kareem",
            "room": "Room 1",
            "fee_amount": str(250 + (i * 25)),
            "payment_status": "Paid" if i < 4 else "Unpaid",
            "payment_method": "Cash",
            "invoice_no": f"INV-2025-{2000 + i}",
            "diagnosis": "",
            "treatment_plan": "",
            "prescription": "",
            "lab_tests": "",
            "vaccines_given": "",
            "followup_datetime": "",
            "reminder_channel": "WhatsApp",
            "reminder_sent": "",
            "reminder_last_opened": "",
            "portal_token": token,
            "owner_confirmed": "",
            "owner_update_message": "",
            "owner_update_datetime": "",
            "ai_last_applied_at": "",
            "notes": "",
            "created_at": now_str(),
            "updated_at": now_str()
        })


# =========================
# AUTH ROUTES
# =========================
@app.route("/")
def index():
    return redirect(url_for("home") if session.get("logged_in") else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = (request.form.get("password") or "").strip()

        auth = auth_user(u, p)
        if auth:
            session["logged_in"] = True
            session["username"] = auth["username"]
            session["role"] = auth["role"]
            return redirect(url_for("home"))

        flash("Invalid credentials.")
        return redirect(url_for("login"))
    return render_template("login.html", app_title=APP_TITLE, title=f"{APP_TITLE} | Login", logged_in=False)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/home")
def home():
    gate = require_login()
    if gate: return gate
    return render_template("home.html",
                           title=f"{APP_TITLE} | Home",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Home",
                           subtitle="Futuristic Journey: Smart Intake → WhatsApp Journey → Owner Portal → Calendar → Popups",
                           active="home"
                           )


# =========================

# =========================
# CONFIG (ADMIN ONLY)
# =========================
@app.route("/config")
def config():
    gate = require_admin()
    if gate: return gate

    users = get_users(include_inactive=True)
    vets_rows = get_vets(include_inactive=True)
    rooms_rows = get_rooms(include_inactive=True)
    services_rows = get_services(include_inactive=True)

    return render_template("config.html",
                           title=f"{APP_TITLE} | Config",
                           header="Configuration",
                           subtitle="Admin-only settings",
                           active="config",
                           users=users,
                           vets_rows=vets_rows,
                           rooms_rows=rooms_rows,
                           services=services_rows,
                           wa_templates=get_whatsapp_templates(include_inactive=True),
                           wa_scenarios=WHATSAPP_SCENARIOS,
                           wa_booking_types=(['Any'] + APPOINTMENT_TYPES),
                           roles=ROLES,
                           role_perms=get_role_permissions_map(),
                           perms_catalog=PERMISSIONS_CATALOG
                           )


@app.route("/config/users/add", methods=["POST"])
def config_add_user():
    gate = require_admin()
    if gate: return gate

    username = norm_username(request.form.get("username"))
    password = (request.form.get("password") or "").strip()
    role = (request.form.get("role") or "user").strip().lower()

    if not username or not password:
        flash("Username and password are required.")
        return redirect(url_for("config"))

    if role not in ROLES:
        flash("Invalid role selected.")
        return redirect(url_for("config"))

    if username == norm_username(ADMIN_USER):
        flash("Admin user is protected (use the main admin login).")
        return redirect(url_for("config"))

    if get_user_by_username(username):
        flash("User already exists.")
        return redirect(url_for("config"))

    append_row(USERS_XLSX, {
        "id": str(uuid.uuid4()),
        "username": username,
        "password": password,
        "role": role,
        "active": "1",
        "created_at": now_str(),
        "updated_at": now_str()
    })
    flash("User added successfully.")
    return redirect(url_for("config"))


@app.route("/config/users/<user_id>/toggle", methods=["POST"])
def config_toggle_user(user_id):
    gate = require_admin()
    if gate: return gate

    users = read_all(USERS_XLSX)
    rec = None
    for u in users:
        if str(u.get("id")) == str(user_id):
            rec = u
            break
    if not rec:
        flash("User not found.")
        return redirect(url_for("config"))

    if norm_username(rec.get("username", "")) == norm_username(ADMIN_USER):
        flash("Admin user is protected.")
        return redirect(url_for("config"))

    active = _boolish(rec.get("active", "1"))
    update_row_by_id(USERS_XLSX, user_id, {
        "active": "0" if active else "1",
        "updated_at": now_str()
    })
    flash("User updated.")
    return redirect(url_for("config"))


@app.route("/config/vets/add", methods=["POST"])
def config_add_vet():
    gate = require_admin()
    if gate: return gate

    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Vet name is required.")
        return redirect(url_for("config"))

    # prevent duplicates (case-insensitive)
    existing = [v.get("name", "").strip().lower() for v in get_vets(include_inactive=True)]
    if name.lower() in existing:
        flash("Vet already exists.")
        return redirect(url_for("config"))

    append_row(VETS_XLSX, {
        "id": str(uuid.uuid4()),
        "name": name,
        "active": "1",
        "created_at": now_str(),
        "updated_at": now_str()
    })
    flash("Vet added.")
    return redirect(url_for("config"))


@app.route("/config/vets/<vet_id>/toggle", methods=["POST"])
def config_toggle_vet(vet_id):
    gate = require_admin()
    if gate: return gate

    rows = read_all(VETS_XLSX)
    rec = None
    for r in rows:
        if str(r.get("id")) == str(vet_id):
            rec = r
            break
    if not rec:
        flash("Vet not found.")
        return redirect(url_for("config"))

    active = _boolish(rec.get("active", "1"))
    update_row_by_id(VETS_XLSX, vet_id, {
        "active": "0" if active else "1",
        "updated_at": now_str()
    })
    flash("Vet updated.")
    return redirect(url_for("config"))


@app.route("/config/rooms/add", methods=["POST"])
def config_add_room():
    gate = require_admin()
    if gate: return gate

    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Room name is required.")
        return redirect(url_for("config"))

    existing = [r.get("name", "").strip().lower() for r in get_rooms(include_inactive=True)]
    if name.lower() in existing:
        flash("Room already exists.")
        return redirect(url_for("config"))

    append_row(ROOMS_XLSX, {
        "id": str(uuid.uuid4()),
        "name": name,
        "active": "1",
        "created_at": now_str(),
        "updated_at": now_str()
    })
    flash("Room added.")
    return redirect(url_for("config"))


@app.route("/config/rooms/<room_id>/toggle", methods=["POST"])
def config_toggle_room(room_id):
    gate = require_admin()
    if gate: return gate

    rows = read_all(ROOMS_XLSX)
    rec = None
    for r in rows:
        if str(r.get("id")) == str(room_id):
            rec = r
            break
    if not rec:
        flash("Room not found.")
        return redirect(url_for("config"))


@app.route("/config/services/add", methods=["POST"])
def config_add_service():
    gate = require_admin()
    if gate: return gate
    name = (request.form.get("service_name") or "").strip()
    cost = (request.form.get("service_cost") or "").strip()
    fee = (request.form.get("service_fee") or "").strip()
    if not name:
        flash("Service name is required.")
        return redirect(url_for("config"))
    try:
        cost_f = float(cost) if cost != "" else 0.0
    except Exception:
        cost_f = 0.0
    try:
        fee_f = float(fee) if fee != "" else 0.0
    except Exception:
        fee_f = 0.0
    append_row(SERVICES_XLSX, {
        "id": str(uuid.uuid4()),
        "name": name,
        "cost": round(cost_f, 2),
        "fee": round(fee_f, 2),
        "active": "1",
        "created_at": now_str(),
        "updated_at": now_str()
    })
    flash("Service added.")
    return redirect(url_for("config"))


@app.route("/config/services/<service_id>/toggle", methods=["POST"])
def config_toggle_service(service_id):
    gate = require_admin()
    if gate: return gate
    recs = get_services(include_inactive=True)
    rec = None
    for r in recs:
        if str(r.get("id", "")) == str(service_id):
            rec = r
            break
    if not rec:
        flash("Service not found.")
        return redirect(url_for("config"))

    active = _boolish(rec.get("active", "1"))
    update_row_by_id(SERVICES_XLSX, service_id, {
        "active": "0" if active else "1",
        "updated_at": now_str()
    })
    flash("Service updated.")
    return redirect(url_for("config"))


@app.route("/config/whatsapp/add", methods=["POST"])
def config_add_whatsapp_template():
    gate = require_admin()
    if gate: return gate

    name = (request.form.get("wa_name") or "").strip()
    scenario = (request.form.get("wa_scenario") or "Appointment").strip()
    booking_type = (request.form.get("wa_booking_type") or "Any").strip() or "Any"
    template_text = (request.form.get("wa_template_text") or "").strip()
    is_default = "1" if request.form.get("wa_is_default") else "0"

    if not name or not template_text:
        flash("WhatsApp template name and text are required.")
        return redirect(url_for("config"))

    tid = str(uuid.uuid4())
    now = now_str()

    append_row(WHATSAPP_TEMPLATES_XLSX, {
        "id": tid,
        "name": name,
        "scenario": scenario,
        "booking_type": booking_type,
        "template_text": template_text,
        "active": "1",
        "is_default": is_default,
        "created_at": now,
        "updated_at": now,
    })

    if is_default == "1":
        _whatsapp_clear_other_defaults(tid, scenario, booking_type)

    flash("WhatsApp template added.")
    return redirect(url_for("config"))


@app.route("/config/whatsapp/set_default/<template_id>", methods=["POST"])
def config_set_default_whatsapp_template(template_id):
    gate = require_admin()
    if gate: return gate

    rows = read_all(WHATSAPP_TEMPLATES_XLSX)
    t = find_by_id(rows, template_id)
    if not t:
        flash("Template not found.")
        return redirect(url_for("config"))

    scenario = (t.get("scenario") or "").strip()
    booking_type = (t.get("booking_type") or "Any").strip() or "Any"

    update_row_by_id(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS, template_id,
                     {"is_default": "1", "updated_at": now_str()})
    _whatsapp_clear_other_defaults(template_id, scenario, booking_type)

    flash("Default WhatsApp template updated.")
    return redirect(url_for("config"))


@app.route("/config/whatsapp/toggle/<template_id>", methods=["POST"])
def config_toggle_whatsapp_template(template_id):
    gate = require_admin()
    if gate: return gate

    rows = read_all(WHATSAPP_TEMPLATES_XLSX)
    t = find_by_id(rows, template_id)
    if not t:
        flash("Template not found.")
        return redirect(url_for("config"))

    active = "0" if _boolish(t.get("active", "1")) else "1"
    update_row_by_id(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS, template_id,
                     {"active": active, "updated_at": now_str()})

    flash("Template updated.")
    return redirect(url_for("config"))


@app.route("/config/whatsapp/update/<template_id>", methods=["POST"])
def config_update_whatsapp_template(template_id):
    gate = require_admin()
    if gate: return gate

    name = (request.form.get("wa_name") or "").strip()
    scenario = (request.form.get("wa_scenario") or "Appointment").strip()
    booking_type = (request.form.get("wa_booking_type") or "Any").strip() or "Any"
    template_text = (request.form.get("wa_template_text") or "").strip()
    is_default = "1" if request.form.get("wa_is_default") else "0"
    active = "1" if request.form.get("wa_active") else "0"

    if not name or not template_text:
        flash("Template name and text are required.")
        return redirect(url_for("config"))

    update_row_by_id(WHATSAPP_TEMPLATES_XLSX, WHATSAPP_TEMPLATES_HEADERS, template_id, {
        "name": name,
        "scenario": scenario,
        "booking_type": booking_type,
        "template_text": template_text,
        "active": active,
        "is_default": is_default,
        "updated_at": now_str(),
    })

    if is_default == "1":
        _whatsapp_clear_other_defaults(template_id, scenario, booking_type)

    flash("Template saved.")
    return redirect(url_for("config"))


@app.route("/config/services/<service_id>/update", methods=["POST"])
def config_update_service(service_id):
    gate = require_admin()
    if gate: return gate
    name = (request.form.get("service_name") or "").strip()
    cost = (request.form.get("service_cost") or "").strip()
    fee = (request.form.get("service_fee") or "").strip()
    if not name:
        flash("Service name is required.")
        return redirect(url_for("config"))
    try:
        fee_f = float(fee) if fee != "" else 0.0
    except Exception:
        fee_f = 0.0
    update_row_by_id(SERVICES_XLSX, service_id, {
        "name": name,
        "cost": round(cost_f, 2),
        "fee": round(fee_f, 2),
        "updated_at": now_str()
    })
    flash("Service saved.")
    return redirect(url_for("config"))

    active = _boolish(rec.get("active", "1"))
    update_row_by_id(ROOMS_XLSX, room_id, {
        "active": "0" if active else "1",
        "updated_at": now_str()
    })
    flash("Room updated.")
    return redirect(url_for("config"))


# OWNERS
# =========================
@app.route("/owners")
def owners():
    gate = require_login()
    if gate: return gate
    q = (request.args.get("q") or "").strip().lower()
    rows = read_all(OWNERS_XLSX)
    if q:
        rows = [o for o in rows if
                any(q in str(o.get(k, "")).lower() for k in ["owner_name", "phone", "email", "address", "notes"])]
    rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    return render_template("owners.html",
                           title=f"{APP_TITLE} | Owners",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Owners",
                           subtitle="Client database",
                           active="owners",
                           owners=rows,
                           q=q
                           )


@app.route("/owners/new", methods=["GET", "POST"])
def owner_new():
    gate = require_login()
    if gate: return gate
    if request.method == "POST":
        oid = str(uuid.uuid4())
        append_row(OWNERS_XLSX, OWNERS_HEADERS, {
            "id": oid,
            "owner_name": (request.form.get("owner_name") or "").strip(),
            "phone": (request.form.get("phone") or "").strip(),
            "email": (request.form.get("email") or "").strip(),
            "address": (request.form.get("address") or "").strip(),
            "preferred_contact": (request.form.get("preferred_contact") or "WhatsApp").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })
        flash("Owner saved.")
        return redirect(url_for("owners"))
    return render_template("owner_form.html",
                           title=f"{APP_TITLE} | Add Owner",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Add Owner",
                           subtitle="Create new owner profile",
                           active="owners",
                           owner=None
                           )


@app.route("/owners/<owner_id>/edit", methods=["GET", "POST"])
def owner_edit(owner_id):
    gate = require_login()
    if gate: return gate
    rows = read_all(OWNERS_XLSX)
    o = find_by_id(rows, owner_id)
    if not o:
        flash("Owner not found.")
        return redirect(url_for("owners"))
    if request.method == "POST":
        update_row_by_id(OWNERS_XLSX, OWNERS_HEADERS, owner_id, {
            "owner_name": (request.form.get("owner_name") or "").strip(),
            "phone": (request.form.get("phone") or "").strip(),
            "email": (request.form.get("email") or "").strip(),
            "address": (request.form.get("address") or "").strip(),
            "preferred_contact": (request.form.get("preferred_contact") or "").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "updated_at": now_str()
        })
        flash("Owner updated.")
        return redirect(url_for("owners"))
    return render_template("owner_form.html",
                           title=f"{APP_TITLE} | Edit Owner",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Edit Owner",
                           subtitle="Update owner profile",
                           active="owners",
                           owner=o
                           )


@app.route("/owners/<owner_id>/delete")
def owner_delete(owner_id):
    gate = require_login()
    if gate: return gate
    delete_row_by_id(OWNERS_XLSX, owner_id)
    flash("Owner deleted.")
    return redirect(url_for("owners"))


# =========================
# PETS (+ Health Snapshot)
# =========================
def health_snapshot(pet: dict, bookings_rows: list) -> dict:
    """
    Lightweight 'digital twin' snapshot (0-100), not medical.
    Signals: chronic/allergies/vaccines presence + booking follow-through.
    """
    score = 85
    note = "Stable profile."

    chronic = str(pet.get("chronic_conditions", "") or "").strip()
    allergies = str(pet.get("allergies", "") or "").strip()
    vacc = str(pet.get("vaccinations_summary", "") or "").strip()

    if chronic:
        score -= 18
        note = "Chronic condition present."
    if allergies:
        score -= 8
    if not vacc:
        score -= 10
        note = "Vaccination summary missing."

    pid = str(pet.get("id", ""))
    pet_bookings = [b for b in bookings_rows if str(b.get("pet_id", "")) == pid]
    noshow = len([b for b in pet_bookings if str(b.get("status", "")) == "No-Show"])
    if noshow:
        score -= min(15, noshow * 7)

    score = max(0, min(100, score))
    bucket = "Good" if score >= 70 else "Watch" if score >= 45 else "Critical"
    return {"score": score, "bucket": bucket, "note": note}


@app.route("/pets")
def pets():
    gate = require_login()
    if gate: return gate
    q = (request.args.get("q") or "").strip().lower()
    owners_rows = read_all(OWNERS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    rows = read_all(PETS_XLSX)

    view = []
    for p in rows:
        owner = ob.get(str(p.get("owner_id", "")), {})
        snap = health_snapshot(p, bookings_rows)
        view.append({
            **p,
            "owner_name": safe_get(owner, "owner_name"),
            "health_score": snap["score"],
            "health_bucket": snap["bucket"],
            "health_note": snap["note"]
        })

    if q:
        view = [p for p in view if any(
            q in str(p.get(k, "")).lower() for k in ["pet_name", "species", "breed", "sex", "notes", "owner_name"])]

    view.sort(key=lambda x: str(x.get("pet_name", "")).lower())
    return render_template("pets.html",
                           title=f"{APP_TITLE} | Pets",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Pets",
                           subtitle="Pet profiles + health snapshot",
                           active="pets",
                           pets=view,
                           q=q
                           )


@app.route("/pets/new", methods=["GET", "POST"])
def pet_new():
    gate = require_login()
    if gate: return gate
    owners_rows = read_all(OWNERS_XLSX)
    owners_rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    if request.method == "POST":
        pid = str(uuid.uuid4())
        row = {h: "" for h in PETS_HEADERS}
        row.update({
            "id": pid,
            "pet_name": (request.form.get("pet_name") or "").strip(),
            "species": (request.form.get("species") or "").strip(),
            "breed": (request.form.get("breed") or "").strip(),
            "sex": (request.form.get("sex") or "").strip(),
            "dob": (request.form.get("dob") or "").strip(),
            "age_years": (request.form.get("age_years") or "").strip(),
            "weight_kg": (request.form.get("weight_kg") or "").strip(),
            "allergies": (request.form.get("allergies") or "").strip(),
            "chronic_conditions": (request.form.get("chronic_conditions") or "").strip(),
            "vaccinations_summary": (request.form.get("vaccinations_summary") or "").strip(),
            "owner_id": (request.form.get("owner_id") or "").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })
        append_row(PETS_XLSX, PETS_HEADERS, row)
        flash("Pet saved.")
        return redirect(url_for("pets"))
    return render_template("pet_form.html",
                           title=f"{APP_TITLE} | Add Pet",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Add Pet",
                           subtitle="Create new pet profile",
                           active="pets",
                           pet=None,
                           owners=owners_rows
                           )


@app.route("/pets/<pet_id>/edit", methods=["GET", "POST"])
def pet_edit(pet_id):
    gate = require_login()
    if gate: return gate
    owners_rows = read_all(OWNERS_XLSX)
    owners_rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    pets_rows = read_all(PETS_XLSX)
    p = find_by_id(pets_rows, pet_id)
    if not p:
        flash("Pet not found.")
        return redirect(url_for("pets"))
    if request.method == "POST":
        update_row_by_id(PETS_XLSX, PETS_HEADERS, pet_id, {
            "pet_name": (request.form.get("pet_name") or "").strip(),
            "species": (request.form.get("species") or "").strip(),
            "breed": (request.form.get("breed") or "").strip(),
            "sex": (request.form.get("sex") or "").strip(),
            "dob": (request.form.get("dob") or "").strip(),
            "age_years": (request.form.get("age_years") or "").strip(),
            "weight_kg": (request.form.get("weight_kg") or "").strip(),
            "allergies": (request.form.get("allergies") or "").strip(),
            "chronic_conditions": (request.form.get("chronic_conditions") or "").strip(),
            "vaccinations_summary": (request.form.get("vaccinations_summary") or "").strip(),
            "owner_id": (request.form.get("owner_id") or "").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "updated_at": now_str()
        })
        flash("Pet updated.")
        return redirect(url_for("pets"))
    return render_template("pet_form.html",
                           title=f"{APP_TITLE} | Edit Pet",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Edit Pet",
                           subtitle="Update pet profile",
                           active="pets",
                           pet=p,
                           owners=owners_rows
                           )


@app.route("/pets/<pet_id>/delete")
def pet_delete(pet_id):
    gate = require_login()
    if gate: return gate
    delete_row_by_id(PETS_XLSX, pet_id)
    flash("Pet deleted.")
    return redirect(url_for("pets"))


# =========================
# QUICK 1-CLICK BOOKING STATUS (5 statuses)
# =========================
QUICK_STATUSES = ["Scheduled", "Checked-in", "In Treatment", "Completed", "Cancelled"]

STATUS_BTN_CLASS = {
    "Scheduled": "warn",
    "Checked-in": "primary",
    "In Treatment": "warn",
    "Completed": "good",
    "Cancelled": "bad",
}


@app.context_processor
def inject_quick_status_helpers():
    # available in all templates
    try:
        vets_list = [v.get("name") for v in get_vets(include_inactive=False)]
        rooms_list = [r.get("name") for r in get_rooms(include_inactive=False)]
    except Exception:
        # fail-safe if files not ready yet
        vets_list = ["ahmed", "zaineb", "hatem", "hayaa"]
        rooms_list = ["Room 1", "Room 2", "Room 3", "Room 4"]

    return {
        "quick_statuses": QUICK_STATUSES,
        "status_btn_class": STATUS_BTN_CLASS,
        "app_title": APP_TITLE,
        "logged_in": bool(session.get("logged_in")),
        "is_admin": (session.get("role") == "admin"),
        "vets": vets_list,
        "rooms": rooms_list
    }


@app.route("/bookings/<booking_id>/status/<path:new_status>")
def booking_set_status(booking_id, new_status):
    gate = require_login()
    if gate: return gate

    new_status = (new_status or "").strip()
    if new_status not in QUICK_STATUSES:
        flash("Invalid status.")
        return redirect(request.referrer or url_for("booking_view", booking_id=booking_id))

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("bookings"))

    ok = update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
        "status": new_status,
        "updated_at": now_str()
    })

    flash(f"Status updated to: {new_status}" if ok else "Failed to update status.")
    return redirect(request.referrer or url_for("booking_view", booking_id=booking_id))


# =========================
# BOOKINGS
# =========================
@app.route("/bookings")
def bookings():
    gate = require_login()
    if gate: return gate

    q = (request.args.get("q") or "").strip().lower()
    status = (request.args.get("status") or "").strip()
    atype = (request.args.get("atype") or "").strip()

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    rows = read_all(BOOKINGS_XLSX)
    view = [decorate_booking(b, ob, pb) for b in rows]

    if q:
        view = [x for x in view if q in (" ".join([
            str(x.get("owner_name", "")), str(x.get("pet_name", "")),
            str(x.get("appointment_type", "")), str(x.get("status", "")),
            str(x.get("reason", "")), str(x.get("symptoms", ""))
        ])).lower()]
    if status:
        view = [x for x in view if str(x.get("status", "")) == status]
    if atype:
        view = [x for x in view if str(x.get("appointment_type", "")) == atype]

    view.sort(key=lambda x: parse_dt(str(x.get("appointment_start", ""))) or datetime.min, reverse=True)

    return render_template("bookings.html",
                           title=f"{APP_TITLE} | Exams",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Bookings",
                           subtitle="Seamless Journey (WA + Portal + Calendar + Popups)",
                           active="bookings",
                           bookings=view,
                           q=q, status=status, atype=atype,
                           statuses=STATUS_FLOW,
                           types=APPOINTMENT_TYPES
                           )


@app.route("/bookings/new", methods=["GET", "POST"])
def booking_new():
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    owners_rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    pets_rows = read_all(PETS_XLSX)
    pets_rows.sort(key=lambda x: str(x.get("pet_name", "")).lower())

    if request.method == "POST":
        bid = str(uuid.uuid4())
        start = normalize_dt(request.form.get("appointment_start") or "")
        dur = str((request.form.get("duration_min") or "30").strip())
        start_dt = parse_dt(start) or datetime.now()
        if not start:
            start = start_dt.strftime("%Y-%m-%d %H:%M")
        end_dt = start_dt + timedelta(minutes=int(float(dur)))
        token = uuid.uuid4().hex

        row = {h: "" for h in BOOKINGS_HEADERS}
        row.update({
            "id": bid,
            "appointment_start": start,
            "duration_min": dur,
            "appointment_end": end_dt.strftime("%Y-%m-%d %H:%M"),
            "owner_id": (request.form.get("owner_id") or "").strip(),
            "pet_id": (request.form.get("pet_id") or "").strip(),
            "visit_weight_kg": (request.form.get("visit_weight_kg") or "").strip(),
            "visit_temp_c": (request.form.get("visit_temp_c") or "").strip(),
            "appointment_type": (request.form.get("appointment_type") or "Consultation").strip(),
            "priority": (request.form.get("priority") or "Normal").strip(),
            "status": (request.form.get("status") or "Scheduled").strip(),
            "channel": (request.form.get("channel") or "Walk-in").strip(),
            "reason": (request.form.get("reason") or "").strip(),
            "symptoms": (request.form.get("symptoms") or "").strip(),
            "vet_name": (request.form.get("vet_name") or "").strip(),
            "room": (request.form.get("room") or "").strip(),
            # Services (JSON list)
            "services_json": (request.form.get("services_json") or "").strip(),
            "service_name": (request.form.get("service_name") or "").strip(),
            "service_fee": "",  # calculated
            "paid_amount": (request.form.get("paid_amount") or "").strip(),
            "due_amount": "",  # calculated
            "fee_amount": "",  # calculated

            "payment_status": "",
            "invoice_no": "",
            "followup_datetime": "",
            "reminder_channel": (request.form.get("reminder_channel") or "WhatsApp").strip(),
            "reminder_sent": "",
            "reminder_last_opened": "",
            "portal_token": token,
            "owner_confirmed": "",
            "owner_update_message": "",
            "owner_update_datetime": "",
            "ai_last_applied_at": "",
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str(),
        })

        # Payment channel required (Cash / Visa / Instapay)
        payment_channel = normalize_payment_channel(
            request.form.get("payment_channel") or request.form.get("payment_method") or ""
        )
        if not payment_channel:
            flash("Payment channel is required (Cash / Visa / Instapay).")
            return redirect(url_for("booking_new"))
        row["payment_channel"] = payment_channel
        row["payment_method"] = payment_channel

        # --- Services/Fee calculation (server-side safety) ---
        raw_services = (row.get("services_json") or "").strip()
        if not raw_services:
            row["services_json"] = "[]"
        else:
            try:
                data = json.loads(raw_services)
                if not isinstance(data, list):
                    row["services_json"] = "[]"
            except Exception:
                row["services_json"] = "[]"

        services_norm = parse_services_json(row["services_json"])
        svc_fee = services_subtotal(services_norm)

        # If service_name empty, build a compact summary
        if (not (row.get("service_name") or "").strip()) and services_norm:
            first = str((services_norm[0] or {}).get("name", "") or "").strip()
            if first:
                row["service_name"] = first if len(services_norm) == 1 else f"{first} +{len(services_norm) - 1}"

        paid = _safe_money(row.get("paid_amount", 0))

        discount = validated_discount(svc_fee, request.form.get("discount"))
        net_fee = round(svc_fee - discount, 2)

        vat_calc = round(net_fee * float(VAT_RATE), 2)
        total_calc = round(net_fee + vat_calc, 2)
        due_calc2 = round(total_calc - paid, 2)
        if due_calc2 < 0:
            due_calc2 = 0.0

        row["service_fee"] = f"{svc_fee:.2f}" if (svc_fee or services_norm) else ""
        row["discount"] = f"{discount:.2f}" if discount else ""
        row["fee_amount"] = f"{net_fee:.2f}" if (svc_fee or services_norm or discount) else ""
        row["paid_amount"] = f"{paid:.2f}" if paid else ""
        row["due_amount"] = f"{due_calc2:.2f}" if (svc_fee or paid or discount) else ""

        if total_calc <= 0:
            row["payment_status"] = ""
        elif paid <= 0:
            row["payment_status"] = "Unpaid"
        elif paid + 0.0001 >= total_calc:
            row["payment_status"] = "Paid"
        else:
            row["payment_status"] = "Partial"

        append_row(BOOKINGS_XLSX, BOOKINGS_HEADERS, row)
        flash("Exam created.")
        return redirect(url_for("bookings"))

    return render_template("booking_form.html",
                           title=f"{APP_TITLE} | New Exam",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="New Exam",
                           subtitle="Smart Intake + Full journey automation",
                           active="bookings",
                           booking=None,
                           default_start=datetime.now().strftime("%Y-%m-%d %H:%M"),
                           owners=owners_rows,
                           pets=pets_rows,
                           statuses=STATUS_FLOW,
                           types=APPOINTMENT_TYPES,
                           priorities=PRIORITIES,
                           channels=CHANNELS,
                           payment_statuses=PAYMENT_STATUSES,
                           payment_methods=PAYMENT_METHODS,
                           reminder_channels=REMINDER_CHANNELS,
                           default_vet=session.get("username", ""),
                           vets=active_vet_names(),
                           rooms=active_room_names(),
                           services=active_services(),
                           vat_rate=float(VAT_RATE),
                           easy_mode=session.get("easy_mode", "")
                           )


@app.route("/bookings/<booking_id>")
@app.route("/bookings/<booking_id>")
def booking_view(booking_id):
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("bookings"))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}
    b = decorate_booking(b, ob, pb)

    return render_template("booking_view.html",
                           title=f"{APP_TITLE} | Booking",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Booking",
                           subtitle="AI Copilot + Owner Portal + WhatsApp Journey",
                           active="bookings",
                           b=b
                           )


@app.route("/bookings/<booking_id>/edit", methods=["GET", "POST"])
def booking_edit(booking_id):
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("bookings"))

    owners_rows = read_all(OWNERS_XLSX)
    owners_rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    pets_rows = read_all(PETS_XLSX)
    pets_rows.sort(key=lambda x: str(x.get("pet_name", "")).lower())

    if request.method == "POST":
        start = normalize_dt(request.form.get("appointment_start") or "")
        dur = str((request.form.get("duration_min") or "30").strip())
        start_dt = parse_dt(start) or datetime.now()
        if not start:
            start = start_dt.strftime("%Y-%m-%d %H:%M")
        end_dt = start_dt + timedelta(minutes=int(float(dur)))

        # Payment channel required (Cash / Visa / Instapay)
        payment_channel = normalize_payment_channel(
            request.form.get("payment_channel") or request.form.get("payment_method") or ""
        )
        if not payment_channel:
            flash("Payment channel is required (Cash / Visa / Instapay).")
            return redirect(url_for("booking_edit", booking_id=booking_id))

        # --- Services + totals (server-side) ---
        raw_services = (request.form.get("services_json") or "").strip()
        if not raw_services:
            raw_services = "[]"
        else:
            try:
                data = json.loads(raw_services)
                if not isinstance(data, list):
                    raw_services = "[]"
            except Exception:
                raw_services = "[]"

        services_norm = parse_services_json(raw_services)
        svc_fee = services_subtotal(services_norm)

        svc_name = (request.form.get("service_name") or "").strip()
        if (not svc_name) and services_norm:
            first = str((services_norm[0] or {}).get("name", "") or "").strip()
            if first:
                svc_name = first if len(services_norm) == 1 else f"{first} +{len(services_norm) - 1}"

        paid = round(to_float(request.form.get("paid_amount"), 0.0), 2)
        discount = validated_discount(svc_fee, request.form.get("discount"))
        net_fee = round(svc_fee - discount, 2)

        vat_calc = round(net_fee * float(VAT_RATE), 2)
        total_calc = round(net_fee + vat_calc, 2)
        due_calc2 = round(total_calc - paid, 2)
        if due_calc2 < 0:
            due_calc2 = 0.0

        if total_calc <= 0:
            pay_status = ""
        elif paid <= 0:
            pay_status = "Unpaid"
        elif paid + 0.0001 >= total_calc:
            pay_status = "Paid"
        else:
            pay_status = "Partial"

        update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
            "appointment_start": start,
            "duration_min": dur,
            "appointment_end": end_dt.strftime("%Y-%m-%d %H:%M"),
            "owner_id": (request.form.get("owner_id") or "").strip(),
            "pet_id": (request.form.get("pet_id") or "").strip(),
            "visit_weight_kg": (request.form.get("visit_weight_kg") or "").strip(),
            "visit_temp_c": (request.form.get("visit_temp_c") or "").strip(),
            "appointment_type": (request.form.get("appointment_type") or "").strip(),
            "priority": (request.form.get("priority") or "").strip(),
            "status": (request.form.get("status") or "").strip(),
            "channel": (request.form.get("channel") or "").strip(),
            "reason": (request.form.get("reason") or "").strip(),
            "symptoms": (request.form.get("symptoms") or "").strip(),
            "vet_name": (request.form.get("vet_name") or "").strip(),
            "room": (request.form.get("room") or "").strip(),
            "services_json": raw_services,
            "service_name": svc_name,
            "service_fee": f"{svc_fee:.2f}" if (svc_fee or services_norm) else "",
            "discount": f"{discount:.2f}" if discount else "",
            "paid_amount": f"{paid:.2f}" if paid else "",
            "due_amount": f"{due_calc2:.2f}" if (svc_fee or paid or discount) else "",
            "fee_amount": f"{net_fee:.2f}" if (svc_fee or services_norm or discount) else "",
            "payment_status": pay_status,
            "payment_method": payment_channel,
            "payment_channel": payment_channel,
            "reminder_channel": (request.form.get("reminder_channel") or "").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "updated_at": now_str()
        })
        flash("Exam updated.")
        return redirect(url_for("booking_view", booking_id=booking_id))

    return render_template("booking_form.html",
                           title=f"{APP_TITLE} | Edit Exam",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Edit Exam",
                           subtitle="Update appointment + keep portal & AI journey",
                           active="bookings",
                           booking=b,
                           default_start=datetime.now().strftime("%Y-%m-%d %H:%M"),
                           owners=owners_rows,
                           pets=pets_rows,
                           statuses=STATUS_FLOW,
                           types=APPOINTMENT_TYPES,
                           priorities=PRIORITIES,
                           channels=CHANNELS,
                           payment_statuses=PAYMENT_STATUSES,
                           payment_methods=PAYMENT_METHODS,
                           reminder_channels=REMINDER_CHANNELS,
                           default_vet=session.get("username", ""),
                           vets=active_vet_names(),
                           rooms=active_room_names(),
                           services=active_services(),
                           vat_rate=float(VAT_RATE)
                           )


@app.route("/bookings/<booking_id>/delete")
def booking_delete(booking_id):
    gate = require_login()
    if gate: return gate
    delete_row_by_id(BOOKINGS_XLSX, booking_id)
    flash("Exam deleted.")
    return redirect(url_for("bookings"))


@app.route("/bookings/<booking_id>/remind")
def booking_remind(booking_id):
    """
    One-click Journey:
    - Ensures portal token exists
    - Opens WhatsApp with portal link
    - Logs reminder_last_opened
    - Creates reminder record (Opened)
    """
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("bookings"))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}
    db = decorate_booking(b, ob, pb)

    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
        "reminder_last_opened": now_str(),
        "updated_at": now_str()
    })

    rid = str(uuid.uuid4())
    append_row(REMINDERS_XLSX, REMINDERS_HEADERS, {
        "id": rid,
        "booking_id": booking_id,
        "owner_id": str(b.get("owner_id", "")),
        "pet_id": str(b.get("pet_id", "")),
        "reminder_type": "Appointment",
        "service_name": "",
        "channel": "WhatsApp",
        "status": "Opened",
        "scheduled_for": normalize_dt(str(b.get("appointment_start", ""))),
        "opened_at": now_str(),
        "sent_at": "",
        "message": booking_message_template(db.get("owner_name", ""), db.get("pet_name", ""),
                                            db.get("appointment_start", ""), db.get("portal_link", "")),
        "created_at": now_str(),
        "updated_at": now_str()
    })

    return redirect(db["whatsapp_link"])


@app.route("/bookings/<booking_id>/apply_ai")
def booking_apply_ai(booking_id):
    """
    Applies AI suggestions to booking:
    - priority, type, duration (optional)
    - appends AI summary into notes
    """
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("bookings"))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    owner = ob.get(str(b.get("owner_id", "")), {})
    pet = pb.get(str(b.get("pet_id", "")), {})

    out = ai_copilot(pet, owner, b)

    # Apply suggestions (non-destructive)
    new_notes = (str(b.get("notes", "") or "") + "\n\n" +
                 f"[AI Applied @ {now_str()}]\n" + out.get("plan_text", "") + "\n").strip()

    updates = {
        "priority": out.get("suggested_priority", b.get("priority", "")),
        "appointment_type": out.get("suggested_type", b.get("appointment_type", "")),
        "duration_min": str(out.get("suggested_duration_min", b.get("duration_min", "30"))),
        "ai_last_applied_at": now_str(),
        "notes": new_notes,
        "updated_at": now_str()
    }
    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, updates)
    flash("AI suggestions applied (priority/type/duration + notes).")
    return redirect(url_for("booking_view", booking_id=booking_id))


# =========================
# OWNER PORTAL ROUTE
# =========================
@app.route("/portal/<token>", methods=["GET", "POST"])
def portal(token):
    b = find_booking_by_token(token)
    if not b:
        return render_template("portal.html", app_title=APP_TITLE, msg="Invalid or expired link.",
                               pet_name="—", appointment_start="—", wa_link="#")

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    owner = ob.get(str(b.get("owner_id", "")), {})
    pet = pb.get(str(b.get("pet_id", "")), {})

    pet_name = safe_get(pet, "pet_name")
    appt = normalize_dt(str(b.get("appointment_start", "")))
    owner_name = safe_get(owner, "owner_name")

    msg = ""
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        contact = (request.form.get("contact") or "WhatsApp").strip()
        message = (request.form.get("message") or "").strip()

        # update booking fields
        confirmed = str(b.get("owner_confirmed", "") or "").strip()
        if action == "confirm":
            confirmed = "Yes"
            msg = "Confirmed. Thank you."
        elif action == "reschedule":
            confirmed = "Reschedule Requested"
            msg = "Reschedule request submitted. We will contact you."
        else:
            msg = "Update submitted. Thank you."

        update_text = f"[{action.upper()} via {contact}] {message}".strip()
        update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, str(b.get("id", "")), {
            "owner_confirmed": confirmed,
            "owner_update_message": update_text,
            "owner_update_datetime": now_str(),
            "updated_at": now_str()
        })

    portal_link = url_for("portal", token=token, _external=True)
    svc_items = parse_services_json(b.get("services_json", ""))
    svc_name = str(b.get("service_name", "") or "").strip()
    if not svc_name and svc_items:
        first = str((svc_items[0] or {}).get("name", "") or "").strip()
        if first:
            svc_name = first if len(svc_items) == 1 else f"{first} +{len(svc_items) - 1}"
    details = format_booking_details(b, services=svc_items)

    wa_msg = booking_message_template(
        owner_name,
        pet_name,
        appt,
        portal_link=portal_link,
        service_name=svc_name,
        booking_details=details,
        booking_type=str(b.get("appointment_type", "Any") or "Any"),
    )

    wa_link = whatsapp_link(safe_get(owner, "phone"), wa_msg)

    return render_template("portal.html",
                           app_title=APP_TITLE,
                           msg=msg,
                           pet_name=pet_name,
                           appointment_start=appt,
                           wa_link=wa_link
                           )


# =========================
# REMINDERS
# =========================
@app.route("/reminders")
def reminders():
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    rem_rows = read_all(REMINDERS_XLSX)

    ob = {str(o.get("id", "")): o for o in owners_rows}
    pb = {str(p.get("id", "")): p for p in pets_rows}

    # Global: upcoming appointments (next 24h)
    now_dt = datetime.now()
    cutoff = now_dt + timedelta(hours=24)
    upcoming = []
    for b in bookings_rows:
        dt = parse_dt(str(b.get("appointment_start", "")))
        if dt and now_dt <= dt <= cutoff and str(b.get("status", "")) not in ("Cancelled", "No-Show"):
            upcoming.append(decorate_booking(b, ob, pb))
    upcoming.sort(key=lambda x: parse_dt(str(x.get("appointment_start", ""))) or datetime.max)

    # Global: reminder records
    reminder_rows = []
    for r in rem_rows:
        pet = pb.get(str(r.get("pet_id", "")), {})
        owner = ob.get(str(r.get("owner_id", "")), {})
        reminder_rows.append({
            **r,
            "pet_name": safe_get(pet, "pet_name"),
            "whatsapp_link": whatsapp_link(owner.get("phone", ""), str(r.get("message", "")))
        })
    reminder_rows.sort(key=lambda x: (str(x.get("scheduled_for", "")) or str(x.get("created_at", ""))), reverse=True)

    # Pet 360 (search + dropdown)
    q = (request.args.get("q") or "").strip()
    crit = (request.args.get("crit") or "any").strip().lower()
    selected_pet_id = (request.args.get("pet_id") or "").strip()

    def digits(s: str) -> str:
        return re.sub(r"\D+", "", s or "")

    pet_matches = []
    if q:
        q_l = q.lower()
        q_digits = digits(q)
        seen = set()

        for p in pets_rows:
            pid = str(p.get("id", "")).strip()
            if not pid:
                continue

            pet_name = str(p.get("pet_name") or p.get("name") or "").strip()
            species = str(p.get("species") or "").strip()
            breed = str(p.get("breed") or "").strip()
            microchip_id = str(p.get("microchip_id") or "").strip()

            owner = ob.get(str(p.get("owner_id", "")).strip(), {})
            owner_name = str(owner.get("owner_name") or owner.get("name") or "").strip()
            owner_phone = str(owner.get("phone") or owner.get("mobile") or "").strip()
            owner_email = str(owner.get("email") or "").strip()

            hay_any = f"{pet_name} {species} {breed} {microchip_id} {pid} {owner_name} {owner_email}".lower()
            phone_digits = digits(owner_phone)

            ok = False

            # Criteria-based match
            if crit in ("any", "all", "", None):
                if q_l and q_l in hay_any:
                    ok = True
                if (not ok) and q_digits and phone_digits and q_digits in phone_digits:
                    ok = True

            elif crit in ("owner", "owner_name", "name"):
                if q_l and q_l in (owner_name or "").lower():
                    ok = True

            elif crit in ("phone", "owner_phone", "mobile"):
                if q_digits and phone_digits and q_digits in phone_digits:
                    ok = True

            elif crit in ("pet", "pet_name"):
                hay_pet = f"{pet_name} {species} {breed} {microchip_id}".lower()
                if q_l and q_l in hay_pet:
                    ok = True

            elif crit in ("pet_id", "id"):
                hay_id = f"{pid} {microchip_id}".lower()
                if q_l and q_l in hay_id:
                    ok = True

            else:
                # Unknown criteria → fallback to ANY
                if q_l and q_l in hay_any:
                    ok = True
                if (not ok) and q_digits and phone_digits and q_digits in phone_digits:
                    ok = True

            if ok and pid not in seen:
                seen.add(pid)
                label = f"{pet_name or 'Unnamed Pet'} — {owner_name or 'Unknown Owner'} ({owner_phone or 'No Phone'}) [Pet ID: {pid}]"
                pet_matches.append({
                    "pet_id": pid,
                    "pet_name": pet_name,
                    "owner_name": owner_name,
                    "owner_phone": owner_phone,
                    "label": label
                })

        pet_matches.sort(key=lambda x: (x.get("pet_name") or "").lower())

    pet360 = None
    pet_fin = {"total": "0.00", "paid": "0.00", "due": "0.00", "bookings_count": 0, "open_count": 0, "open_items": []}
    pet_history = []
    pet_reminders_all = []
    pet_reminders_upcoming = []

    if selected_pet_id:
        pet = pb.get(selected_pet_id) or find_by_id(pets_rows, selected_pet_id) or {}
        owner = ob.get(str(pet.get("owner_id", "")).strip(), {}) if pet else {}

        pet360 = {
            "pet_id": selected_pet_id,
            "pet_name": str(pet.get("pet_name") or "").strip(),
            "species": str(pet.get("species") or "").strip(),
            "breed": str(pet.get("breed") or "").strip(),
            "sex": str(pet.get("sex") or "").strip(),
            "age": str(pet.get("age_years") or pet.get("dob") or "").strip(),
            "weight": str(pet.get("weight_kg") or "").strip(),
            "owner_name": str(owner.get("owner_name") or "").strip(),
            "owner_phone": str(owner.get("phone") or "").strip(),
            "owner_email": str(owner.get("email") or "").strip(),
            "owner_address": str(owner.get("address") or "").strip(),
            "microchip_id": str(pet.get("microchip_id") or "").strip(),
            "allergies": str(pet.get("allergies") or "").strip(),
            "chronic_conditions": str(pet.get("chronic_conditions") or "").strip(),
            "vaccinations_summary": str(pet.get("vaccinations_summary") or pet.get("vaccination_status") or "").strip(),
            "pet_notes": str(pet.get("notes") or pet.get("medical_notes") or "").strip(),
        }

        # Bookings history + financial totals
        pet_bookings = [b for b in bookings_rows if str(b.get("pet_id", "")).strip() == selected_pet_id]

        def b_dt(b):
            return parse_dt(str(b.get("appointment_start", ""))) or datetime.min

        pet_bookings.sort(key=b_dt, reverse=True)

        total_sum = 0.0
        paid_sum = 0.0
        due_sum = 0.0
        open_items = []

        for b in pet_bookings:
            services = parse_services_json(b.get("services_json") or "[]")
            subtotal = services_subtotal(services)
            vat = round(subtotal * VAT_RATE, 2)
            total = round(subtotal + vat, 2)

            paid = round(to_float(b.get("paid_amount"), 0.0), 2)
            due = round(to_float(b.get("due_amount"), max(total - paid, 0.0)), 2)
            if due < 0:
                due = 0.0

            total_sum += total
            paid_sum += paid
            due_sum += due

            services_summary = ", ".join([f"{s.get('name')} x{s.get('qty', 1)}" for s in services]) if services else ""

            pet_history.append({
                "booking_id": str(b.get("id", "")).strip(),
                "appointment_start": str(b.get("appointment_start") or "").strip(),
                "appointment_type": str(b.get("appointment_type") or "").strip(),
                "status": str(b.get("status") or "").strip(),
                "vet_name": str(b.get("vet_name") or "").strip(),
                "room": str(b.get("room") or "").strip(),
                "services_summary": services_summary,
                "reason": str(b.get("reason") or "").strip(),
                "symptoms": str(b.get("symptoms") or "").strip(),
                "diagnosis": str(b.get("diagnosis") or "").strip(),
                "treatment_plan": str(b.get("treatment_plan") or "").strip(),
                "prescription": str(b.get("prescription") or "").strip(),
                "lab_tests": str(b.get("lab_tests") or "").strip(),
                "vaccines_given": str(b.get("vaccines_given") or "").strip(),
                "followup_datetime": str(b.get("followup_datetime") or "").strip(),
                "notes": str(b.get("notes") or "").strip(),
                "invoice_no": str(b.get("invoice_no") or "").strip(),
                "payment_status": str(b.get("payment_status") or "").strip(),
                "paid": f"{paid:.2f}",
                "total": f"{total:.2f}",
                "due": f"{due:.2f}",
            })

            if due > 0.01:
                open_items.append({
                    "booking_id": str(b.get("id", "")).strip(),
                    "appointment_start": str(b.get("appointment_start") or "").strip(),
                    "invoice_no": str(b.get("invoice_no") or "").strip(),
                    "payment_status": str(b.get("payment_status") or "").strip(),
                    "due": f"{due:.2f}",
                })

        pet_fin = {
            "total": f"{total_sum:.2f}",
            "paid": f"{paid_sum:.2f}",
            "due": f"{due_sum:.2f}",
            "bookings_count": len(pet_bookings),
            "open_count": len(open_items),
            "open_items": open_items[:25],
        }

        # Derived medical summary (latest visit snapshot)
        if pet_history:
            latest = pet_history[0]
            pet360["last_visit"] = latest.get("appointment_start", "")
            pet360["last_diagnosis"] = latest.get("diagnosis", "")
            pet360["last_treatment_plan"] = latest.get("treatment_plan", "")
            pet360["last_vaccines"] = latest.get("vaccines_given", "")
            pet360["last_followup"] = latest.get("followup_datetime", "")

        # Reminders for this pet
        for r in rem_rows:
            if str(r.get("pet_id", "")).strip() != selected_pet_id:
                continue
            pet_obj = pb.get(str(r.get("pet_id", "")), {})
            owner_obj = ob.get(str(r.get("owner_id", "")), {})
            pet_reminders_all.append({
                **r,
                "pet_name": safe_get(pet_obj, "pet_name"),
                "whatsapp_link": whatsapp_link(owner_obj.get("phone", ""), str(r.get("message", "")))
            })

        def r_dt(r):
            return parse_dt(str(r.get("scheduled_for", ""))) or datetime.min

        pet_reminders_all.sort(key=r_dt, reverse=True)

        for rr in pet_reminders_all:
            dt = parse_dt(str(rr.get("scheduled_for", "")))
            if dt and dt >= now_dt:
                pet_reminders_upcoming.append(rr)

        pet_reminders_upcoming.sort(key=lambda x: parse_dt(str(x.get("scheduled_for", ""))) or datetime.max)
        pet_reminders_upcoming = pet_reminders_upcoming[:50]

    return render_template(
        "reminders.html",
        title=f"{APP_TITLE} | Reminders",
        app_title=APP_TITLE,
        logged_in=True,
        header="Reminders Center",
        subtitle="Upcoming + Reminder records + Portal journey",
        active="reminders",
        upcoming=upcoming,
        reminder_rows=reminder_rows,
        q=q,
        crit=crit,
        pet_matches=pet_matches,
        selected_pet_id=selected_pet_id,
        pet360=pet360,
        pet_fin=pet_fin,
        pet_history=pet_history,
        pet_reminders_all=pet_reminders_all,
        pet_reminders_upcoming=pet_reminders_upcoming,
    )


@app.route("/reminders/<reminder_id>/mark_sent")
def reminder_mark_sent(reminder_id):
    gate = require_login()
    if gate: return gate
    ok = update_row_by_id(REMINDERS_XLSX, REMINDERS_HEADERS, reminder_id, {
        "status": "Sent",
        "sent_at": now_str(),
        "updated_at": now_str()
    })
    flash("Reminder marked as sent." if ok else "Reminder not found.")
    return redirect(url_for("reminders"))


# =========================
# CALENDAR (.ics)
# =========================
@app.route("/calendar/<booking_id>.ics")
def calendar_ics(booking_id):
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        return Response("Exam not found.", status=404)

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}
    db = decorate_booking(b, ob, pb)

    start_dt = parse_dt(str(db.get("appointment_start", ""))) or datetime.now()
    dur = int(float(db.get("duration_min") or 30))
    end_dt = start_dt + timedelta(minutes=dur)

    summary = f"Vet Appointment - {db.get('pet_name', '')}"
    desc = booking_message_template(db.get("owner_name", ""), db.get("pet_name", ""), db.get("appointment_start", ""),
                                    db.get("portal_link", ""))
    loc = safe_get(ob.get(str(b.get("owner_id", "")), {}), "address")

    ics = ics_content(summary, start_dt, end_dt, desc, location=loc)
    filename = f"appointment_{booking_id}.ics"
    return Response(ics, mimetype="text/calendar",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                    )


# =========================
# API: UPCOMING (Popups)
# =========================
@app.route("/api/upcoming")
def api_upcoming():
    gate = require_login()
    if gate: return gate

    minutes = int(request.args.get("minutes") or 30)
    now = datetime.now()
    cutoff = now + timedelta(minutes=minutes)

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    items = []
    for b in bookings_rows:
        dt = parse_dt(str(b.get("appointment_start", "")))
        if not dt:
            continue
        if now <= dt <= cutoff and str(b.get("status", "")) not in ("Completed", "Cancelled", "No-Show"):
            db = decorate_booking(b, ob, pb)
            ns = no_show_risk(str(b.get("owner_id", "")), str(b.get("appointment_start", "")))
            items.append({
                "id": db["id"],
                "appointment_start": db.get("appointment_start", ""),
                "owner_name": db.get("owner_name", ""),
                "pet_name": db.get("pet_name", ""),
                "whatsapp_link": url_for("booking_remind", booking_id=db["id"]),
                "ics_link": db.get("ics_link", ""),
                "portal_link": db.get("portal_link", ""),
                "page_link": url_for("booking_view", booking_id=db["id"]),
                "no_show_score": ns["score"],
                "no_show_bucket": ns["bucket"]
            })

    items.sort(key=lambda x: parse_dt(x.get("appointment_start", "")) or datetime.max)
    return jsonify({"items": items})


# =========================
# API: AI INTAKE
# =========================
@app.route("/api/intake", methods=["POST"])
def api_intake():
    gate = require_login()
    if gate: return gate
    try:
        payload = request.get_json(force=True)
        text = (payload.get("text") or "").strip()
        if not text:
            return jsonify({"error": "No text provided."})
        return jsonify(ai_extract_from_text(text))
    except Exception as e:
        return jsonify({"error": f"Failed: {e}"}), 400


# =========================
# API: AI COPILOT
# =========================
@app.route("/api/copilot/<booking_id>")
def api_copilot(booking_id):
    gate = require_login()
    if gate: return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        return jsonify({"error": "Exam not found."})

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    owner = ob.get(str(b.get("owner_id", "")), {})
    pet = pb.get(str(b.get("pet_id", "")), {})
    out = ai_copilot(pet, owner, b)
    return jsonify(out)


# =========================
# AI COPILOT PAGE
# =========================
@app.route("/copilot")
def copilot_page():
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    items = [decorate_booking(b, ob, pb) for b in bookings_rows]
    items.sort(key=lambda x: parse_dt(str(x.get("appointment_start", ""))) or datetime.min, reverse=True)

    selected_id = (request.args.get("booking_id") or "").strip()
    selected = next((x for x in items if str(x.get("id", "")) == selected_id), None)

    copilot_text = ""
    if selected:
        # build text from API engine
        b_raw = find_by_id(bookings_rows, selected_id)
        owner = ob.get(str(b_raw.get("owner_id", "")), {})
        pet = pb.get(str(b_raw.get("pet_id", "")), {})
        out = ai_copilot(pet, owner, b_raw)
        copilot_text = out.get("plan_text", "") + "\n\nQuestions:\n- " + "\n- ".join(out.get("questions", [])) + \
                       "\n\nTests:\n- " + "\n- ".join(out.get("tests", [])) + \
                       f"\n\nNo-show risk: {out.get('no_show_risk', {}).get('bucket', '')} ({out.get('no_show_risk', {}).get('score', '')}%)\n" + \
                       "\n" + out.get("disclaimer", "")

    return render_template("copilot.html",
                           title=f"{APP_TITLE} | AI Copilot",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="AI Copilot",
                           subtitle="Decision-support triage + checklists + tests + no-show risk",
                           active="copilot",
                           items=items[:60],
                           selected=selected,
                           copilot_text=copilot_text
                           )


# =========================
# DASHBOARD + CHARTS
# =========================
def chart_data():
    pets = read_all(PETS_XLSX)
    bookings = read_all(BOOKINGS_XLSX)

    status_counts = Counter([str(b.get("status") or "Unknown") for b in bookings])
    status_labels = list(status_counts.keys()) or ["No data"]
    status_values = list(status_counts.values()) or [0]

    today = date.today()
    days = [(today - timedelta(days=i)) for i in range(13, -1, -1)]
    trend_map = {d.strftime("%Y-%m-%d"): 0 for d in days}
    for b in bookings:
        dt = parse_dt(str(b.get("appointment_start", "")))
        if dt:
            k = dt.strftime("%Y-%m-%d")
            if k in trend_map:
                trend_map[k] += 1
    trend_labels = list(trend_map.keys())
    trend_values = list(trend_map.values())

    species_counts = Counter([str(p.get("species") or "Unknown") for p in pets])
    species_labels = list(species_counts.keys()) or ["No data"]
    species_values = list(species_counts.values()) or [0]

    cur = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    months = [(cur - timedelta(days=30 * i)).strftime("%Y-%m") for i in range(5, -1, -1)]
    rev_map = {m: 0.0 for m in months}
    for b in bookings:
        if str(b.get("payment_status", "")).lower() == "paid":
            dt = parse_dt(str(b.get("appointment_start", "")))
            if dt:
                m = dt.strftime("%Y-%m")
                if m in rev_map:
                    rev_map[m] += to_float(b.get("fee_amount"))
    rev_labels = list(rev_map.keys())
    rev_values = [round(v, 2) for v in rev_map.values()]

    now = datetime.now()
    upcoming = 0
    cutoff = now + timedelta(hours=24)
    for b in bookings:
        dt = parse_dt(str(b.get("appointment_start", "")))
        if dt and now <= dt <= cutoff and str(b.get("status", "")) not in ("Cancelled", "No-Show"):
            upcoming += 1

    return {
        "status_labels": status_labels, "status_values": status_values,
        "trend_labels": trend_labels, "trend_values": trend_values,
        "species_labels": species_labels, "species_values": species_values,
        "rev_labels": rev_labels, "rev_values": rev_values,
        "kpi": {"total": len(bookings), "upcoming": upcoming, "revenue": round(sum(rev_values), 2)}
    }


def png(fig):
    buf = BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=140, transparent=True)
    plt.close(fig)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")


@app.route("/dashboard")
def dashboard():
    gate = require_login()
    if gate: return gate
    d = chart_data()
    return render_template("dashboard.html",
                           title=f"{APP_TITLE} | Dashboard",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Dashboard",
                           subtitle="Offline charts + KPIs",
                           active="dashboard",
                           kpi=d["kpi"]
                           )


@app.route("/report")
def report():
    gate = require_login()
    if gate:
        return gate

    # -------- Criteria --------
    # Date range (defaults: last 30 days)
    def _parse_date_only(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    today = date.today()
    default_start = today - timedelta(days=30)

    start_s = request.args.get("start", "").strip()
    end_s = request.args.get("end", "").strip()

    start_d = _parse_date_only(start_s) or default_start
    end_d = _parse_date_only(end_s) or today

    if start_d > end_d:
        start_d, end_d = end_d, start_d

    view = (request.args.get("view", "all") or "all").strip()

    vet_f = (request.args.get("vet", "") or "").strip()
    room_f = (request.args.get("room", "") or "").strip()
    status_f = (request.args.get("status", "") or "").strip()
    service_f = (request.args.get("service", "") or "").strip()

    owner_q = (request.args.get("owner_q", "") or "").strip().lower()
    phone_q = re.sub(r"\D+", "", (request.args.get("phone_q", "") or "").strip())
    pet_q = (request.args.get("pet_q", "") or "").strip().lower()

    min_due_s = (request.args.get("min_due", "") or "").strip()
    try:
        min_due = float(min_due_s) if min_due_s else None
    except Exception:
        min_due = None

    # Lookup tables for filtering
    owners_by_id = {str(o.get("id", "")): o for o in read_all(OWNERS_XLSX)}
    pets_by_id = {str(p.get("id", "")): p for p in read_all(PETS_XLSX)}

    def _appt_date(b):
        dt = parse_dt(str(b.get("appointment_start", "") or b.get("appointment_sta", "") or ""))
        return dt.date() if dt else None

    def _owner_matches(b):
        if not owner_q and not phone_q:
            return True
        oid = str(b.get("owner_id", "") or "")
        o = owners_by_id.get(oid)
        if not o:
            return False
        nm = (str(o.get("owner_name", "") or o.get("name", "") or "")).lower()
        ph = re.sub(r"\D+", "", str(o.get("phone", "") or o.get("mobile", "") or ""))
        if owner_q and owner_q not in nm:
            return False
        if phone_q and phone_q not in ph:
            return False
        return True

    def _pet_matches(b):
        if not pet_q:
            return True
        pid = str(b.get("pet_id", "") or "")
        p = pets_by_id.get(pid)
        if not p:
            return False
        nm = (str(p.get("pet_name", "") or p.get("name", "") or "")).lower()
        return pet_q in nm

    def _services_list(b):
        # services_json preferred
        sj = b.get("services_json")
        if sj:
            try:
                j = json.loads(sj) if isinstance(sj, str) else sj
                if isinstance(j, list):
                    return j
            except Exception:
                pass
        # fallback to single service name/fee if present
        nm = str(b.get("service_name", "") or "").strip()
        fee = to_float(b.get("service_fee"), 0.0)
        if nm:
            return [{"name": nm, "fee": fee}]
        return []

    def _service_matches(b):
        if not service_f:
            return True
        for s in _services_list(b):
            if str(s.get("name", "")).strip() == service_f:
                return True
        return False

    def _due_amount(b, computed_fee=None):
        fee = computed_fee if computed_fee is not None else round(to_float(b.get("total_fee", 0.0), 2), 2)
        paid = round(to_float(b.get("paid_amount", 0.0), 2), 2)
        due = round(to_float(b.get("due_amount", 0.0), 2), 2)
        if due <= 0 and fee > 0:
            due = round(max(fee - paid, 0.0), 2)
        return due

    # -------- Load + filter bookings --------
    all_bookings = read_all(BOOKINGS_XLSX)

    filtered = []
    for b in all_bookings:
        ad = _appt_date(b)
        if not ad:
            continue
        if ad < start_d or ad > end_d:
            continue

        if vet_f and str(b.get("vet_name", "") or "").strip() != vet_f:
            continue
        if room_f and str(b.get("room_name", "") or "").strip() != room_f:
            continue
        if status_f and str(b.get("status", "") or "").strip() != status_f:
            continue
        if not _owner_matches(b):
            continue
        if not _pet_matches(b):
            continue
        if not _service_matches(b):
            continue

        services = _services_list(b)
        fee = round(sum([to_float(s.get("fee"), 0.0) for s in services]), 2)
        if fee <= 0:
            fee = round(to_float(b.get("total_fee", 0.0), 2), 2)
        due = _due_amount(b, computed_fee=fee)

        if min_due is not None and due < min_due:
            continue

        # cache for later computations
        b["_appt_date"] = ad
        b["_services"] = services
        b["_fee"] = fee

        # Discount + net revenue (after discount)
        discount = round(to_float(b.get("discount"), 0.0), 2)
        if discount < 0:
            discount = 0.0
        if discount > fee:
            discount = fee
        fee_field = round(to_float(b.get("fee_amount"), -1.0), 2)
        net_fee = fee_field if fee_field >= 0 else round(fee - discount, 2)

        paid = round(to_float(b.get("paid_amount", 0.0), 2), 2)
        due = round(to_float(b.get("due_amount", 0.0), 2), 2)
        if due <= 0 and net_fee > 0:
            due = round(max((net_fee + round(net_fee * float(VAT_RATE), 2)) - paid, 0.0), 2)

        b["_discount"] = discount
        b["_net_fee"] = net_fee
        b["_due"] = due
        filtered.append(b)

    # Collections for filter dropdowns (based on active config)
    vets = active_vet_names()
    rooms = active_room_names()
    statuses = STATUS_FLOW
    services_cfg = get_services(include_inactive=False)
    services_list = [s.get("name", "") for s in services_cfg if (s.get("name") or "").strip()]

    # -------- Build the 8 reports (based on filtered) --------
    total_exams = len(filtered)

    owner_ids = set()
    pet_ids = set()
    total_fee = 0.0
    total_paid = 0.0
    total_due = 0.0
    total_discounts = 0.0
    channel_totals = {}

    trend_map = {}  # date -> {count, revenue}
    status_counter = {}
    service_counter = {}
    vet_counter = {}
    room_counter = {}

    completed_count = 0

    for b in filtered:
        oid = str(b.get("owner_id", "") or "")
        pid = str(b.get("pet_id", "") or "")
        if oid:
            owner_ids.add(oid)
        if pid:
            pet_ids.add(pid)

        st = str(b.get("status", "") or "Unknown").strip() or "Unknown"
        status_counter[st] = status_counter.get(st, 0) + 1
        if st.lower() == "completed":
            completed_count += 1

        ad = b.get("_appt_date")
        fee = b.get("_fee", 0.0)
        paid = round(to_float(b.get("paid_amount"), 0.0), 2)
        due = b.get("_due", 0.0)

        total_fee += b.get("_net_fee", fee)
        total_paid += paid
        total_due += due
        total_discounts += b.get("_discount", 0.0)
        ch = (b.get("payment_channel") or b.get("payment_method") or "").strip() or "Unspecified"
        channel_totals[ch] = channel_totals.get(ch, 0.0) + b.get("_net_fee", fee)

        if ad:
            key = ad.isoformat()
            if key not in trend_map:
                trend_map[key] = {"count": 0, "revenue": 0.0}
            trend_map[key]["count"] += 1
            trend_map[key]["revenue"] += fee

        # Service mix
        for s in b.get("_services", []):
            nm = str(s.get("name", "") or "").strip()
            if not nm:
                continue
            if nm not in service_counter:
                service_counter[nm] = {"count": 0, "revenue": 0.0}
            service_counter[nm]["count"] += 1
            service_counter[nm]["revenue"] += to_float(s.get("fee"), 0.0)

        # Vet performance
        vet = str(b.get("vet_name", "") or "").strip() or "Unassigned"
        if vet not in vet_counter:
            vet_counter[vet] = {"count": 0, "completed": 0, "revenue": 0.0}
        vet_counter[vet]["count"] += 1
        if st.lower() == "completed":
            vet_counter[vet]["completed"] += 1
        vet_counter[vet]["revenue"] += fee

        # Room utilization
        rm = str(b.get("room_name", "") or "").strip() or "Unassigned"
        if rm not in room_counter:
            room_counter[rm] = {"count": 0, "revenue": 0.0}
        room_counter[rm]["count"] += 1
        room_counter[rm]["revenue"] += fee

    completion_rate = round((completed_count / total_exams) * 100, 1) if total_exams else 0.0
    collection_rate = round((total_paid / total_fee) * 100, 1) if total_fee else 0.0

    # Trend rows (sorted)
    trend_rows = []
    for k in sorted(trend_map.keys()):
        trend_rows.append({"date": k, "count": trend_map[k]["count"], "revenue": round(trend_map[k]["revenue"], 2)})

    # Status rows
    status_rows = []
    for st, cnt in sorted(status_counter.items(), key=lambda x: x[1], reverse=True):
        pct = (cnt / total_exams * 100) if total_exams else 0
        status_rows.append({"status": st, "count": cnt, "pct": pct})

    # Financial summary object
    fin = {
        "revenue_total": round(total_fee, 2),  # after discount
        "discount_total": round(total_discounts, 2),
        "collected_total": round(total_paid, 2),
        "due_total": round(total_due, 2),
        "collection_rate": collection_rate,
        "channel_totals": channel_totals
    }

    # Daily Closure (end-of-day summary)
    closure = {}
    for b in filtered:
        ad = b.get("_appt_date")
        if not ad:
            continue
        day_key = ad.strftime("%Y-%m-%d")
        rec = closure.get(day_key) or {"date": day_key, "count": 0, "revenue": 0.0, "discounts": 0.0,
                                       "cash": 0.0, "visa": 0.0, "instapay": 0.0, "other": 0.0}
        rec["count"] += 1
        rev = float(b.get("_net_fee", 0.0) or 0.0)
        disc = float(b.get("_discount", 0.0) or 0.0)
        rec["revenue"] += rev
        rec["discounts"] += disc
        ch = (b.get("payment_channel") or b.get("payment_method") or "").strip().lower()
        if ch == "cash":
            rec["cash"] += rev
        elif ch == "visa":
            rec["visa"] += rev
        elif ch == "instapay":
            rec["instapay"] += rev
        else:
            rec["other"] += rev
        closure[day_key] = rec

    closure_rows = sorted(closure.values(), key=lambda x: x["date"], reverse=True)

    # Top services
    top_services = sorted(
        [{"name": k, "count": v["count"], "revenue": round(v["revenue"], 2)} for k, v in service_counter.items()],
        key=lambda x: (x["revenue"], x["count"]),
        reverse=True
    )[:10]

    # Vet rows
    vet_rows = []
    for vet, v in sorted(vet_counter.items(), key=lambda x: x[1]["count"], reverse=True):
        cr = (v["completed"] / v["count"] * 100) if v["count"] else 0
        vet_rows.append({"vet": vet, "count": v["count"], "completed_pct": cr, "revenue": round(v["revenue"], 2)})

    # Room rows
    room_rows = []
    for rm, v in sorted(room_counter.items(), key=lambda x: x[1]["count"], reverse=True):
        room_rows.append({"room": rm, "count": v["count"], "revenue": round(v["revenue"], 2)})

    # Due rows (top 20)
    due_rows = []
    for b in filtered:
        due = b.get("_due", 0.0)
        if due <= 0:
            continue
        ad = b.get("_appt_date")
        owner = owners_by_id.get(str(b.get("owner_id", "")))
        pet = pets_by_id.get(str(b.get("pet_id", "")))
        due_rows.append({
            "date": ad.isoformat() if ad else "",
            "owner": str(owner.get("owner_name", "") if owner else "") or str(
                owner.get("name", "") if owner else "") or "",
            "phone": str(owner.get("phone", "") if owner else "") or str(
                owner.get("mobile", "") if owner else "") or "",
            "pet": str(pet.get("pet_name", "") if pet else "") or str(pet.get("name", "") if pet else "") or "",
            "vet": str(b.get("vet_name", "") or ""),
            "status": str(b.get("status", "") or ""),
            "due": round(due, 2)
        })
    due_rows.sort(key=lambda x: (x["due"], x["date"]), reverse=True)
    due_rows = due_rows[:20]

    # -------- Futuristic points --------
    # 1) 7-day forecast (simple moving average)
    last_days = sorted(trend_map.keys())[-14:]  # use up to 14 days
    avg_count = 0.0
    avg_rev = 0.0
    if last_days:
        avg_count = sum([trend_map[d]["count"] for d in last_days]) / len(last_days)
        avg_rev = sum([trend_map[d]["revenue"] for d in last_days]) / len(last_days)
    forecast = []
    for i in range(1, 8):
        fd = (end_d + timedelta(days=i)).isoformat()
        forecast.append({"date": fd, "exams": round(avg_count, 1), "revenue": round(avg_rev, 2)})
    forecast_totals = {"exams": round(avg_count * 7, 1), "revenue": round(avg_rev * 7, 2)}

    # 2) Early-warning signals (week-over-week)
    def _sum_for_range(d1, d2):
        c = 0
        r = 0.0
        canc = 0
        due_sum = 0.0
        for b in filtered:
            ad = b.get("_appt_date")
            if not ad:
                continue
            if d1 <= ad <= d2:
                c += 1
                r += b.get("_fee", 0.0)
                if str(b.get("status", "") or "").strip().lower() == "cancelled":
                    canc += 1
                due_sum += b.get("_due", 0.0)
        return {"count": c, "revenue": r, "cancelled": canc, "due": due_sum}

    last7_start = end_d - timedelta(days=6)
    prev7_end = last7_start - timedelta(days=1)
    prev7_start = prev7_end - timedelta(days=6)

    w1 = _sum_for_range(last7_start, end_d)
    w0 = _sum_for_range(prev7_start, prev7_end)

    alerts = []

    def _pct_change(a, b):
        if b == 0:
            return None
        return (a - b) / b * 100.0

    # cancellation rate change
    w1_cr = (w1["cancelled"] / w1["count"] * 100) if w1["count"] else 0
    w0_cr = (w0["cancelled"] / w0["count"] * 100) if w0["count"] else 0
    ch = _pct_change(w1_cr, w0_cr)
    if ch is not None and ch >= 25:
        alerts.append(f"Cancellation rate increased by {ch:.1f}% week-over-week ({w0_cr:.1f}% → {w1_cr:.1f}%).")

    # due growth
    ch_due = _pct_change(w1["due"], w0["due"])
    if ch_due is not None and ch_due >= 25:
        alerts.append(f"Outstanding due increased by {ch_due:.1f}% week-over-week ({w0['due']:.2f} → {w1['due']:.2f}).")

    # low collection
    if collection_rate < 70 and total_fee > 0:
        alerts.append(
            f"Collection rate is low at {collection_rate:.1f}%. Consider prioritizing reminder calls for top due owners.")

    if not alerts:
        alerts.append("No significant early-warning signals detected in the selected criteria window.")

    # -------- Render --------
    return render_template(
        "report.html",
        title="Reports",
        header="Reports",
        subtitle="Filter-driven executive reports",
        active="report",
        # criteria
        start=start_d.isoformat(),
        end=end_d.isoformat(),
        view=view,
        vet_f=vet_f,
        room_f=room_f,
        status_f=status_f,
        service_f=service_f,
        owner_q=request.args.get("owner_q", ""),
        phone_q=request.args.get("phone_q", ""),
        pet_q=request.args.get("pet_q", ""),
        min_due=min_due_s,
        # dropdown data
        vets=vets,
        rooms=rooms,
        statuses=statuses,
        services=services_list,
        # reports
        total_exams=total_exams,
        unique_owners=len(owner_ids),
        unique_pets=len(pet_ids),
        completion_rate=completion_rate,
        fin=fin,
        closure_rows=closure_rows,
        trend_rows=trend_rows,
        status_rows=status_rows,
        top_services=top_services,
        vet_rows=vet_rows,
        room_rows=room_rows,
        due_rows=due_rows,
        # futuristic
        forecast=forecast,
        forecast_totals=forecast_totals,
        alerts=alerts
    )


@app.route("/charts/status.png")
def chart_status_png():
    gate = require_login()
    if gate: return gate
    d = chart_data()
    fig = plt.figure()
    plt.pie(d["status_values"], labels=d["status_labels"], wedgeprops=dict(width=0.45))
    plt.title("Bookings by Status")
    return png(fig)


@app.route("/charts/trend.png")
def chart_trend_png():
    gate = require_login()
    if gate: return gate
    d = chart_data()
    fig = plt.figure()
    plt.bar(d["trend_labels"], d["trend_values"])
    plt.xticks(rotation=45, ha="right")
    plt.title("Appointments Trend (last 14 days)")
    return png(fig)


@app.route("/charts/species.png")
def chart_species_png():
    gate = require_login()
    if gate: return gate
    d = chart_data()
    fig = plt.figure()
    plt.pie(d["species_values"], labels=d["species_labels"])
    plt.title("Pets by Species")
    return png(fig)


@app.route("/charts/revenue.png")
def chart_revenue_png():
    gate = require_login()
    if gate: return gate
    d = chart_data()
    fig = plt.figure()
    plt.plot(d["rev_labels"], d["rev_values"], marker="o")
    plt.xticks(rotation=45, ha="right")
    plt.title("Paid Revenue (last 6 months)")
    return png(fig)


# =========================
# HISTORY
# =========================
@app.route("/history")
def history():
    gate = require_login()
    if gate: return gate
    q = (request.args.get("q") or "").strip().lower()

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}
    bookings_rows = read_all(BOOKINGS_XLSX)
    view = [decorate_booking(b, ob, pb) for b in bookings_rows]

    view = [x for x in view if (parse_dt(str(x.get("appointment_start", ""))) or datetime.max) < datetime.now()]
    if q:
        view = [x for x in view if q in (" ".join([
            str(x.get("owner_name", "")), str(x.get("pet_name", "")),
            str(x.get("appointment_type", "")), str(x.get("status", "")),
            str(x.get("reason", "")), str(x.get("symptoms", ""))
        ])).lower()]
    view.sort(key=lambda x: parse_dt(str(x.get("appointment_start", ""))) or datetime.min, reverse=True)

    return render_template("history.html",
                           title=f"{APP_TITLE} | History",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="History",
                           subtitle="Past bookings and audit view",
                           active="history",
                           rows=view,
                           q=q
                           )


# =====================================================================
# ADD-ON PACK (NO CHANGES TO YOUR CURRENT CODE)
# - Printable Invoice (Ready to Print)
# - Smoother Exam page + button
# - Enhanced Reminders (Auto-schedule + Open WhatsApp from reminder row)
# Paste this block BEFORE the "# ========================= # STARTUP" section
# =====================================================================

# -------------------------
# GLOBAL TEMPLATE HELPERS
# -------------------------
@app.context_processor
def inject_runtime_helpers():
    # now_key is a sortable string like "2025-12-14 20:15"
    return {"now_key": datetime.now().strftime("%Y-%m-%d %H:%M")}


# -------------------------
# INVOICE CONFIG
# -------------------------
INVOICE_PREFIX = "EV"
CLINIC_NAME = APP_TITLE
CLINIC_ADDRESS = "Cairo, Egypt"
CLINIC_PHONE = "+20 000 000 0000"
CLINIC_EMAIL = "contact@elitevetclinic.local"
VAT_RATE = 0.00  # set 0.14 if you want 14% (Egypt VAT example). Keep 0 if not needed.


def next_slot(minutes_ahead=60, step_min=15):
    """Returns a datetime rounded up to next 'step_min' minutes."""
    dt = datetime.now() + timedelta(minutes=minutes_ahead)
    dt = dt.replace(second=0, microsecond=0)
    m = dt.minute
    r = (step_min - (m % step_min)) % step_min
    return dt + timedelta(minutes=r)


def ensure_invoice_no_for_booking(booking_id: str) -> str:
    """If invoice_no is missing, generate one and save it into BOOKINGS_XLSX."""
    bookings = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings, booking_id)
    if not b:
        return ""

    inv = str(b.get("invoice_no", "") or "").strip()
    if inv:
        return inv

    # Deterministic, readable invoice number
    start_dt = parse_dt(str(b.get("appointment_start", ""))) or datetime.now()
    inv = f"{INVOICE_PREFIX}-{start_dt.strftime('%Y%m%d')}-{str(booking_id)[:6].upper()}"

    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
        "invoice_no": inv,
        "updated_at": now_str()
    })
    return inv


def _safe_money(v):
    x = to_float(v, 0.0)
    return round(x, 2)


# -------------------------
# ROUTE: PRINTABLE INVOICE
# -------------------------
@app.route("/invoice/<booking_id>")
def invoice_print(booking_id):
    gate = require_login()
    if gate:
        return gate

    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Booking not found.")
        return redirect(url_for("bookings"))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o.get("id", "")): o for o in owners_rows}
    pb = {str(p.get("id", "")): p for p in pets_rows}

    owner = ob.get(str(b.get("owner_id", "")), {})
    pet = pb.get(str(b.get("pet_id", "")), {})

    invoice_no = ensure_invoice_no_for_booking(booking_id)

    appt_start = normalize_dt(str(b.get("appointment_start", "")))
    appt_end = normalize_dt(str(b.get("appointment_end", "")))
    dur = int(float(b.get("duration_min") or 30))

    # Line items (multi-service)
    services = parse_services_json(b.get("services_json", ""))
    line_items = []
    if services:
        for it in services:
            if not isinstance(it, dict):
                continue
            desc = str(it.get("name") or "").strip()
            if not desc:
                continue
            try:
                qty = max(1, int(float(it.get("qty", 1) or 1)))
            except Exception:
                qty = 1
            unit = _safe_money(it.get("fee", 0))
            lt = it.get("line_total")
            if lt is None or str(lt).strip() == "":
                lt = unit * qty
            lt = _safe_money(lt)
            line_items.append({
                "desc": desc,
                "details": "",
                "qty": qty,
                "unit_price": unit,
                "line_total": lt
            })

    if not line_items:
        # Fallback legacy single-service fields
        desc = str((b.get("service_name") or b.get("appointment_type") or "Service"))
        details = str(b.get("reason", "") or "").strip()
        lt = _safe_money(b.get("service_fee") or b.get("fee_amount") or 0)
        line_items = [{
            "desc": desc,
            "details": details,
            "qty": 1,
            "unit_price": lt,
            "line_total": lt
        }]

    subtotal = round(sum(_safe_money(x.get("line_total", 0)) for x in line_items), 2)

    discount_type = normalize_discount_type(b.get("discount_type") or "value")
    discount_value = str(b.get("discount_value") or "").strip()
    discount_amt = round(compute_discount_amount(subtotal, discount_type, discount_value, b.get("discount")), 2)

    net_fee = round(subtotal - discount_amt, 2)
    vat = round(net_fee * float(VAT_RATE), 2)
    total = round(net_fee + vat, 2)

    paid_amt = _safe_money(b.get("paid_amount", 0))
    due_amt = round(total - paid_amt, 2)
    if due_amt < 0:
        due_amt = 0.0

    return render_template(
        "invoice_print.html",
        title=f"{APP_TITLE} | Invoice {invoice_no}",
        app_title=APP_TITLE,
        logged_in=True,
        booking=b,
        owner=owner,
        pet=pet,
        invoice_no=invoice_no,
        appt_start=appt_start,
        appt_end=appt_end,
        dur=dur,
        subtotal=subtotal,
        discount_amt=discount_amt,
        discount_type=discount_type,
        discount_value=discount_value,
        net_fee=net_fee,
        vat=vat,
        total=total,
        paid=paid_amt,
        due=due_amt,
        payment_channel=str(b.get("payment_channel", "") or ""),
        payment_method=str(b.get("payment_method", "") or ""),
        vat_rate=float(VAT_RATE),
        line_items=line_items,
        clinic={
            "name": CLINIC_NAME,
            "address": CLINIC_ADDRESS,
            "phone": CLINIC_PHONE,
        }
    )


# -------------------------
# ROUTE: SMOOTHER BOOKING (FAST WIZARD)
# -------------------------
@app.route("/smoother", methods=["GET", "POST"])
def smoother_booking():
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    owners_rows.sort(key=lambda x: str(x.get("owner_name", "")).lower())
    pets_rows = read_all(PETS_XLSX)
    pets_rows.sort(key=lambda x: str(x.get("pet_name", "")).lower())

    if request.method == "POST":
        bid = str(uuid.uuid4())

        start = normalize_dt(request.form.get("appointment_start") or "")
        if not start:
            start = datetime.now().strftime("%Y-%m-%d %H:%M")

        dur = str((request.form.get("duration_min") or "30").strip())
        start_dt = parse_dt(start) or datetime.now()
        end_dt = start_dt + timedelta(minutes=int(float(dur)))

        token = uuid.uuid4().hex

        row = {h: "" for h in BOOKINGS_HEADERS}
        row.update({
            "id": bid,
            "appointment_start": start,
            "duration_min": dur,
            "appointment_end": end_dt.strftime("%Y-%m-%d %H:%M"),
            "owner_id": (request.form.get("owner_id") or "").strip(),
            "pet_id": (request.form.get("pet_id") or "").strip(),
            "visit_weight_kg": (request.form.get("visit_weight_kg") or "").strip(),
            "visit_temp_c": (request.form.get("visit_temp_c") or "").strip(),
            "appointment_type": (request.form.get("appointment_type") or "Consultation").strip(),
            "priority": (request.form.get("priority") or "Normal").strip(),
            "status": (request.form.get("status") or "Scheduled").strip(),
            "channel": (request.form.get("channel") or "WhatsApp").strip(),
            "reason": (request.form.get("reason") or "").strip(),
            "symptoms": (request.form.get("symptoms") or "").strip(),
            "vet_name": (request.form.get("vet_name") or "").strip(),
            "room": (request.form.get("room") or "").strip(),
            "services_json": (request.form.get("services_json") or "").strip(),
            "service_name": (request.form.get("service_name") or "").strip(),
            "service_fee": "",  # will be computed from services_json

            "discount_type": (request.form.get("discount_type") or "value").strip(),
            "discount_value": (request.form.get("discount_value") or "").strip(),

            "paid_amount": (request.form.get("paid_amount") or "").strip(),
            "due_amount": (request.form.get("due_amount") or "").strip(),
            "fee_amount": (request.form.get("service_fee") or "").strip(),
            "payment_status": "",
            "payment_method": (request.form.get("payment_channel") or request.form.get("payment_method") or "").strip(),
            "payment_channel": (request.form.get("payment_channel") or "").strip(),
            "invoice_no": "",
            "followup_datetime": "",
            "reminder_channel": (request.form.get("reminder_channel") or "WhatsApp").strip(),
            "reminder_sent": "",
            "reminder_last_opened": "",
            "portal_token": token,
            "owner_confirmed": "",
            "owner_update_message": "",
            "owner_update_datetime": "",
            "ai_last_applied_at": "",
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })

        # --- Service/Fee calculation (server-side safety) ---
        services = []
        raw_services = (row.get("services_json") or "").strip()
        if not raw_services:
            raw_services = "[]"
            row["services_json"] = raw_services
        try:
            services = json.loads(raw_services)
            if not isinstance(services, list):
                services = []
        except Exception:
            services = []
            row["services_json"] = "[]"

        subtotal = 0.0
        for it in services:
            if not isinstance(it, dict):
                continue
            qty = it.get("qty", 1) or 1
            try:
                qty = max(1, int(float(qty)))
            except Exception:
                qty = 1
            fee = to_float(it.get("fee", 0.0), 0.0)
            line_total = it.get("line_total", None)
            if line_total is None or str(line_total).strip() == "":
                line_total = fee * qty
            line = to_float(line_total, fee * qty)
            subtotal += line

        svc_fee = round(subtotal, 2)

        if (not (row.get("service_name") or "").strip()) and services:
            first = str((services[0] or {}).get("name", "") or "").strip()
            if first:
                row["service_name"] = first if len(services) == 1 else f"{first} +{len(services) - 1}"

        paid = _safe_money(row.get("paid_amount", 0))

        # Discount handling (validated server-side): supports value or percentage
        row["discount_type"] = normalize_discount_type(row.get("discount_type") or "value")
        row["discount_value"] = str(row.get("discount_value", "") or "").strip()
        discount = compute_discount_amount(svc_fee, row.get("discount_type"), row.get("discount_value"),
                                           row.get("discount"))
        net_fee = round(svc_fee - discount, 2)

        vat_calc = round(net_fee * float(VAT_RATE), 2)
        total_calc = round(net_fee + vat_calc, 2)
        due_calc2 = round(total_calc - paid, 2)

        row["service_fee"] = f"{svc_fee:.2f}" if svc_fee else ""
        row["discount"] = f"{discount:.2f}" if discount else ""
        row["fee_amount"] = f"{net_fee:.2f}" if (svc_fee or discount) else ""
        row["paid_amount"] = f"{paid:.2f}" if paid else ""
        row["due_amount"] = f"{max(due_calc2, 0):.2f}" if (svc_fee or paid) else ""

        if total_calc <= 0:
            row["payment_status"] = ""
        elif paid <= 0:
            row["payment_status"] = "Unpaid"
        elif paid + 0.0001 >= total_calc:
            row["payment_status"] = "Paid"
        else:
            row["payment_status"] = "Partial"

        append_row(BOOKINGS_XLSX, BOOKINGS_HEADERS, row)

        flash("Smoother Exam created.")

        # Optional: open WhatsApp journey immediately
        open_wa = (request.form.get("open_wa") or "").strip().lower() == "on"
        if open_wa:
            return redirect(url_for("booking_remind", booking_id=bid))

        return redirect(url_for("booking_view", booking_id=bid))

    # GET defaults
    default_start = datetime.now().strftime("%Y-%m-%d %H:%M")
    return render_template(
        "smoother_booking.html",
        title=f"{APP_TITLE} | Smoother Exam",
        app_title=APP_TITLE,
        logged_in=True,
        header="Smoother Exam",
        subtitle="Fast booking wizard + Smart Intake + optional instant WhatsApp Journey",
        active="bookings",
        owners=owners_rows,
        pets=pets_rows,
        statuses=STATUS_FLOW,
        types=APPOINTMENT_TYPES,
        priorities=PRIORITIES,
        channels=CHANNELS,
        payment_statuses=PAYMENT_STATUSES,
        payment_methods=PAYMENT_METHODS,
        reminder_channels=REMINDER_CHANNELS,
        default_vet=session.get("username", ""),
        vets=active_vet_names(),
        rooms=active_room_names(),
        default_start=default_start
    )


# -------------------------
# REMINDERS ENHANCEMENT: OPEN WHATSAPP FROM REMINDER ROW
# -------------------------
@app.route("/reminders/<reminder_id>/open")
def reminder_open(reminder_id):
    gate = require_login()
    if gate: return gate

    rem_rows = read_all(REMINDERS_XLSX)
    r = find_by_id(rem_rows, reminder_id)
    if not r:
        flash("Reminder not found.")
        return redirect(url_for("reminders"))

    owners_rows = read_all(OWNERS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    owner = ob.get(str(r.get("owner_id", "")), {})

    msg = str(r.get("message", "") or "")
    wa = whatsapp_link(owner.get("phone", ""), msg)

    # Mark as opened when user clicks the WA button from reminders page
    update_row_by_id(REMINDERS_XLSX, REMINDERS_HEADERS, reminder_id, {
        "status": "Opened",
        "opened_at": now_str(),
        "updated_at": now_str()
    })

    return redirect(wa)


# -------------------------
# REMINDERS ENHANCEMENT: AUTO-SCHEDULE REMINDERS
# -------------------------
@app.route("/reminders/auto_schedule", methods=["POST"])
def reminders_auto_schedule():
    gate = require_login()
    if gate: return gate

    try:
        hours_before = float((request.form.get("hours_before") or "2").strip())
    except Exception:
        hours_before = 2.0

    try:
        days_ahead = int((request.form.get("days_ahead") or "7").strip())
    except Exception:
        days_ahead = 7

    hours_before = max(0.0, min(168.0, hours_before))
    days_ahead = max(1, min(60, days_ahead))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    rem_rows = read_all(REMINDERS_XLSX)

    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    now_dt = datetime.now()
    cutoff = now_dt + timedelta(days=days_ahead)

    # quick index to avoid duplicates
    existing_keys = set()
    for rr in rem_rows:
        rtype = str(rr.get("reminder_type") or "Appointment").strip() or "Appointment"
        sname = str(rr.get("service_name") or "").strip()
        existing_keys.add((str(rr.get("booking_id", "")), str(rr.get("scheduled_for", "")), rtype, sname))

    created = 0
    for b in bookings_rows:
        st = str(b.get("status", ""))
        if st in ("Completed", "Cancelled", "No-Show"):
            continue

        appt_dt = parse_dt(str(b.get("appointment_start", "")))
        if not appt_dt:
            continue
        if not (now_dt <= appt_dt <= cutoff):
            continue

        sched_dt = appt_dt - timedelta(hours=hours_before)
        if sched_dt < now_dt:
            continue

        scheduled_for = sched_dt.strftime("%Y-%m-%d %H:%M")
        key = (str(b.get("id", "")), scheduled_for, "Appointment", "")
        if key in existing_keys:
            continue

        # Build message with portal link
        db = decorate_booking(b, ob, pb)
        msg = booking_message_template(
            db.get("owner_name", ""),
            db.get("pet_name", ""),
            db.get("appointment_start", ""),
            portal_link=db.get("portal_link", "")
        )

        rid = str(uuid.uuid4())
        append_row(REMINDERS_XLSX, REMINDERS_HEADERS, {
            "id": rid,
            "booking_id": str(b.get("id", "")),
            "owner_id": str(b.get("owner_id", "")),
            "pet_id": str(b.get("pet_id", "")),
            "reminder_type": "Appointment",
            "service_name": "",
            "channel": "WhatsApp",
            "status": "Scheduled",
            "scheduled_for": scheduled_for,
            "opened_at": "",
            "sent_at": "",
            "message": msg,
            "created_at": now_str(),
            "updated_at": now_str()
        })

        existing_keys.add(key)
        created += 1

    flash(f"Auto-scheduled {created} reminder(s). (Hours before: {hours_before}, days ahead: {days_ahead})")
    return redirect(url_for("reminders"))


# -------------------------
# TEMPLATE OVERRIDES (NO EDITS TO YOUR ORIGINAL TEMPLATES REQUIRED)
# We overwrite only these templates by updating TEMPLATES dict.
# -------------------------
TEMPLATES.update({
    "home.html": r"""
{% extends "base.html" %}
{% block content %}
  <div class="grid two">
    <a class="card" href="{{ url_for('bookings') }}">
      <h3>Booking Journey</h3>
      <div class="muted">Create appointment → WhatsApp Journey → Owner Portal → Calendar → Popups.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill good">WhatsApp Journey</span>
        <span class="pill">Owner Portal</span>
        <span class="pill warn">AI Intake</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('smoother_booking') }}">
      <h3>Smoother Exam</h3>
      <div class="muted">Fast wizard: Smart Intake → minimal fields → optional instant WhatsApp Journey.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill warn">Fast</span>
        <span class="pill good">Smart Intake</span>
        <span class="pill">One-click WA</span>
      </div>
    </a>
<a class="card" href="{{ url_for('dashboard') }}">
      <h3>Dashboard</h3>
      <div class="muted">Offline charts + KPIs, track performance.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Charts</span>
        <span class="pill good">PNG</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('owners') }}">
      <h3>Owners & Pets</h3>
      <div class="muted">Editable master data.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Owners</span>
        <span class="pill">Pets</span>
        <span class="pill warn">Edit</span>
      </div>
    </a>
  </div>
{% endblock %}
""",

    "bookings.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Bookings</h3>
  <div class="muted">Reminder Journey: WhatsApp + Owner Portal + Calendar + Popups. + Invoice Print.</div>

  <form method="get" class="form" style="margin-top:10px;">
    <div class="full">
      <label>Search</label>
      <input name="q" value="{{ q }}" placeholder="owner / pet / status / type / reason">
    </div>
    <div>
      <label>Status</label>
      <select name="status">
        <option value="">All</option>
        {% for s in statuses %}
          <option value="{{ s }}" {% if status==s %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>
    <div>
      <label>Type</label>
      <select name="atype">
        <option value="">All</option>
        {% for t in types %}
          <option value="{{ t }}" {% if atype==t %}selected{% endif %}>{{ t }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="full row-actions">
      <button class="btn primary" type="submit">Filter</button>
      <a class="btn" href="{{ url_for('booking_new') }}">New Exam</a>
      <a class="btn warn" href="{{ url_for('smoother_booking') }}">Smoother Exam</a>
      <a class="btn" href="{{ url_for('reminders') }}">Reminders Center</a>
    </div>
  </form>

  <div class="hr"></div>

  <div style="overflow:auto">
    <table>
      <thead>
        <tr>
          <th>Date/Time</th><th>Owner</th><th>Pet</th><th>Type</th><th>Status</th><th>Journey</th><th>Action</th>
        </tr>
      </thead>
      <tbody>
        {% for b in bookings %}
        <tr>
          <td><b>{{ b.appointment_start }}</b><div class="small">{{ b.duration_min }} min</div></td>
          <td>{{ b.owner_name }}</td>
          <td>{{ b.pet_name }}</td>
          <td>{{ b.appointment_type }}</td>
          <td>
            {% set st=b.status %}
            <span class="pill {% if st=='Completed' %}good{% elif st in ['Cancelled','No-Show'] %}bad{% else %}warn{% endif %}">{{ st }}</span>
          </td>
          <td style="white-space:nowrap">
            <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp</a>
            <a class="btn" href="{{ b.portal_link }}" target="_blank">Portal</a>
            <a class="btn" href="{{ url_for('calendar_ics', booking_id=b.id) }}">.ics</a>
          </td>
          <td style="white-space:nowrap">
            <a class="btn primary" href="{{ url_for('booking_view', booking_id=b.id) }}">Open</a>
            <a class="btn" href="{{ url_for('booking_edit', booking_id=b.id) }}">Edit</a>
            <a class="btn good" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Completed') }}">Complete</a>
            <a class="btn bad" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Cancelled') }}" onclick="return confirm('Cancel this exam?')">Cancel</a>
            <a class="btn" href="{{ url_for('invoice_print', booking_id=b.id) }}" target="_blank">Invoice</a>
            <a class="btn bad" href="{{ url_for('booking_delete', booking_id=b.id) }}" onclick="return confirm('Delete exam?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
        {% if not bookings %}
          <tr><td colspan="7" class="muted">No exams found.</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

    "booking_view.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="grid two">
  <div class="card">
    <h3>Exam Details</h3>
    <div class="muted">
      <b>{{ b.pet_name }}</b> • Owner: <b>{{ b.owner_name }}</b><br>
      Starts: <b>{{ b.appointment_start }}</b> • Duration: {{ b.duration_min }} min<br>
      Type: {{ b.appointment_type }} • Status: {{ b.status }} • Priority: {{ b.priority }}
    </div>

    <div class="hr"></div>

    <div class="row-actions">
      <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp Journey</a>
      <a class="btn" href="{{ b.portal_link }}" target="_blank">Owner Portal</a>
      <a class="btn" href="{{ b.ics_link }}">Download .ics</a>
            <a class="btn primary" href="{{ b.google_cal_link }}" target="_blank">Google Calendar</a>
      <a class="btn good" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Completed') }}">Mark Completed</a>
      <a class="btn bad" href="{{ url_for('booking_set_status', booking_id=b.id, new_status='Cancelled') }}" onclick="return confirm('Cancel this exam?')">Cancel</a>
      <a class="btn warn" href="{{ url_for('invoice_print', booking_id=b.id) }}" target="_blank">Invoice / Print</a>
    </div>

    <div class="hr"></div>

    <div class="muted">
      Phone: {{ b.owner_phone }}<br>
      Email: {{ b.owner_email }}<br>
      Owner Confirmed: <b>{{ b.owner_confirmed or '—' }}</b><br>
      Owner Update: <span class="small">{{ b.owner_update_message or '—' }}</span><br>
      Owner Update Time: <span class="small">{{ b.owner_update_datetime or '—' }}</span>
    </div>

    <div class="hr"></div>

    <div class="row-actions">
      <a class="btn" href="{{ url_for('booking_edit', booking_id=b.id) }}">Edit Exam</a>
      <a class="btn bad" href="{{ url_for('booking_delete', booking_id=b.id) }}" onclick="return confirm('Delete exam?')">Delete</a>
    </div>
  </div>

  <div class="card">
    <h3>AI Copilot (Futuristic)</h3>
    <div class="muted">Generates triage + risk flags + questions + tests. Decision support only.</div>

    <div class="hr"></div>

    <div id="aiBox" class="muted">Loading AI insights...</div>

    <div class="hr"></div>

    <div class="row-actions">
      <button class="btn warn" type="button" onclick="reloadAI()">Regenerate</button>
      <a class="btn good" id="aiApplyBtn" href="{{ url_for('booking_apply_ai', booking_id=b.id) }}">Apply AI Suggestions</a>
    </div>
  </div>
</div>

<script>
  async function reloadAI(){
    const box = document.getElementById("aiBox");
    box.textContent = "Loading AI insights...";
    const res = await fetch("{{ url_for('api_copilot', booking_id=b.id) }}");
    const data = await res.json();

    if (data.error){
      box.textContent = data.error;
      return;
    }

    const flags = (data.flags||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";
    const qs = (data.questions||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";
    const ts = (data.tests||[]).slice(0,6).map(x=>"- " + x).join("\n") || "- None";

    box.textContent =
      "Triage: " + data.triage + "\n" +
      "Suggested Priority: " + data.suggested_priority + "\n" +
      "Suggested Type: " + data.suggested_type + "\n" +
      "Suggested Duration: " + data.suggested_duration_min + " min\n\n" +
      "No-show risk: " + data.no_show_risk.bucket + " (" + data.no_show_risk.score + "%)\n" +
      (data.no_show_risk.tips && data.no_show_risk.tips.length ? ("Tips:\n" + data.no_show_risk.tips.map(x=>"- "+x).join("\n") + "\n\n") : "\n") +
      "Risk Flags:\n" + flags + "\n\n" +
      "Questions Checklist:\n" + qs + "\n\n" +
      "Suggested Tests:\n" + ts + "\n\n" +
      "Draft:\n" + (data.plan_text || "") + "\n\n" +
      "Note: " + (data.disclaimer || "");
  }

  reloadAI();
</script>
{% endblock %}
""",

    "reminders.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Reminders Center</h3>
  <div class="muted">Auto-schedule reminders + due reminders + upcoming bookings + global records + Pet 360 profile search.</div>

  <div class="hr"></div>

  <!-- Pet 360 Search -->
  <div class="card">
    <h3>Pet 360 (Search)</h3>
    <div class="muted">Search by Owner Name / Owner Phone / Pet Name / Pet ID. Results appear first; then select the correct pet from the dropdown to load the full profile (medical, history, reminders, financial).</div>

    <form method="get" class="form" style="margin-top:10px;">
      <div>
        <label>Criteria</label>
        <select name="crit">
          <option value="any" {% if (crit or 'any')=='any' %}selected{% endif %}>Any Field</option>
          <option value="owner_name" {% if crit=='owner_name' %}selected{% endif %}>Owner Name</option>
          <option value="phone" {% if crit=='phone' %}selected{% endif %}>Owner Phone</option>
          <option value="pet" {% if crit=='pet' %}selected{% endif %}>Pet Name</option>
          <option value="pet_id" {% if crit=='pet_id' %}selected{% endif %}>Pet ID / Microchip</option>
        </select>
      </div>
      <div class="full">
        <label>Search</label>
        <input name="q" value="{{ q or '' }}" placeholder="Type owner / phone / pet / id">
      </div>
      <div class="full row-actions">
        <button class="btn primary" type="submit">Search</button>
        <a class="btn" href="{{ url_for('reminders') }}">Clear</a>
      </div>
    </form>

    {% if q and not pet_matches %}
      <div class="hr"></div>
      <div class="muted">No matching pets found for: <b>{{ q }}</b></div>
    {% endif %}

    {% if pet_matches %}
      <div class="hr"></div>
      <form method="get" class="form">
        <input type="hidden" name="q" value="{{ q or '' }}">
        <input type="hidden" name="crit" value="{{ crit or 'any' }}">
        <div class="full">
          <label>Select Pet</label>
          <select name="pet_id" onchange="this.form.submit()">
            <option value="">Choose...</option>
            {% for p in pet_matches %}
              <option value="{{ p.pet_id }}" {% if selected_pet_id==p.pet_id %}selected{% endif %}>{{ p.label }}</option>
            {% endfor %}
          </select>
          <div class="small muted" style="margin-top:8px">
            {% if not selected_pet_id %}
              Select a pet to show the full profile.
            {% endif %}
          </div>
        </div>
      </form>
    {% endif %}

    {% if pet360 %}
      <div class="hr"></div>

      <div id="petProfile" class="grid two">
        <div class="card">
          <h3>Medical Profile</h3>
          <div class="muted">
            <div><b>Pet:</b> {{ pet360.pet_name }} • {{ pet360.species or '—' }} • {{ pet360.breed or '—' }}</div>
            <div><b>Pet ID:</b> {{ pet360.pet_id }} • <b>Microchip:</b> {{ pet360.microchip_id or '—' }}</div>
            <div><b>Sex:</b> {{ pet360.sex or '—' }} • <b>Age:</b> {{ pet360.age or '—' }} • <b>Weight:</b> {{ pet360.weight or '—' }}</div>

            <div class="hr"></div>

            <div><b>Allergies:</b> {{ pet360.allergies or '—' }}</div>
            <div><b>Chronic Conditions:</b> {{ pet360.chronic_conditions or '—' }}</div>
            <div><b>Vaccinations Summary:</b> {{ pet360.vaccinations_summary or '—' }}</div>
            <div><b>Notes:</b> {{ pet360.pet_notes or '—' }}</div>

            <div class="hr"></div>

            <div class="muted small"><b>Latest Visit Snapshot</b></div>
            <div><b>Last Visit:</b> {{ pet360.last_visit or '—' }}</div>
            <div><b>Last Diagnosis:</b> {{ pet360.last_diagnosis or '—' }}</div>
            <div><b>Last Treatment:</b> {{ pet360.last_treatment_plan or '—' }}</div>
            <div><b>Last Vaccines:</b> {{ pet360.last_vaccines or '—' }}</div>
            <div><b>Follow-up:</b> {{ pet360.last_followup or '—' }}</div>
          </div>

          <div class="hr"></div>

          <div class="muted">
            <div><b>Owner:</b> {{ pet360.owner_name }} ({{ pet360.owner_phone or '—' }})</div>
            <div><b>Email:</b> {{ pet360.owner_email or '—' }}</div>
            <div><b>Address:</b> {{ pet360.owner_address or '—' }}</div>
          </div>
        </div>

        <div class="card">
          <h3>Financial Status</h3>
          <div class="muted">
            <div><b>Total (incl. VAT):</b> {{ pet_fin.total }}</div>
            <div><b>Paid:</b> {{ pet_fin.paid }}</div>
            <div><b>Outstanding (Due):</b> {{ pet_fin.due }}</div>
            <div class="hr"></div>
            <div><b>Visits/Bookings:</b> {{ pet_fin.bookings_count }} • <b>Unpaid/Partial:</b> {{ pet_fin.open_count }}</div>
          </div>

          {% if pet_fin.open_items %}
            <div class="hr"></div>
            <div class="muted small"><b>Outstanding Items</b></div>
            <div style="overflow:auto;margin-top:10px">
              <table>
                <thead><tr><th>Date</th><th>Invoice</th><th>Status</th><th style="text-align:right">Due</th><th>Action</th></tr></thead>
                <tbody>
                  {% for x in pet_fin.open_items %}
                    <tr>
                      <td><b>{{ x.appointment_start }}</b></td>
                      <td>{{ x.invoice_no or '—' }}</td>
                      <td>{{ x.payment_status or '—' }}</td>
                      <td style="text-align:right"><b>{{ x.due }}</b></td>
                      <td><a class="btn primary" href="{{ url_for('booking_view', booking_id=x.booking_id) }}">Open</a></td>
                    </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
          {% endif %}

          <div class="hr"></div>

          <details>
            <summary><b>All Invoices / Payments</b> ({{ pet_history|length }})</summary>
            <div style="overflow:auto;margin-top:10px">
              <table>
                <thead><tr><th>Date</th><th>Invoice</th><th>Payment</th><th style="text-align:right">Total</th><th style="text-align:right">Paid</th><th style="text-align:right">Due</th><th>Action</th></tr></thead>
                <tbody>
                  {% for h in pet_history %}
                    <tr>
                      <td><b>{{ h.appointment_start }}</b></td>
                      <td>{{ h.invoice_no or '—' }}</td>
                      <td>{{ h.payment_status or '—' }}</td>
                      <td style="text-align:right">{{ h.total }}</td>
                      <td style="text-align:right">{{ h.paid }}</td>
                      <td style="text-align:right"><b>{{ h.due }}</b></td>
                      <td><a class="btn primary" href="{{ url_for('booking_view', booking_id=h.booking_id) }}">Open</a></td>
                    </tr>
                  {% endfor %}
                  {% if not pet_history %}
                    <tr><td colspan="7" class="muted">No invoices/visits found yet.</td></tr>
                  {% endif %}
                </tbody>
              </table>
            </div>
          </details>
        </div>
      </div>

      <div class="hr"></div>

      <details open>
        <summary><b>Upcoming Reminders</b> ({{ pet_reminders_upcoming|length }})</summary>
        <div style="overflow:auto;margin-top:10px">
          <table>
            <thead><tr><th>Scheduled</th><th>Service</th><th>Type</th><th>Status</th><th>Message</th><th>Action</th></tr></thead>
            <tbody>
              {% for r in pet_reminders_upcoming %}
                <tr>
                  <td><b>{{ r.scheduled_for }}</b></td>
                  <td>{{ r.service_name }}</td>
                  <td>{{ r.reminder_type }}</td>
                  <td><span class="pill {% if r.status=='Sent' %}good{% elif r.status in ['Opened','Scheduled'] %}warn{% endif %}">{{ r.status }}</span></td>
                  <td class="small">{{ (r.message or '')[:80] }}{% if (r.message or '')|length > 80 %}...{% endif %}</td>
                  <td style="white-space:nowrap">
                    <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                    <a class="btn good" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a>
                  </td>
                </tr>
              {% endfor %}
              {% if not pet_reminders_upcoming %}
                <tr><td colspan="6" class="muted">No upcoming reminders.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </details>

      <div class="hr"></div>

      <details>
        <summary><b>All Reminders</b> ({{ pet_reminders_all|length }})</summary>
        <div style="overflow:auto;margin-top:10px">
          <table>
            <thead><tr><th>Scheduled</th><th>Service</th><th>Type</th><th>Status</th><th>Channel</th><th>Message</th><th>Action</th></tr></thead>
            <tbody>
              {% for r in pet_reminders_all %}
                <tr>
                  <td><b>{{ r.scheduled_for }}</b></td>
                  <td>{{ r.service_name }}</td>
                  <td>{{ r.reminder_type }}</td>
                  <td><span class="pill {% if r.status=='Sent' %}good{% elif r.status in ['Opened','Scheduled'] %}warn{% endif %}">{{ r.status }}</span></td>
                  <td>{{ r.channel }}</td>
                  <td class="small">{{ (r.message or '')[:80] }}{% if (r.message or '')|length > 80 %}...{% endif %}</td>
                  <td style="white-space:nowrap">
                    <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                    <a class="btn good" href="{{ r.whatsapp_link }}" target="_blank">WhatsApp</a>
                    <a class="btn" href="{{ url_for('reminder_mark_sent', reminder_id=r.id) }}">Mark Sent</a>
                  </td>
                </tr>
              {% endfor %}
              {% if not pet_reminders_all %}
                <tr><td colspan="7" class="muted">No reminders found for this pet.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </details>

      <div class="hr"></div>

      <details>
        <summary><b>Medical History</b> ({{ pet_history|length }} visits)</summary>
        {% for h in pet_history %}
          <div class="hr"></div>
          <details>
            <summary>
              <b>{{ h.appointment_start }}</b> — {{ h.appointment_type }} • {{ h.status }}
              &nbsp;|&nbsp; Vet: {{ h.vet_name or '—' }}
              &nbsp;|&nbsp; Total: {{ h.total }} • Due: {{ h.due }}
            </summary>
            <div style="margin-top:10px" class="muted">
              <div><b>Services:</b> {{ h.services_summary or '—' }}</div>
              <div><b>Reason:</b> {{ h.reason or '—' }}</div>
              <div><b>Symptoms:</b> {{ h.symptoms or '—' }}</div>
              <div class="hr"></div>
              <div><b>Diagnosis:</b> {{ h.diagnosis or '—' }}</div>
              <div><b>Treatment Plan:</b> {{ h.treatment_plan or '—' }}</div>
              <div><b>Prescription:</b> {{ h.prescription or '—' }}</div>
              <div><b>Lab Tests:</b> {{ h.lab_tests or '—' }}</div>
              <div><b>Vaccines Given:</b> {{ h.vaccines_given or '—' }}</div>
              <div><b>Follow-up:</b> {{ h.followup_datetime or '—' }}</div>
              <div class="hr"></div>
              <div><b>Invoice:</b> {{ h.invoice_no or '—' }} • <b>Payment:</b> {{ h.payment_status or '—' }} • <b>Paid:</b> {{ h.paid }}</div>
              <div><b>Notes:</b> {{ h.notes or '—' }}</div>
              <div class="row-actions" style="margin-top:10px">
                <a class="btn primary" href="{{ url_for('booking_view', booking_id=h.booking_id) }}">Open Visit</a>
              </div>
            </div>
          </details>
        {% endfor %}
        {% if not pet_history %}
          <div class="hr"></div>
          <div class="muted">No visits found for this pet yet.</div>
        {% endif %}
      </details>
    {% endif %}
  </div>

  <div class="hr"></div>

  <!-- Enhanced Reminder Controls -->
  <div class="grid two">
    <div class="card">
      <h3>Auto-Schedule Reminders</h3>
      <div class="muted">Creates “Scheduled” reminders for upcoming bookings. No duplicates for same booking/time.</div>

      <form method="post" action="{{ url_for('reminders_auto_schedule') }}" class="form" style="margin-top:10px;">
        <div>
          <label>Hours Before Appointment</label>
          <select name="hours_before">
            <option value="24">24 hours</option>
            <option value="6">6 hours</option>
            <option value="2" selected>2 hours</option>
            <option value="1">1 hour</option>
            <option value="0">At time</option>
          </select>
        </div>
        <div>
          <label>Days Ahead</label>
          <select name="days_ahead">
            <option value="1">1 day</option>
            <option value="3">3 days</option>
            <option value="7" selected>7 days</option>
            <option value="14">14 days</option>
          </select>
        </div>
        <div class="full row-actions">
          <button class="btn warn" type="submit">Auto-Schedule</button>
          <a class="btn" href="{{ url_for('bookings') }}">Back to Bookings</a>
        </div>
      </form>

      <div class="hr"></div>

      {% set ns = namespace(total=0, scheduled=0, opened=0, sent=0, due=0) %}
      {% for r in reminder_rows %}
        {% set ns.total = ns.total + 1 %}
        {% if r.status == 'Scheduled' %}{% set ns.scheduled = ns.scheduled + 1 %}{% endif %}
        {% if r.status == 'Opened' %}{% set ns.opened = ns.opened + 1 %}{% endif %}
        {% if r.status == 'Sent' %}{% set ns.sent = ns.sent + 1 %}{% endif %}
        {% if r.scheduled_for and r.scheduled_for <= now_key and r.status != 'Sent' %}
          {% set ns.due = ns.due + 1 %}
        {% endif %}
      {% endfor %}

      <div class="muted">
        Total Records: <b>{{ ns.total }}</b><br>
        Scheduled: <b>{{ ns.scheduled }}</b> • Due Now: <b>{{ ns.due }}</b><br>
        Opened: <b>{{ ns.opened }}</b> • Sent: <b>{{ ns.sent }}</b>
      </div>
    </div>

    <div class="card">
      <h3>Due Now (Based on Scheduled Time)</h3>
      <div class="muted">Reminders where scheduled_for ≤ now and not sent.</div>

      <div style="overflow:auto;margin-top:10px">
        <table>
          <thead><tr><th>Scheduled</th><th>Pet</th><th>Status</th><th>Action</th></tr></thead>
          <tbody>
            {% set found = false %}
            {% for r in reminder_rows %}
              {% if r.scheduled_for and r.scheduled_for <= now_key and r.status != 'Sent' %}
                {% set found = true %}
                <tr>
                  <td><b>{{ r.scheduled_for }}</b></td>
                  <td>{{ r.pet_name }}</td>
                  <td><span class="pill warn">{{ r.status }}</span></td>
                  <td style="white-space:nowrap">
                    <a class="btn good" href="{{ url_for('reminder_open', reminder_id=r.id) }}" target="_blank">WhatsApp</a>
                    <a class="btn" href="{{ url_for('reminder_mark_sent', reminder_id=r.id) }}">Mark Sent</a>
                    <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                  </td>
                </tr>
              {% endif %}
            {% endfor %}
            {% if not found %}
              <tr><td colspan="4" class="muted">No due reminders right now.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <div class="hr"></div>

  <div class="grid two">
    <div class="card">
      <h3>Upcoming in next 24 hours</h3>
      <div style="overflow:auto;margin-top:10px">
        <table>
          <thead><tr><th>Time</th><th>Pet</th><th>Owner</th><th>Action</th></tr></thead>
          <tbody>
            {% for b in upcoming %}
              <tr>
                <td><b>{{ b.appointment_start }}</b></td>
                <td>{{ b.pet_name }}</td>
                <td>{{ b.owner_name }}</td>
                <td style="white-space:nowrap">
                  <a class="btn primary" href="{{ url_for('booking_view', booking_id=b.id) }}">Open</a>
                  <a class="btn good" href="{{ url_for('booking_remind', booking_id=b.id) }}" target="_blank">WhatsApp</a>
                  <a class="btn" href="{{ b.portal_link }}" target="_blank">Portal</a>
                  <a class="btn" href="{{ b.ics_link }}">.ics</a>
                  <a class="btn" href="{{ url_for('invoice_print', booking_id=b.id) }}" target="_blank">Invoice</a>
                </td>
              </tr>
            {% endfor %}
            {% if not upcoming %}
              <tr><td colspan="4" class="muted">No upcoming appointments.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <h3>Reminder Records</h3>
      <div style="overflow:auto;margin-top:10px">
        <table>
          <thead><tr><th>Scheduled</th><th>Pet</th><th>Status</th><th>Action</th></tr></thead>
          <tbody>
            {% for r in reminder_rows %}
              <tr>
                <td><b>{{ r.scheduled_for }}</b></td>
                <td>{{ r.pet_name }}</td>
                <td>
                  <span class="pill {% if r.status=='Sent' %}good{% elif r.status in ['Opened','Scheduled'] %}warn{% endif %}">{{ r.status }}</span>
                </td>
                <td style="white-space:nowrap">
                  <a class="btn primary" href="{{ url_for('booking_view', booking_id=r.booking_id) }}">Booking</a>
                  <a class="btn good" href="{{ url_for('reminder_open', reminder_id=r.id) }}" target="_blank">WhatsApp</a>
                  <a class="btn" href="{{ url_for('reminder_mark_sent', reminder_id=r.id) }}">Mark Sent</a>
                </td>
              </tr>
            {% endfor %}
            {% if not reminder_rows %}
              <tr><td colspan="4" class="muted">No reminder records yet.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

    "smoother_booking.html": r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Smoother Exam (Fast Wizard)</h3>
  <div class="form">
    <input type="hidden" id="vat_rate" value="{{ vat_rate }}">
<div class="full row-actions">
<span class="small">Offline AI. Decision support only.</span>
    </div>
  </div>
<div class="hr"></div>
<form method="post" class="form" id="smoothForm">
    <div>
      <label>Appointment Start</label>
      <input name="appointment_start" id="appointment_start" value="{{ default_start }}" placeholder="YYYY-MM-DD HH:MM" required>
    </div>

    <div>
      <label>Duration (min)</label>
      <input name="duration_min" id="duration_min" value="30" required>
    </div>

    <div>
      <label>Owner</label>
      <select name="owner_id" id="owner_id" required onchange="filterPets()">
        <option value="">Select...</option>
        {% for o in owners %}
          <option value="{{ o.id }}">{{ o.owner_name }} ({{ o.phone }})</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Pet</label>
      <select name="pet_id" id="pet_id" required>
        <option value="">Select...</option>
        {% for p in pets %}
          <option value="{{ p.id }}" data-owner="{{ p.owner_id }}">{{ p.pet_name }} ({{ p.species }})</option>
        {% endfor %}
      </select>
      <div class="small" style="margin-top:8px;display:flex;gap:8px;flex-wrap:wrap"><button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetHistory(document.getElementById('pet_id').value)">History</button><button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetReminders(document.getElementById('pet_id').value)">Remind</button></div>

      <div class="small">Pets list filters automatically when you choose an owner.</div>
    </div>


    <div>
      <label>Weight (kg)</label>
      <input name="visit_weight_kg" id="visit_weight_kg" value="{{ booking.visit_weight_kg if booking else '' }}" placeholder="e.g. 5.40" required>
    </div>

    <div>
      <label>Temperature (°C)</label>
      <input name="visit_temp_c" id="visit_temp_c" value="{{ booking.visit_temp_c if booking else '' }}" placeholder="e.g. 38.5" required>
    </div>

    <div>
      <label>Appointment Type</label>
      <select name="appointment_type" id="appointment_type">
        {% for t in types %}
          <option>{{ t }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Priority</label>
      <select name="priority" id="priority">
        {% for p in priorities %}
          <option>{{ p }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Status</label>
      <select name="status" id="status">
        {% for s in statuses %}
          <option {% if s=='Scheduled' %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Channel</label>
      <select name="channel" id="channel">
        {% for c in channels %}
          <option {% if c=='WhatsApp' %}selected{% endif %}>{{ c }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label>Reason</label>
      <textarea name="reason" id="reason" placeholder="Main reason / request" style="min-height:44px;height:44px;resize:none;overflow:hidden"></textarea>
    </div>

    <div class="full">
      <label>Symptoms</label>
      <input name="symptoms" id="symptoms" placeholder="Optional symptoms">
    </div>

    <div>
      <label>Vet Name</label>
      <select name="vet_name" required>
        {% for v in vets %}
          <option value="{{ v }}" {% if (default_vet|lower)==(v|lower) %}selected{% endif %}>{{ v }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Room</label>
      <select name="room" required>
        {% for r in rooms %}
          <option value="{{ r }}">{{ r }}</option>
        {% endfor %}
      </select>
    </div>



    <div>
      <label>Service</label>
      <select name="service_name" id="service_name" onchange="syncServiceFee()">
        {% for s in services %}
          <option value="{{ s.name }}" data-fee="{{ s.fee }}">{{ s.name }} ({{ s.fee }})</option>
        {% endfor %}
      </select>
    </div>
<input type="hidden" name="services_json" id="services_json" value="">
<input type="hidden" name="service_name" id="service_name" value="">

<div class="full">
  <label>Services</label>
  <div class="small muted" style="margin-top:-6px;margin-bottom:10px;">
    Select services (up to 10). Fee will auto-fill. A new row appears when you select a service.
  </div>

  <div style="overflow:auto;border:1px solid var(--line);border-radius:12px;">
    <table id="svcTable" class="tbl" style="width:100%;border-collapse:collapse;">
      <thead>
        <tr style="background:var(--soft);">
          <th style="text-align:left;padding:10px;">Service</th>
          <th style="text-align:left;padding:10px;width:140px;">Fee</th>
          <th style="text-align:left;padding:10px;width:110px;">Qty</th>
          <th style="text-align:left;padding:10px;width:150px;">Line Total</th>
          <th style="padding:10px;width:90px;"></th>
        </tr>
      </thead>
      <tbody id="svcBody"></tbody>
    </table>
  </div>
  <div style="margin-top:10px;display:flex;gap:10px;align-items:center;">
    <button type="button" class="btn" id="addSvcBtn">+ Add service</button>
    <div class="muted small">Subtotal is calculated automatically.</div>
  </div>

  <template id="svcRowTpl">
    <tr class="svcRow">
      <td style="padding:10px;">
        <select class="svcSel">
          <option value="">Select service…</option>
          {% for s in services %}
            <option value="{{ s.name }}" data-fee="{{ s.fee }}">{{ s.name }} ({{ s.fee }})</option>
          {% endfor %}
        </select>
      </td>
      <td style="padding:10px;">
        <input type="number" step="0.01" min="0" class="svcFee" placeholder="0.00">
      </td>
      <td style="padding:10px;">
        <input type="number" min="1" value="1" class="svcQty">
      </td>
      <td style="padding:10px;">
        <span class="svcLine">0.00</span>
      </td>
      <td style="padding:10px;">
        <button type="button" class="btn bad svcRemove" title="Remove">✕</button>
      </td>
    </tr>
  </template>
</div>

<div>
  <label>Subtotal (auto)</label>
  <input name="service_fee" id="service_fee" placeholder="Subtotal" value="" readonly>
</div>

<div>
  <label>Paid</label>
  <input name="paid_amount" id="paid_amount" placeholder="Paid amount" value="" oninput="recalcTotals()">
</div>

<div>
  <label>Due (auto)</label>
  <input name="due_amount" id="due_amount" placeholder="Due amount" value="" readonly>
</div>
<div>
      <label>Reminder Channel</label>
      <select name="reminder_channel">
        {% for rc in reminder_channels %}
          <option {% if rc=='WhatsApp' %}selected{% endif %}>{{ rc }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label>Notes</label>
      <textarea name="notes"></textarea>
    </div>

    <div class="full">
      <label style="display:flex;gap:10px;align-items:center">
        <input type="checkbox" name="open_wa" style="width:auto;transform:scale(1.1)">
        Open WhatsApp Journey after create
      </label>
      <div class="small">If enabled, the system will create the booking then open WhatsApp with the reminder message.</div>
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">Create (Smoother)</button>
      <a class="btn" href="{{ url_for('bookings') }}">Back</a>
    </div>
  </form>
</div>

<script>

  function _num(v){
    try{
      const x = parseFloat(String(v||"").replace(/[^0-9.\-]/g,""));
      return isNaN(x)?0:x;
    }catch(e){ return 0; }
  }

  function syncServiceFee(){
    const sel = document.getElementById("service_name");
    const feeEl = document.getElementById("service_fee");
    if(!sel || !feeEl) return;
    const opt = sel.options[sel.selectedIndex];
    const fee = opt ? opt.getAttribute("data-fee") : "";
    if(!feeEl.value){
      feeEl.value = fee || "";
    }
    calcDue();
  }

  function calcDue(){
    const fee = _num(document.getElementById("service_fee")?.value);
    const paid = _num(document.getElementById("paid_amount")?.value);
    const vatRate = _num(document.getElementById("vat_rate")?.value);
    const total = fee + (fee * vatRate);
    const due = total - paid;
    const dueEl = document.getElementById("due_amount");
    if(dueEl) dueEl.value = (Math.max(due,0)).toFixed(2);
  }

  window.addEventListener("load", ()=>{
    syncServiceFee();
    calcDue();
  });


  // Reason field: auto-expand after 15 words (keeps compact height for short text)
  function autoExpandReason(){
    const el = document.getElementById("reason");
    if (!el) return;
    const words = (el.value || "").trim().match(/\S+/g);
    const wc = words ? words.length : 0;

    if (wc <= 15){
      el.style.minHeight = "44px";
      el.style.height = "44px";
      el.style.overflow = "hidden";
      el.style.resize = "none";
      return;
    }

    el.style.resize = "vertical";
    el.style.overflow = "hidden";
    el.style.minHeight = "92px";
    el.style.height = "auto";
    el.style.height = Math.max(el.scrollHeight, 92) + "px";
  }

  window.addEventListener("load", ()=>{
    const el = document.getElementById("reason");
    if (el){
      el.addEventListener("input", autoExpandReason);
      autoExpandReason();
    }
  });

  function filterPets(){
    const ownerId = document.getElementById("owner_id").value;
    const petSel = document.getElementById("pet_id");
    const opts = petSel.querySelectorAll("option");
    let firstVisible = "";
    opts.forEach((o, idx)=>{
      if (idx===0) return;
      const ok = !ownerId || (o.dataset.owner === ownerId);
      o.style.display = ok ? "block" : "none";
      if (ok && !firstVisible) firstVisible = o.value;
    });
    const cur = petSel.value;
    const curOpt = petSel.querySelector(`option[value="${cur}"]`);
    if (curOpt && curOpt.style.display === "none"){
      petSel.value = firstVisible || "";
    }
  }
  filterPets();
    const res = await fetch("{{ url_for('api_intake') }}", {
      method:"POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({text:t})
    });
    const data = await res.json();
    if (data.error){
      alert(data.error);
      return;
    }
    document.getElementById("appointment_type").value = data.appointment_type || "Consultation";
    document.getElementById("priority").value = data.priority || "Normal";
    document.getElementById("reason").value = data.reason || "";
    autoExpandReason();
    document.getElementById("symptoms").value = data.symptoms || "";
  }

// ========= Services table (multi-line) =========
const svcBody = document.getElementById("svcBody");
const rowTpl = document.getElementById("svcRowTpl");
const addBtn = document.getElementById("addSvcBtn");
const servicesJsonEl = document.getElementById("services_json");
const subtotalEl = document.getElementById("service_fee");
const paidEl = document.getElementById("paid_amount");
const dueEl = document.getElementById("due_amount");

const MAX_SERVICES_ROWS = 10;

function money(v, d=0){
  if (v === null || v === undefined) return d;
  if (typeof v === "number") return isFinite(v) ? v : d;
  const s = String(v).replace(/,/g,"").trim();
  if (!s) return d;
  const x = parseFloat(s);
  return isFinite(x) ? x : d;
}

function buildServicesPayload(){
  const rows = Array.from(svcBody.querySelectorAll("tr.svcRow"));
  const out = [];
  rows.forEach(r=>{
    const sel = r.querySelector(".svcSel");
    const feeEl = r.querySelector(".svcFee");
    const qtyEl = r.querySelector(".svcQty");
    const name = sel ? sel.value : "";
    if (!name) return;
    const fee = money(feeEl ? feeEl.value : 0);
    let qty = 1;
    try{ qty = Math.max(1, parseInt(qtyEl.value || "1")); }catch(e){ qty = 1; }
    const line_total = round2(fee * qty);
    out.push({name, fee: round2(fee), qty, line_total});
  });
  return out;
}

function round2(x){ return Math.round((money(x)*100))/100; }

function recalcTotals(){
  if (!svcBody || !subtotalEl) return;
  const payload = buildServicesPayload();
  if (servicesJsonEl) servicesJsonEl.value = JSON.stringify(payload);
  // best-effort for legacy single service_name
  if (document.getElementById("service_name")){
    const sn = document.getElementById("service_name");
    if (payload.length === 0) sn.value = "";
    else {
      const first = payload[0].name;
      sn.value = (payload.length===1) ? first : (first + " +" + (payload.length-1));
    }
  }
  const subtotal = round2(payload.reduce((s,it)=> s + money(it.line_total), 0));
  const paid = money(paidEl ? paidEl.value : 0);
  const due = round2(subtotal - paid);
  subtotalEl.value = subtotal.toFixed(2);
  if (dueEl) dueEl.value = due.toFixed(2);
}

// Backward compatibility
function syncServiceFee(){ recalcTotals(); }
function calcDue(){ recalcTotals(); }

function calcLine(node){
  const feeEl = node.querySelector(".svcFee");
  const qtyEl = node.querySelector(".svcQty");
  const lineEl = node.querySelector(".svcLine");
  const fee = money(feeEl ? feeEl.value : 0);
  let qty = 1;
  try{ qty = Math.max(1, parseInt(qtyEl.value || "1")); }catch(e){ qty = 1; }
  if (lineEl) lineEl.textContent = round2(fee * qty).toFixed(2);
}

function addRow(pref){
  if (!svcBody || !rowTpl) return null;
  const current = svcBody.querySelectorAll("tr.svcRow").length;
  if (current >= MAX_SERVICES_ROWS) return null;

  const node = rowTpl.content.firstElementChild.cloneNode(true);
  const sel = node.querySelector(".svcSel");
  const feeEl = node.querySelector(".svcFee");
  const qtyEl = node.querySelector(".svcQty");
  const rm = node.querySelector(".svcRemove");

  if (pref){
    if (pref.name && sel){
      sel.value = pref.name;
      const opt = sel.options[sel.selectedIndex];
      if (opt && opt.dataset && opt.dataset.fee && (!pref.fee || pref.fee==0)){
        feeEl.value = opt.dataset.fee;
      } else if (pref.fee !== undefined){
        feeEl.value = pref.fee;
      }
    }
    if (pref.fee !== undefined && feeEl && !feeEl.value) feeEl.value = pref.fee;
    if (pref.qty && qtyEl) qtyEl.value = pref.qty;
  }

  function onChange(){
    const opt = sel.options[sel.selectedIndex];
    if (opt && opt.dataset && opt.dataset.fee && (!feeEl.value || feeEl.value==="0" || feeEl.value==="0.00")){
      feeEl.value = opt.dataset.fee;
    }
    calcLine(node);

    const rows = Array.from(svcBody.querySelectorAll("tr.svcRow"));
    const isLast = rows.length ? (rows[rows.length - 1] === node) : true;
    if (isLast && sel.value){
      const countNow = svcBody.querySelectorAll("tr.svcRow").length;
      if (countNow < MAX_SERVICES_ROWS) addRow(); // add trailing empty
    }
    recalcTotals();
  }

  sel.addEventListener("change", onChange);
  feeEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });
  qtyEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });

  rm.addEventListener("click", ()=>{
    node.remove();
    if (svcBody.querySelectorAll("tr.svcRow").length === 0) addRow();
    recalcTotals();
  });

  svcBody.appendChild(node);
  calcLine(node);
  recalcTotals();
  return node;
}

if (addBtn){
  addBtn.addEventListener("click", ()=> addRow());
}

(function initServices(){
  // If page contains pre-loaded services_json, load it. Else start empty row.
  let existing = [];
  try{
    const raw = (servicesJsonEl && servicesJsonEl.value) ? servicesJsonEl.value : "";
    if (raw) existing = JSON.parse(raw);
  } catch(e){ existing = []; }

  if (existing && existing.length){
    existing.forEach(it=> addRow(it));
    addRow(); // trailing
  } else {
    addRow();
  }
  recalcTotals();
})();
</script>
{% endblock %}
""",

    "invoice_print.html": r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{{ app_title }} | Invoice {{ invoice_no }}</title>
  <style>
    :root{ --ink:#111; --muted:#666; --line:#e6e6e6; --bg:#ffffff; --shade:#f7f7f7; }
    *{ box-sizing:border-box; }
    body{ margin:0; font-family: Arial, Helvetica, sans-serif; color:var(--ink); background:#f2f3f5; }
    .wrap{ max-width:920px; margin:24px auto; padding:0 16px; }
    .sheet{ background:var(--bg); border:1px solid var(--line); border-radius:12px; overflow:hidden; box-shadow:0 6px 18px rgba(0,0,0,.06); }
    .head{ padding:22px 24px; border-bottom:1px solid var(--line); display:flex; gap:18px; align-items:flex-start; justify-content:space-between; }
    .clinic h1{ margin:0; font-size:20px; letter-spacing:.2px; }
    .clinic .meta{ margin-top:6px; font-size:12.5px; color:var(--muted); line-height:1.5; }
    .inv{ text-align:right; }
    .inv .no{ font-size:18px; font-weight:700; }
    .inv .small{ margin-top:6px; font-size:12.5px; color:var(--muted); line-height:1.5; }
    .bar{ padding:14px 24px; background:var(--shade); border-bottom:1px solid var(--line); display:flex; flex-wrap:wrap; gap:14px 24px; }
    .kv{ min-width:190px; }
    .k{ font-size:12px; color:var(--muted); margin-bottom:4px; }
    .v{ font-size:13.5px; font-weight:600; }
    .body{ padding:18px 24px 24px; }
    table{ width:100%; border-collapse:collapse; }
    th,td{ padding:10px 10px; border-bottom:1px solid var(--line); vertical-align:top; }
    th{ background:#fff; font-size:12px; text-transform:uppercase; letter-spacing:.08em; color:var(--muted); text-align:left; }
    td.right, th.right{ text-align:right; }
    .muted{ color:var(--muted); }
    .totals{ margin-top:18px; display:flex; gap:18px; justify-content:flex-end; }
    .totbox{ width:340px; border:1px solid var(--line); border-radius:10px; overflow:hidden; }
    .totbox .row{ display:flex; justify-content:space-between; padding:10px 12px; border-bottom:1px solid var(--line); }
    .totbox .row:last-child{ border-bottom:none; }
    .totbox .row.total{ background:var(--shade); font-weight:800; }
    .totbox .row.due{ background:#fff0f0; font-weight:800; }
    .foot{ padding:16px 24px; border-top:1px solid var(--line); display:flex; align-items:center; justify-content:space-between; gap:12px; flex-wrap:wrap; }
    .actions a, .actions button{
      display:inline-block; border:1px solid var(--line); background:#fff; padding:9px 12px; border-radius:10px;
      text-decoration:none; color:var(--ink); font-weight:700; cursor:pointer;
    }
    .actions .primary{ background:#111; color:#fff; border-color:#111; }
    @media print{
      body{ background:#fff; }
      .wrap{ max-width:100%; margin:0; padding:0; }
      .sheet{ border:none; border-radius:0; box-shadow:none; }
      .actions{ display:none !important; }
    }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="sheet">

      <div class="head">
        <div class="clinic">
          <h1>{{ clinic.name }}</h1>
          <div class="meta">
            {% if clinic.address %}{{ clinic.address }}<br>{% endif %}
            {% if clinic.phone %}{{ clinic.phone }}{% endif %}
            {% if clinic.email %}{% if clinic.phone %} | {% endif %}{{ clinic.email }}{% endif %}
          </div>
        </div>
        <div class="inv">
          <div class="no">INVOICE #{{ invoice_no }}</div>
          <div class="small">
            Exam ID: {{ booking.id }}<br>
            Appointment: {{ appt_start }}{% if appt_end %} – {{ appt_end }}{% endif %}
          </div>
        </div>
      </div>

      <div class="bar">
        <div class="kv">
          <div class="k">Owner</div>
          <div class="v">{{ owner.owner_name if owner else '-' }}</div>
          <div class="muted" style="font-size:12px;margin-top:2px;">
            {% if owner and owner.phone %}{{ owner.phone }}{% endif %}
            {% if owner and owner.email %}{% if owner.phone %} | {% endif %}{{ owner.email }}{% endif %}
          </div>
        </div>
        <div class="kv">
          <div class="k">Pet</div>
          <div class="v">{{ pet.pet_name if pet else '-' }}</div>
          <div class="muted" style="font-size:12px;margin-top:2px;">
            {% if pet and pet.species %}{{ pet.species }}{% endif %}
            {% if pet and pet.breed %}{% if pet.species %} | {% endif %}{{ pet.breed }}{% endif %}
          </div>
        </div>
        <div class="kv">
          <div class="k">Vet</div>
          <div class="v">{{ booking.vet_name or '-' }}</div>
        </div>
        <div class="kv">
          <div class="k">Room</div>
          <div class="v">{{ booking.room or '-' }}</div>
        </div>
        <div class="kv">
          <div class="k">Status</div>
          <div class="v">{{ booking.status or 'Scheduled' }}</div>
        </div>
      </div>

      <div class="body">
        <table>
          <thead>
            <tr>
              <th style="width:52%;">Service</th>
              <th class="right" style="width:12%;">Qty</th>
              <th class="right" style="width:18%;">Unit</th>
              <th class="right" style="width:18%;">Line Total</th>
            </tr>
          </thead>
          <tbody>
            {% if line_items and line_items|length > 0 %}
              {% for it in line_items %}
                <tr>
                  <td>
                    <b>{{ it.desc }}</b>
                    {% if it.details %}<div class="muted" style="font-size:12px;margin-top:2px;">{{ it.details }}</div>{% endif %}
                  </td>
                  <td class="right">{{ it.qty }}</td>
                  <td class="right">{{ "%.2f"|format(it.unit_price) }}</td>
                  <td class="right"><b>{{ "%.2f"|format(it.line_total) }}</b></td>
                </tr>
              {% endfor %}
            {% else %}
              <tr>
                <td colspan="4" class="muted">No services recorded.</td>
              </tr>
            {% endif %}
          </tbody>
        </table>

        <div class="totals">
          <div class="totbox">
            <div class="row"><span class="muted">Subtotal (Services)</span><span>{{ "%.2f"|format(fee) }}</span></div>
            <div class="row"><span class="muted">VAT ({{ (vat_rate*100)|round(0) }}%)</span><span>{{ "%.2f"|format(vat) }}</span></div>
            <div class="row total"><span>Total</span><span>{{ "%.2f"|format(total) }}</span></div>
            <div class="row"><span class="muted">Paid</span><span>{{ "%.2f"|format(paid) }}</span></div>
            <div class="row due"><span>Due</span><span>{{ "%.2f"|format(due) }}</span></div>
          </div>
        </div>

      </div>

      <div class="foot">
        <div class="muted" style="font-size:12px;">
          Thank you. This invoice is generated by the clinic system.
        </div>
        <div class="actions">
          <a href="{{ url_for('booking_details', booking_id=booking.id) }}">Back to Exam</a>
          <button class="primary" onclick="window.print()">Print</button>
        </div>
      </div>

    </div>
  </div>
</body>
</html>
"""
})

# =====================================================================
# END OF ADD-ON PACK
# =====================================================================


# =====================================================================
# EASY BOOKING WIZARD ADD-ON (NO DELETION OF YOUR EXISTING FEATURES)
# - Easy Booking: New Customer (Owner -> Pet -> Booking -> Reminder -> Invoice)
# - Old Customer: Search by Phone -> Select/Add Pet -> Booking -> Reminder -> Invoice
# - Auto Serial: TZ-00001, TZ-00002, ... stored in owners.xlsx as "customer_sn"
#
# Paste this block BEFORE your final STARTUP section (before write_templates/init_storage/app.run)
# =====================================================================

# -------------------------
# Wizard / Serial Settings
# -------------------------
EASY_OWNER_SN_PREFIX = "TZ"
EASY_OWNER_SN_WIDTH = 5
OWNERS_HEADERS_PLUS_SN = OWNERS_HEADERS + ["customer_sn"]


def ensure_owner_sn_schema():
    # non-destructive: adds customer_sn column to owners.xlsx if missing
    ensure_headers(OWNERS_XLSX, OWNERS_HEADERS_PLUS_SN)


def _sn_to_int(sn: str):
    try:
        sn = (sn or "").strip()
        m = re.match(rf"^{re.escape(EASY_OWNER_SN_PREFIX)}-(\d+)$", sn)
        return int(m.group(1)) if m else None
    except Exception:
        return None


def next_customer_sn() -> str:
    """
    Generates next SN like TZ-00001, TZ-00002...
    Uses max existing in owners.xlsx, so it stays consistent even if app restarts.
    """
    ensure_owner_sn_schema()
    rows = read_all(OWNERS_XLSX)
    mx = 0
    for r in rows:
        n = _sn_to_int(str(r.get("customer_sn", "") or ""))
        if n and n > mx:
            mx = n
    return f"{EASY_OWNER_SN_PREFIX}-{mx + 1:0{EASY_OWNER_SN_WIDTH}d}"


def phone_key(s: str) -> str:
    d = "".join([c for c in (s or "") if c.isdigit()])
    if d.startswith("00"):
        d = d[2:]
    return d


def phone_variants(s: str):
    d = phone_key(s)
    out = set()
    if not d:
        return out
    out.add(d)

    # Egypt-ish normalization (best effort)
    if d.startswith("0") and len(d) == 11:
        out.add("20" + d[1:])
    if d.startswith("20") and len(d) == 12:
        out.add("0" + d[2:])

    # last digits matching (handles +20 vs 0 vs formatting)
    if len(d) >= 10:
        out.add(d[-10:])
    if len(d) >= 11:
        out.add(d[-11:])
    if len(d) >= 12:
        out.add(d[-12:])

    return out


def owners_match_by_phone(owner_phone: str, query_phone: str) -> bool:
    a = phone_variants(owner_phone)
    b = phone_variants(query_phone)
    if not a or not b:
        return False
    if a.intersection(b):
        return True
    # fallback: suffix match
    for x in a:
        for y in b:
            if x.endswith(y) or y.endswith(x):
                return True
    return False


def find_owners_by_phone(query_phone: str) -> list:
    ensure_owner_sn_schema()
    rows = read_all(OWNERS_XLSX)
    matches = []
    for o in rows:
        if owners_match_by_phone(str(o.get("phone", "")), query_phone):
            matches.append(o)
    # prefer exact-ish matches
    qp = phone_key(query_phone)
    matches.sort(key=lambda x: 0 if phone_key(str(x.get("phone", ""))) == qp else 1)
    return matches


def easy_clear():
    for k in list(session.keys()):
        if k.startswith("easy_"):
            session.pop(k, None)


def easy_require(keys: list):
    for k in keys:
        if not session.get(k):
            flash("Easy Booking session is missing. Please start again.")
            return redirect(url_for("easy_landing"))
    return None


# -------------------------
# Templates (Wizard Pages)
# -------------------------
TEMPLATES["easy_landing.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Easy Booking</h3>
  <div class="muted">Fast wizard: Owner → Pet → Booking → Reminder → Invoice</div>
  <div class="hr"></div>

  <div class="grid two">
    <a class="card" href="{{ url_for('easy_new_start') }}">
      <h3>Easy Booking - New Customer</h3>
      <div class="muted">Create owner with auto SN (TZ-00001...) then continue steps.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill good">Owner</span>
        <span class="pill">Pet</span>
        <span class="pill warn">Booking</span>
        <span class="pill">Reminder</span>
        <span class="pill">Invoice</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('easy_old_start') }}">
      <h3>Old Customer</h3>
      <div class="muted">Search by phone number, select pet, and continue booking steps.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Phone Search</span>
        <span class="pill">Pet Select</span>
        <span class="pill warn">Booking</span>
        <span class="pill">Reminder</span>
        <span class="pill">Invoice</span>
      </div>
    </a>
  </div>

  <div class="hr"></div>
  <div class="row-actions">
    <a class="btn" href="{{ url_for('home') }}">Back</a>
    <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel Session</a>
  </div>
</div>
{% endblock %}
"""

TEMPLATES["easy_owner_step.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 1/5 - Owner (New Customer)</h3>
  <div class="muted">Auto Serial: <b>{{ sn }}</b></div>
  <div class="hr"></div>

  <form method="post" class="form">
    <div class="full">
      <label>Customer SN</label>
      <input name="customer_sn" value="{{ sn }}" readonly>
    </div>

    <div class="full">
      <label>Owner Name</label>
      <input name="owner_name" required>
    </div>

    <div>
      <label>Phone (Search key)</label>
      <input name="phone" placeholder="+20..." required>
    </div>

    <div>
      <label>Email</label>
      <input name="email">
    </div>

    <div class="full">
      <label>Address</label>
      <input name="address">
    </div>

    <div>
      <label>Preferred Contact</label>
      <select name="preferred_contact">
        {% for x in ['Phone','WhatsApp','Email','SMS'] %}
          <option>{{ x }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Notes</label>
      <input name="notes" placeholder="Optional">
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">Next: Pet</button>
      <a class="btn" href="{{ url_for('easy_landing') }}">Back</a>
      <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
    </div>
  </form>
</div>
{% endblock %}
"""

TEMPLATES["easy_pet_step.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 2/5 - Pet</h3>
  <div class="muted">Owner: <b>{{ owner.owner_name }}</b> • Phone: {{ owner.phone }} • SN: {{ owner.customer_sn or '—' }}</div>
  <div class="hr"></div>

  <form method="post" class="form">
    <div class="full">
      <label>Pet Name</label>
      <input name="pet_name" required>
    </div>

    <div>
      <label>Species</label>
      <input name="species" placeholder="Dog / Cat / ...">
    </div>

    <div>
      <label>Breed</label>
      <input name="breed" placeholder="Mixed / ...">
    </div>

    <div>
      <label>Sex</label>
      <select name="sex">
        {% for x in ['Male','Female'] %}
          <option>{{ x }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>DOB</label>
      <input name="dob" placeholder="YYYY-MM-DD">
    </div>

    <div>
      <label>Age (years)</label>
      <input name="age_years" placeholder="Optional">
    </div>

    <div>
      <label>Weight (kg)</label>
      <input name="weight_kg" required>
    </div>

    <div class="full">
      <label>Allergies</label>
      <input name="allergies" placeholder="Optional">
    </div>

    <div class="full">
      <label>Chronic Conditions</label>
      <input name="chronic_conditions" placeholder="Optional">
    </div>

    <div class="full">
      <label>Vaccinations Summary</label>
      <input name="vaccinations_summary" placeholder="Optional">
    </div>

    <div class="full">
      <label>Notes</label>
      <textarea name="notes" placeholder="Optional"></textarea>
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">Next: Booking</button>
      <a class="btn" href="{{ back_url }}">Back</a>
      <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
    </div>
  </form>
</div>
{% endblock %}
"""

TEMPLATES["easy_old_search.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 1/5 - Old Customer</h3>
  <div class="muted">Search by Owner Name, Phone, or Pet Name (you can use one or more fields).</div>

  <div class="hr"></div>

  <form method="post" class="form">
    <div>
      <label>Owner Name</label>
      <input name="owner_name" value="{{ owner_name or '' }}" placeholder="e.g. Ahmed Mohamed">
    </div>

    <div>
      <label>Phone</label>
      <input name="phone" value="{{ phone or '' }}" placeholder="e.g. 010... or +971...">
    </div>

    <div>
      <label>Pet Name</label>
      <input name="pet_name" value="{{ pet_name or '' }}" placeholder="e.g. Luna">
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit">Search</button>
      <a class="btn" href="{{ url_for('easy_landing') }}">Back</a>
      <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
      <span class="small">Tip: partial text works (contains search).</span>
    </div>
  </form>

  {% if matches is not none %}
    <div class="hr"></div>
    {% if matches %}
      <div class="muted">Select the correct owner:</div>
      <div class="hr"></div>
      <ul class="list">
        {% for o in matches %}
          <li class="list-item">
            <div>
              <b>{{ o.owner_name }}</b> — {{ o.phone }}{% if o.email %} — {{ o.email }}{% endif %}<br>
              <span class="small">SN: {{ o.customer_sn or '—' }}</span>
            </div>
            <form method="post" action="{{ url_for('easy_old_pick_owner') }}">
              <input type="hidden" name="owner_id" value="{{ o.id }}">
              <button class="btn warn" type="submit">Select</button>
            </form>
          </li>
        {% endfor %}
      </ul>
    {% else %}
      <div class="muted">No matches found.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <a class="btn good" href="{{ url_for('easy_new_start') }}">Create as New Customer</a>
      </div>
    {% endif %}
  {% endif %}
</div>
{% endblock %}
"""

TEMPLATES["easy_old_pets.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 1/4 - Select Pet</h3>
  <div class="muted">Owner: <b>{{ owner.owner_name }}</b> • Phone: {{ owner.phone }} • SN: {{ owner.customer_sn or '—' }}</div>
  <div class="hr"></div>

  {% if pets %}
    <form method="post" class="form">
      <div class="full">
        <label>Choose Pet</label>
        <select name="pet_id" required>
          {% for p in pets %}
            <option value="{{ p.id }}">{{ p.pet_name }} ({{ p.species }})</option>
          {% endfor %}
        </select>
      </div>

      <div class="full row-actions">
        <button class="btn good" type="submit">Next: Booking</button>
        <a class="btn" href="{{ url_for('easy_old_pet_new') }}">Add New Pet</a>
        <a class="btn" href="{{ url_for('easy_old_search') }}">Back</a>
        <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
      </div>
    </form>
  {% else %}
    <div class="muted">No pets found for this owner.</div>
    <div class="hr"></div>
    <div class="row-actions">
      <a class="btn good" href="{{ url_for('easy_old_pet_new') }}">Add New Pet</a>
      <a class="btn" href="{{ url_for('easy_old_search') }}">Back</a>
      <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
    </div>
  {% endif %}
</div>
{% endblock %}
"""

TEMPLATES["easy_booking_step.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 3/5 - Booking</h3>

  <div class="small" style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;margin:6px 0 12px 0">
    <div><b>Owner:</b> {{ owner.owner_name }}{% if owner.phone %} ({{ owner.phone }}){% endif %}</div>
    <div><b>Pet:</b> {{ pet.pet_name }}{% if pet.species %} ({{ pet.species }}){% endif %}</div>
    <div style="display:flex;gap:8px;align-items:center">
      <button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetHistory('{{ pet.id }}')">History</button>
      <button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetReminders('{{ pet.id }}')">Remind</button>
    </div>
  </div>

<div class="hr"></div>
<form method="post" class="form">
    <div>
      <label>Appointment Start</label>
      <input name="appointment_start" value="{{ default_start }}" placeholder="YYYY-MM-DD HH:MM" required>
    </div>

    <div>
      <label>Duration (min)</label>
      <input name="duration_min" value="30" required>
    </div>

    <div>
      <label>Weight (kg)</label>
      <input name="visit_weight_kg" value="{{ pet.weight_kg or '' }}" placeholder="e.g. 5.40" required>
    </div>

    <div>
      <label>Temperature (°C)</label>
      <input name="visit_temp_c" value="" placeholder="e.g. 38.5" required>
    </div>

    <div>
      <label>Appointment Type</label>
      <select name="appointment_type">
        {% for t in types %}
          <option>{{ t }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Priority</label>
      <select name="priority">
        {% for p in priorities %}
          <option>{{ p }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Channel</label>
      <select name="channel">
        {% for c in channels %}
          <option>{{ c }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Status</label>
      <select name="status">
        {% for s in statuses %}
          <option {% if s=='Scheduled' %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label style="display:flex;justify-content:space-between;align-items:center;gap:10px">
        <span>Reason</span>
        {% if easy_mode=='old' %}
          <span class="pill" style="cursor:pointer" onclick="saveReasonDraft('{{ owner.id }}','{{ pet.id }}')">Draft</span>
        {% endif %}
      </label>
      <textarea name="reason" id="reason" data-draft-owner="{{ owner.id }}" data-draft-pet="{{ pet.id }}" data-draft-autosave="{% if easy_mode=='old' %}1{% else %}0{% endif %}" placeholder="Main reason / request" style="min-height:44px;height:44px;resize:none;overflow:hidden"></textarea>
      {% if easy_mode=='old' %}
        <div class="small" style="margin-top:6px">
          Draft: <a href="#" onclick="saveReasonDraft('{{ owner.id }}','{{ pet.id }}');return false;">Save</a> ·
          <a href="#" onclick="loadReasonDraft('{{ owner.id }}','{{ pet.id }}');return false;">Load</a> ·
          <a href="#" onclick="clearReasonDraft('{{ owner.id }}','{{ pet.id }}');return false;">Clear</a>
        </div>
        <script>
          window.addEventListener("load", function(){ try{ loadReasonDraft('{{ owner.id }}','{{ pet.id }}'); }catch(e){} });
        </script>
      {% endif %}
    </div>

    <div class="full">
      <label>Symptoms</label>
      <input name="symptoms" placeholder="Optional">
    </div>

    <div>
      <label>Vet Name</label>
      <select name="vet_name" required>
        {% for v in vets %}
          <option value="{{ v }}" {% if (default_vet|lower)==(v|lower) %}selected{% endif %}>{{ v }}</option>
        {% endfor %}
      </select>
    </div>

    <div>
      <label>Room</label>
      <select name="room" required>
        {% for r in rooms %}
          <option value="{{ r }}">{{ r }}</option>
        {% endfor %}
      </select>
    </div>



    <div>
      <label>Service</label>
      <select name="service_name" id="service_name" onchange="syncServiceFee()">
        {% for s in services %}
          <option value="{{ s.name }}" data-fee="{{ s.fee }}">{{ s.name }} ({{ s.fee }})</option>
        {% endfor %}
      </select>
    </div>
<input type="hidden" name="services_json" id="services_json" value="">
<input type="hidden" name="service_name" id="service_name" value="">

<div class="full">
  <label>Services</label>
  <div class="small muted" style="margin-top:-6px;margin-bottom:10px;">
    Select services (up to 10). Fee will auto-fill. A new row appears when you select a service.
  </div>

  <div style="overflow:auto;border:1px solid var(--line);border-radius:12px;">
    <table id="svcTable" class="tbl" style="width:100%;border-collapse:collapse;">
      <thead>
        <tr style="background:var(--soft);">
          <th style="text-align:left;padding:10px;">Service</th>
          <th style="text-align:left;padding:10px;width:140px;">Fee</th>
          <th style="text-align:left;padding:10px;width:110px;">Qty</th>
          <th style="text-align:left;padding:10px;width:150px;">Line Total</th>
          <th style="padding:10px;width:90px;"></th>
        </tr>
      </thead>
      <tbody id="svcBody"></tbody>
    </table>
  </div>
  <div style="margin-top:10px;display:flex;gap:10px;align-items:center;">
    <button type="button" class="btn" id="addSvcBtn">+ Add service</button>
    <div class="muted small">Subtotal is calculated automatically.</div>
  </div>

  <template id="svcRowTpl">
    <tr class="svcRow">
      <td style="padding:10px;">
        <select class="svcSel">
          <option value="">Select service…</option>
          {% for s in services %}
            <option value="{{ s.name }}" data-fee="{{ s.fee }}">{{ s.name }} ({{ s.fee }})</option>
          {% endfor %}
        </select>
      </td>
      <td style="padding:10px;">
        <input type="number" step="0.01" min="0" class="svcFee" placeholder="0.00">
      </td>
      <td style="padding:10px;">
        <input type="number" min="1" value="1" class="svcQty">
      </td>
      <td style="padding:10px;">
        <span class="svcLine">0.00</span>
      </td>
      <td style="padding:10px;">
        <button type="button" class="btn bad svcRemove" title="Remove">✕</button>
      </td>
    </tr>
  </template>
</div>

<div>
  <label>Subtotal (auto)</label>
  <input name="service_fee" id="service_fee" placeholder="Subtotal" value="" readonly>
</div>
<div>
  <label>Discount Type</label>
  <select name="discount_type" id="discount_type" onchange="recalcTotals()">
    <option value="value" selected>Value</option>
    <option value="percent">Percentage</option>
  </select>
</div>

<div>
  <label>Discount Value</label>
  <input name="discount_value" id="discount_value" placeholder="0" value="0" oninput="recalcTotals()">
  <div class="small muted">If <b>Percentage</b> is selected, the discount is applied to the subtotal.</div>
</div>

<div>
  <label>Discount Applied (auto)</label>
  <input name="discount" id="discount" placeholder="0.00" value="0" readonly>
</div>

<div>
  <label>Final Total (auto)</label>
  <input name="final_total" id="final_total" placeholder="Final total" value="" readonly>
</div>

<div>
  <label>Payment Channel</label>
  <select name="payment_channel" id="payment_channel" required>
    <option value="" selected disabled>Select...</option>
    <option value="Cash">Cash</option>
    <option value="Visa">Visa</option>
    <option value="Instapay">Instapay</option>
  </select>
</div>


<div>
  <label>Paid</label>
  <input name="paid_amount" id="paid_amount" placeholder="Paid amount" value="" oninput="recalcTotals()">
</div>

<div>
  <label>Due (auto)</label>
  <input name="due_amount" id="due_amount" placeholder="Due amount" value="" readonly>
</div>
<div>
      <label>Reminder Channel</label>
      <select name="reminder_channel">
        {% for rc in reminder_channels %}
          <option>{{ rc }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="full">
      <label>Notes</label>
      <textarea name="notes" placeholder="Optional"></textarea>
    </div>

    <div class="full row-actions">
      <button class="btn good" type="submit" name="next_step" value="reminder">Reminder</button>
        <button class="btn primary" type="submit" name="next_step" value="invoice">Invoice</button>
      <a class="btn" href="{{ back_url }}">Back</a>
      <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
    </div>
  </form>
</div>

  <script>
  // Reason field: auto-expand after 15 words (keeps compact height for short text)
  function autoExpandReason(){
    const el = document.getElementById("reason");
    if (!el) return;
    const words = (el.value || "").trim().match(/\S+/g);
    const wc = words ? words.length : 0;

    if (wc <= 15){
      el.style.minHeight = "44px";
      el.style.height = "44px";
      el.style.overflow = "hidden";
      el.style.resize = "none";
      return;
    }

    el.style.resize = "vertical";
    el.style.overflow = "hidden";
    el.style.minHeight = "92px";
    el.style.height = "auto";
    el.style.height = Math.max(el.scrollHeight, 92) + "px";
  }

  window.addEventListener("load", ()=>{
    const el = document.getElementById("reason");
    if (el){
      el.addEventListener("input", autoExpandReason);
      autoExpandReason();
    }
  });

// ========= Services table (multi-line) =========
const svcBody = document.getElementById("svcBody");
const rowTpl = document.getElementById("svcRowTpl");
const addBtn = document.getElementById("addSvcBtn");
const servicesJsonEl = document.getElementById("services_json");
const subtotalEl = document.getElementById("service_fee");
const paidEl = document.getElementById("paid_amount");
const dueEl = document.getElementById("due_amount");

const MAX_SERVICES_ROWS = 10;

function money(v, d=0){
  if (v === null || v === undefined) return d;
  if (typeof v === "number") return isFinite(v) ? v : d;
  const s = String(v).replace(/,/g,"").trim();
  if (!s) return d;
  const x = parseFloat(s);
  return isFinite(x) ? x : d;
}

function buildServicesPayload(){
  const rows = Array.from(svcBody.querySelectorAll("tr.svcRow"));
  const out = [];
  rows.forEach(r=>{
    const sel = r.querySelector(".svcSel");
    const feeEl = r.querySelector(".svcFee");
    const qtyEl = r.querySelector(".svcQty");
    const name = sel ? sel.value : "";
    if (!name) return;
    const fee = money(feeEl ? feeEl.value : 0);
    let qty = 1;
    try{ qty = Math.max(1, parseInt(qtyEl.value || "1")); }catch(e){ qty = 1; }
    const line_total = round2(fee * qty);
    out.push({name, fee: round2(fee), qty, line_total});
  });
  return out;
}

function round2(x){ return Math.round((money(x)*100))/100; }

function recalcTotals(){
  if (!svcBody || !subtotalEl) return;
  const payload = buildServicesPayload();
  if (servicesJsonEl) servicesJsonEl.value = JSON.stringify(payload);
  // best-effort for legacy single service_name
  if (document.getElementById("service_name")){
    const sn = document.getElementById("service_name");
    if (payload.length === 0) sn.value = "";
    else {
      const first = payload[0].name;
      sn.value = (payload.length===1) ? first : (first + " +" + (payload.length-1));
    }
  }
  const subtotal = round2(payload.reduce((s,it)=> s + money(it.line_total), 0));

  const dtypeEl = document.getElementById("discount_type");
  const dvalEl  = document.getElementById("discount_value");
  const discountEl = document.getElementById("discount");

  const dtype = (dtypeEl ? String(dtypeEl.value || "value").toLowerCase() : "value");
  let dval = money(dvalEl ? dvalEl.value : 0);

  let discount = 0;
  if (dtype === "percent" || dtype === "percentage"){
    let pct = dval;
    if (pct < 0) pct = 0;
    if (pct > 100) pct = 100;
    discount = round2(subtotal * (pct / 100));
  } else {
    discount = dval;
  }

  if (discount < 0) discount = 0;
  if (discount > subtotal) discount = subtotal;
  if (discountEl) discountEl.value = discount.toFixed(2);

  const finalTotal = round2(subtotal - discount);
  const paid = money(paidEl ? paidEl.value : 0);
  const due = round2(finalTotal - paid);

  subtotalEl.value = subtotal.toFixed(2);
  const ftEl = document.getElementById("final_total");
  if (ftEl) ftEl.value = finalTotal.toFixed(2);
  if (dueEl) dueEl.value = due.toFixed(2);
}

// Backward compatibility
function syncServiceFee(){ recalcTotals(); }
function calcDue(){ recalcTotals(); }

function calcLine(node){
  const feeEl = node.querySelector(".svcFee");
  const qtyEl = node.querySelector(".svcQty");
  const lineEl = node.querySelector(".svcLine");
  const fee = money(feeEl ? feeEl.value : 0);
  let qty = 1;
  try{ qty = Math.max(1, parseInt(qtyEl.value || "1")); }catch(e){ qty = 1; }
  if (lineEl) lineEl.textContent = round2(fee * qty).toFixed(2);
}

function addRow(pref){
  if (!svcBody || !rowTpl) return null;
  const current = svcBody.querySelectorAll("tr.svcRow").length;
  if (current >= MAX_SERVICES_ROWS) return null;

  const node = rowTpl.content.firstElementChild.cloneNode(true);
  const sel = node.querySelector(".svcSel");
  const feeEl = node.querySelector(".svcFee");
  const qtyEl = node.querySelector(".svcQty");
  const rm = node.querySelector(".svcRemove");

  if (pref){
    if (pref.name && sel){
      sel.value = pref.name;
      const opt = sel.options[sel.selectedIndex];
      if (opt && opt.dataset && opt.dataset.fee && (!pref.fee || pref.fee==0)){
        feeEl.value = opt.dataset.fee;
      } else if (pref.fee !== undefined){
        feeEl.value = pref.fee;
      }
    }
    if (pref.fee !== undefined && feeEl && !feeEl.value) feeEl.value = pref.fee;
    if (pref.qty && qtyEl) qtyEl.value = pref.qty;
  }

  function onChange(){
    const opt = sel.options[sel.selectedIndex];
    if (opt && opt.dataset && opt.dataset.fee && (!feeEl.value || feeEl.value==="0" || feeEl.value==="0.00")){
      feeEl.value = opt.dataset.fee;
    }
    calcLine(node);

    const rows = Array.from(svcBody.querySelectorAll("tr.svcRow"));
    const isLast = rows.length ? (rows[rows.length - 1] === node) : true;
    if (isLast && sel.value){
      const countNow = svcBody.querySelectorAll("tr.svcRow").length;
      if (countNow < MAX_SERVICES_ROWS) addRow(); // add trailing empty
    }
    recalcTotals();
  }

  sel.addEventListener("change", onChange);
  feeEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });
  qtyEl.addEventListener("input", ()=>{ calcLine(node); recalcTotals(); });

  rm.addEventListener("click", ()=>{
    node.remove();
    if (svcBody.querySelectorAll("tr.svcRow").length === 0) addRow();
    recalcTotals();
  });

  svcBody.appendChild(node);
  calcLine(node);
  recalcTotals();
  return node;
}

if (addBtn){
  addBtn.addEventListener("click", ()=> addRow());
}

(function initServices(){
  // If page contains pre-loaded services_json, load it. Else start empty row.
  let existing = [];
  try{
    const raw = (servicesJsonEl && servicesJsonEl.value) ? servicesJsonEl.value : "";
    if (raw) existing = JSON.parse(raw);
  } catch(e){ existing = []; }

  if (existing && existing.length){
    existing.forEach(it=> addRow(it));
    addRow(); // trailing
  } else {
    addRow();
  }
  recalcTotals();
})();
</script>


<script>
  function _num(v){
    try{
      const x = parseFloat(String(v||"").replace(/[^0-9.\-]/g,""));
      return isNaN(x)?0:x;
    }catch(e){ return 0; }
  }
  function syncServiceFee(){
    const sel = document.getElementById("service_name");
    const feeEl = document.getElementById("service_fee");
    if(!sel || !feeEl) return;
    const opt = sel.options[sel.selectedIndex];
    const fee = opt ? opt.getAttribute("data-fee") : "";
    if(!feeEl.value){
      feeEl.value = fee || "";
    }
    calcDue();
  }
  function calcDue(){
    const fee = _num(document.getElementById("service_fee")?.value);
    const paid = _num(document.getElementById("paid_amount")?.value);
    const vatRate = _num(document.getElementById("vat_rate")?.value);
    const total = fee + (fee * vatRate);
    const due = total - paid;
    const dueEl = document.getElementById("due_amount");
    if(dueEl) dueEl.value = (Math.max(due,0)).toFixed(2);
  }
  window.addEventListener("load", ()=>{
    syncServiceFee();
    calcDue();
  });
</script>

{% endblock %}
"""

TEMPLATES["easy_reminder_step.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 4/5 — Service Reminders</h3>
  <div class="muted">Top section = global master service list (affects Booking options). Below = only the services selected for this booking, with a reminder date/time per service.</div>
  <div class="hr"></div>

  <div class="grid two">
    <div class="card">
      <h3>Master Service List (Global)</h3>
      <div class="muted">Add, delete (soft remove), or restore services. Deleted services will disappear from Booking service selection.</div>

      <div class="hr"></div>

      <form method="post" class="form">
        <input type="hidden" name="action" value="master_add" />
        <div>
          <label>Service Name</label>
          <input name="service_name" placeholder="e.g., Vaccination" required />
        </div>
        <div>
          <label>Price</label>
          <input name="service_fee" type="number" step="0.01" value="0" />
        </div>
        <div class="full row-actions">
          <button class="btn good" type="submit">Add Service to Master List</button>
        </div>
      </form>

      <div style="overflow:auto; margin-top:12px">
        <table>
          <thead>
            <tr>
              <th style="min-width:220px">Service</th>
              <th>Price</th>
              <th>Status</th>
              <th style="min-width:140px">Action</th>
            </tr>
          </thead>
          <tbody>
            {% for s in master_services %}
            <tr>
              <td><b>{{ s.name }}</b></td>
              <td>{{ "%.2f"|format(s.fee|float) }}</td>
              <td>
                {% if s.active_bool %}
                  <span class="pill good">Active</span>
                {% else %}
                  <span class="pill warn">Removed</span>
                {% endif %}
              </td>
              <td>
                <form method="post" style="margin:0">
                  <input type="hidden" name="action" value="master_toggle" />
                  <input type="hidden" name="service_id" value="{{ s.id }}" />
                  {% if s.active_bool %}
                    <input type="hidden" name="set_active" value="0" />
                    <button class="btn bad" type="submit" onclick="return confirm('Delete this service from the master list? It will disappear from Booking options.')">Delete</button>
                  {% else %}
                    <input type="hidden" name="set_active" value="1" />
                    <button class="btn good" type="submit">Restore</button>
                  {% endif %}
                </form>
              </td>
            </tr>
            {% endfor %}
            {% if master_services|length == 0 %}
            <tr><td colspan="4" class="muted">No services found.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <h3>Booking Context & Actions</h3>
      <div class="muted">
        <div><b>Owner:</b> {{ b.owner_name }} ({{ b.owner_phone }})</div>
        <div><b>Pet:</b> {{ b.pet_name }}</div>
        <div><b>Appointment:</b> {{ b.appointment_start }}</div>
      </div>

      <div class="small" style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap">
        <button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetHistory('{{ b.pet_id }}')">History</button>
        <button type="button" class="btn" style="padding:8px 10px;font-size:12px" onclick="openPetReminders('{{ b.pet_id }}')">Remind</button>
      </div>

      <div class="hr"></div>

      <div class="muted">
        <div>• Reminder dates are saved per service under this booking.</div>
        <div>• Saved reminders are added/updated in the Reminders Center automatically.</div>
      </div>

      <div class="hr"></div>

      <div class="row-actions">
        <a class="btn" href="{{ url_for('easy_booking_step') }}">Back</a>
        <a class="btn good" href="{{ url_for('easy_invoice_step') }}">Continue to Invoice</a>
        <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Flow? All unsaved work will be lost.')">Cancel</a>
      </div>

      {% if wa_link and appt_msg %}
      <div class="hr"></div>
      <details>
        <summary><b>Appointment Reminder (Optional)</b></summary>
        <div style="margin-top:10px" class="muted">Legacy appointment reminder message (optional).</div>
        <div class="hr"></div>
        <label>Message Preview</label>
        <textarea rows="5" readonly>{{ appt_msg }}</textarea>
        <div class="row-actions" style="margin-top:8px">
          <a class="btn primary" href="{{ wa_link }}" target="_blank">Open WhatsApp</a>
        </div>
      </details>
      {% endif %}
    </div>
  </div>

  <div class="hr"></div>

  <div class="card">
    <h3>Selected Services (This Booking)</h3>
    <div class="muted">This list comes from the Booking page for this customer/pet/booking. Assign a reminder date/time per service.</div>

    <div class="hr"></div>

    <form method="post" class="form">
      <input type="hidden" name="action" value="booking_add" />
      <div class="full">
        <label>Add a service to this booking</label>
        <select name="service_id">
          {% for s in active_services %}
            <option value="{{ s.id }}">{{ s.name }} ({{ "%.2f"|format(s.fee|float) }})</option>
          {% endfor %}
        </select>
      </div>
      <div>
        <label>Qty</label>
        <input name="qty" type="number" min="1" value="1" />
      </div>
      <div class="full row-actions">
        <button class="btn primary" type="submit">Add to Booking List</button>
      </div>
    </form>

    <div class="hr"></div>

    <form method="post">
      <input type="hidden" name="action" value="booking_update" />
      <div style="overflow:auto">
        <table>
          <thead>
            <tr>
              <th style="min-width:220px">Service</th>
              <th>Qty</th>
              <th>Price</th>
              <th>Subtotal</th>
              <th style="min-width:240px">Reminder Date/Time</th>
              <th style="min-width:120px">Remove</th>
            </tr>
          </thead>
          <tbody>
            {% for svc in selected_services %}
            <tr>
              <td><b>{{ svc.name }}</b></td>
              <td>{{ svc.qty }}</td>
              <td>{{ "%.2f"|format(svc.fee|float) }}</td>
              <td>{{ "%.2f"|format(svc.subtotal|float) }}</td>
              <td>
                <input type="datetime-local" name="reminder_at_{{ loop.index0 }}" value="{{ svc.reminder_at_local }}" />
              </td>
              <td>
                <button class="btn bad" type="submit" name="remove_idx" value="{{ loop.index0 }}" onclick="return confirm('Remove this service from the booking list?')">Remove</button>
              </td>
            </tr>
            {% endfor %}
            {% if selected_services|length == 0 %}
            <tr><td colspan="6" class="muted">No services selected yet.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <div class="row-actions" style="margin-top:12px">
        <button class="btn good" type="submit">Save Reminder Dates</button>
        <a class="btn" href="{{ url_for('easy_invoice_step') }}">Continue to Invoice</a>
      </div>

      <div class="muted" style="margin-top:10px">
        Current subtotal: <b>{{ "%.2f"|format(subtotal|float) }}</b> &nbsp;|&nbsp; VAT rate: <b>{{ (vat_rate*100)|round(0) }}%</b>
      </div>
    </form>
  </div>

</div>
{% endblock %}
"""

TEMPLATES["easy_invoice_step.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <h3>Step 5/5 - Invoice</h3>
  <div class="muted">Preview and print invoice, then finish the wizard.</div>
  <div class="hr"></div>

  <div class="grid two">
    <div class="card">
      <h3>Summary</h3>
      <div class="muted">
        Invoice: <b>{{ invoice_no }}</b><br>
        Owner: <b>{{ b.owner_name }}</b> ({{ b.owner_phone }}) • SN: {{ owner_sn or '—' }}<br>
        Pet: <b>{{ b.pet_name }}</b><br>
        Appointment: <b>{{ b.appointment_start }}</b><br>
        Type: {{ b.appointment_type }} • Status: {{ b.status }} • Priority: {{ b.priority }}<br>
        Subtotal: <b>{{ '%.2f'|format(subtotal) }}</b><br>
        Discount: <b>-{{ '%.2f'|format(discount_amt) }}</b>
        <span class="muted">({{ 'Percentage' if discount_type=='percent' else 'Value' }}{% if discount_type=='percent' and discount_value %}: {{ discount_value }}%{% endif %})</span><br>
        Net: <b>{{ '%.2f'|format(net_fee) }}</b><br>
        Paid: <b>{{ '%.2f'|format(paid_amt) }}</b><br>
        Due: <b>{{ '%.2f'|format(due_amt) }}</b><br>
        Payment: <b>{{ payment_channel or '—' }}</b>
      </div>
    </div>

    <div class="card">
      <h3>Actions</h3>
      <div class="row-actions">
        <a class="btn primary" href="{{ url_for('easy_invoice_print', booking_id=b.id) }}" target="_blank">Open Printable Invoice</a>
        <a class="btn" href="{{ url_for('booking_view', booking_id=b.id) }}">Open Exam</a>
        <a class="btn" href="{{ url_for('bookings') }}">All Bookings</a>
      </div>
      <div class="hr"></div>
      <div class="row-actions">
        <a class="btn good" href="{{ url_for('easy_finish') }}">Finish</a>
        <a class="btn" href="{{ back_url }}">Back</a>
        <a class="btn bad" href="{{ url_for('easy_cancel') }}" onclick="return confirm('Cancel Easy Booking session?')">Cancel</a>
      </div>
    </div>
  </div>
</div>
{% endblock %}
"""

TEMPLATES["easy_invoice_print.html"] = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{{ app_title }} | Invoice {{ invoice_no }}</title>
  <style>
    :root{--bg:#0b1220;--panel:#0f1b33;--card:#111f3a;--text:rgba(255,255,255,.92);--muted:rgba(255,255,255,.70);--line:rgba(255,255,255,.12)}
    body{margin:0;font-family:ui-sans-serif,system-ui;background:var(--bg);color:var(--text)}
    .wrap{max-width:980px;margin:0 auto;padding:22px}
    .bar{display:flex;justify-content:space-between;align-items:center;gap:10px;margin-bottom:12px}
    .btn{display:inline-flex;align-items:center;justify-content:center;padding:10px 12px;border-radius:12px;border:1px solid var(--line);background:rgba(255,255,255,.04);color:var(--text);cursor:pointer;text-decoration:none}
    .btn.primary{background:rgba(96,165,250,.16);border-color:rgba(96,165,250,.25)}
    .card{background:linear-gradient(180deg, rgba(255,255,255,.05), rgba(255,255,255,.02));border:1px solid var(--line);border-radius:18px;padding:14px}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    .muted{color:var(--muted);font-size:12px}
    table{width:100%;border-collapse:collapse;margin-top:10px}
    th,td{padding:10px 8px;border-bottom:1px solid rgba(255,255,255,.10);font-size:12px;text-align:left;vertical-align:top}
    h1,h2,h3{margin:0}
    .right{text-align:right}
    @media print{
      .noprint{display:none}
      body{background:#fff;color:#000}
      .card{background:#fff;border-color:#ddd}
      .muted{color:#444}
      th,td{border-bottom:1px solid #ddd}
      a{color:#000;text-decoration:none}
    }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="bar noprint">
      <a class="btn" href="javascript:window.close()">Close</a>
      <button class="btn primary" onclick="window.print()">Print</button>
    </div>

    <div class="card">
      <div class="grid">
        <div>
          <h2>{{ app_title }}</h2>
          <div class="muted">Invoice</div>
        </div>
        <div class="right">
          <h3>{{ invoice_no }}</h3>
          <div class="muted">Date: {{ today }}</div>
        </div>
      </div>

      <div style="height:1px;background:rgba(255,255,255,.12);margin:12px 0"></div>

      <div class="grid">
        <div>
          <h3>Bill To</h3>
          <div class="muted">
            Owner: <b>{{ owner_name }}</b><br>
            Phone: {{ owner_phone }}<br>
            Customer SN: {{ owner_sn or '—' }}<br>
            Address: {{ owner_address or '—' }}
          </div>
        </div>
        <div>
          <h3>Appointment</h3>
          <div class="muted">
            Pet: <b>{{ pet_name }}</b><br>
            Start: {{ appt_start }}<br>
            Type: {{ appt_type }}<br>
            Status: {{ status }}<br>
            Payment Channel: {{ payment_channel or '—' }}<br>
            Payment Method: {{ payment_method or '—' }}<br>
            Payment Status: {{ payment_status or '—' }}
          </div>
        </div>
      </div>

      <div style="height:1px;background:rgba(255,255,255,.12);margin:12px 0"></div>

      <h3>Services</h3>
      <table>
        <thead>
          <tr>
            <th>Description</th>
            <th class="right" style="width:70px">Qty</th>
            <th class="right" style="width:120px">Unit</th>
            <th class="right" style="width:120px">Amount</th>
          </tr>
        </thead>
        <tbody>
          {% for it in line_items %}
          <tr>
            <td><b>{{ it.desc }}</b></td>
            <td class="right">{{ it.qty }}</td>
            <td class="right">{{ "%.2f"|format(it.unit_price) }}</td>
            <td class="right"><b>{{ "%.2f"|format(it.line_total) }}</b></td>
          </tr>
          {% endfor %}
        </tbody>
      </table>

      <div style="display:flex;justify-content:flex-end;margin-top:10px">
        <div style="min-width:320px">
          <table>
            <tr><td class="muted">Subtotal</td><td class="right">{{ subtotal }}</td></tr>
            <tr>
              <td class="muted">Discount
                {% if discount_type == "percent" and discount_value %}
                  ({{ discount_value }}%)
                {% endif %}
              </td>
              <td class="right">-{{ discount_amt }}</td>
            </tr>
            <tr><td class="muted">Net</td><td class="right">{{ net_fee }}</td></tr>
            <tr><td class="muted">VAT ({{ (vat_rate*100)|round(0) }}%)</td><td class="right">{{ vat }}</td></tr>
            <tr><td><b>Total</b></td><td class="right"><b>{{ total }}</b></td></tr>
            <tr><td class="muted">Paid</td><td class="right">{{ paid }}</td></tr>
            <tr><td class="muted">Due</td><td class="right"><b>{{ due }}</b></td></tr>
          </table>
        </div>
      </div>

      {% if reason or notes %}
      <div style="height:1px;background:rgba(255,255,255,.12);margin:12px 0"></div>
      <div class="grid">
        <div>
          <h3>Reason</h3>
          <div class="muted">{{ reason or '—' }}</div>
        </div>
        <div>
          <h3>Notes</h3>
          <div class="muted">{{ notes or '—' }}</div>
        </div>
      </div>
      {% endif %}
    </div>
  </div>
</body>
</html>

"""

TEMPLATES["invoice_print.html"] = r"""
{% extends "base.html" %}
{% block content %}
<div class="card">
  <div style="display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;align-items:flex-start">
    <div>
      <h2>Invoice <span class="hint">#{{ invoice_no }}</span></h2>
      <div class="hint">{{ clinic.name }}</div>
      <div class="hint">{{ clinic.address }}</div>
      <div class="hint">{{ clinic.phone }}</div>
    </div>
    <div class="row-actions">
      <button class="btn" onclick="window.print()">Print</button>
      <a class="btn" href="{{ url_for('bookings') }}">Back</a>
    </div>
  </div>

  <hr>

  <div class="grid">
    <div class="card" style="padding:14px">
      <h3>Customer</h3>
      <div class="hint">Owner: <b>{{ owner.owner_name }}</b></div>
      <div class="hint">Phone: {{ owner.phone }}</div>
      <div class="hint">Pet: <b>{{ pet.pet_name }}</b> {% if pet.species %}({{ pet.species }}){% endif %}</div>
    </div>

    <div class="card" style="padding:14px">
      <h3>Booking</h3>
      <div class="hint">Start: {{ appt_start }}</div>
      <div class="hint">Duration: {{ dur }} min</div>
      <div class="hint">Payment Channel: {{ payment_channel or '—' }}</div>
      <div class="hint">Payment Method: {{ payment_method or '—' }}</div>
    </div>
  </div>

  <h3 class="mt">Services</h3>
  <table class="table mt">
    <thead>
      <tr>
        <th>Description</th>
        <th style="width:80px;text-align:right">Qty</th>
        <th style="width:120px;text-align:right">Unit</th>
        <th style="width:140px;text-align:right">Amount</th>
      </tr>
    </thead>
    <tbody>
      {% for it in line_items %}
      <tr>
        <td>
          <b>{{ it.desc }}</b>
          {% if it.details %}<div class="hint">{{ it.details }}</div>{% endif %}
        </td>
        <td style="text-align:right">{{ it.qty }}</td>
        <td style="text-align:right">{{ "%.2f"|format(it.unit_price) }}</td>
        <td style="text-align:right"><b>{{ "%.2f"|format(it.line_total) }}</b></td>
      </tr>
      {% endfor %}
    </tbody>
  </table>

  <div class="grid mt" style="grid-template-columns:1fr 360px;gap:12px;align-items:start">
    <div></div>
    <div class="card" style="padding:14px">
      <table class="table">
        <tbody>
          <tr><td class="hint">Subtotal</td><td style="text-align:right">{{ "%.2f"|format(subtotal) }}</td></tr>
          <tr>
            <td class="hint">Discount{% if discount_type=='percent' and discount_value %} ({{ discount_value }}%){% endif %}</td>
            <td style="text-align:right">-{{ "%.2f"|format(discount_amt) }}</td>
          </tr>
          <tr><td class="hint">Net</td><td style="text-align:right">{{ "%.2f"|format(net_fee) }}</td></tr>
          <tr><td class="hint">VAT ({{ (vat_rate*100)|round(0) }}%)</td><td style="text-align:right">{{ "%.2f"|format(vat) }}</td></tr>
          <tr><td><b>Total</b></td><td style="text-align:right"><b>{{ "%.2f"|format(total) }}</b></td></tr>
          <tr><td class="hint">Paid</td><td style="text-align:right">{{ "%.2f"|format(paid) }}</td></tr>
          <tr><td class="hint">Due</td><td style="text-align:right"><b>{{ "%.2f"|format(due) }}</b></td></tr>
        </tbody>
      </table>
    </div>
  </div>
</div>
{% endblock %}
"""

# -------------------------
# OPTIONAL: Add "Easy Booking" card to HOME (overrides home.html)
# If you don't want this, remove this block only.
# -------------------------
TEMPLATES["home.html"] = r"""
{% extends "base.html" %}
{% block content %}
  <div class="grid two">
    <a class="card" href="{{ url_for('easy_landing') }}">
      <h3>Easy Booking</h3>
      <div class="muted">New Customer wizard + Old Customer phone search.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill good">Wizard</span>
        <span class="pill">TZ SN</span>
        <span class="pill warn">Fast</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('bookings') }}">
      <h3>Booking Journey</h3>
      <div class="muted">Create appointment → WhatsApp Journey → Owner Portal → Calendar → Popups.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill good">WhatsApp Journey</span>
        <span class="pill">Owner Portal</span>
        <span class="pill warn">AI Intake</span>
      </div>
    </a>
<a class="card" href="{{ url_for('dashboard') }}">
      <h3>Dashboard</h3>
      <div class="muted">Offline charts + KPIs, track performance.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Charts</span>
        <span class="pill good">PNG</span>
      </div>
    </a>

    <a class="card" href="{{ url_for('owners') }}">
      <h3>Owners & Pets</h3>
      <div class="muted">Editable master data.</div>
      <div class="hr"></div>
      <div class="row-actions">
        <span class="pill">Owners</span>
        <span class="pill">Pets</span>
        <span class="pill warn">Edit</span>
      </div>
    </a>
  </div>
{% endblock %}
"""


# -------------------------
# Routes: Landing / Cancel / Finish
# -------------------------
@app.route("/easy")
def easy_landing():
    gate = require_login()
    if gate: return gate
    return render_template("easy_landing.html",
                           title=f"{APP_TITLE} | Easy Booking",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Easy Booking",
                           subtitle="Wizard booking flow (New + Old customers)",
                           active="home"
                           )


@app.route("/easy/cancel")
def easy_cancel():
    gate = require_login()
    if gate: return gate
    easy_clear()
    flash("Easy Booking session cleared.")
    return redirect(url_for("easy_landing"))


@app.route("/easy/finish")
def easy_finish():
    gate = require_login()
    if gate: return gate
    easy_clear()
    flash("Easy Booking completed.")
    return redirect(url_for("bookings"))


# -------------------------
# Routes: New Customer Flow
# -------------------------
@app.route("/easy/new")
def easy_new_start():
    gate = require_login()
    if gate: return gate
    easy_clear()
    session["easy_mode"] = "new"
    return redirect(url_for("easy_new_owner"))


@app.route("/easy/new/owner", methods=["GET", "POST"])
def easy_new_owner():
    gate = require_login()
    if gate: return gate

    ensure_owner_sn_schema()
    sn = next_customer_sn()

    if request.method == "POST":
        ensure_owner_sn_schema()
        oid = str(uuid.uuid4())
        customer_sn = (request.form.get("customer_sn") or sn).strip() or sn

        row = {h: "" for h in OWNERS_HEADERS_PLUS_SN}
        row.update({
            "id": oid,
            "owner_name": (request.form.get("owner_name") or "").strip(),
            "phone": (request.form.get("phone") or "").strip(),
            "email": (request.form.get("email") or "").strip(),
            "address": (request.form.get("address") or "").strip(),
            "preferred_contact": (request.form.get("preferred_contact") or "WhatsApp").strip(),
            "notes": (request.form.get("notes") or "").strip(),
            "customer_sn": customer_sn,
            "created_at": now_str(),
            "updated_at": now_str()
        })
        append_row(OWNERS_XLSX, OWNERS_HEADERS_PLUS_SN, row)

        session["easy_owner_id"] = oid
        session["easy_customer_sn"] = customer_sn
        return redirect(url_for("easy_new_pet"))

    return render_template("easy_owner_step.html",
                           title=f"{APP_TITLE} | Easy Booking - Owner",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Easy Booking - New Customer",
                           subtitle="Step 1: Owner profile",
                           active="home",
                           sn=sn
                           )


@app.route("/easy/new/pet", methods=["GET", "POST"])
def easy_new_pet():
    gate = require_login()
    if gate: return gate

    must = easy_require(["easy_owner_id"])
    if must: return must

    owners_rows = read_all(OWNERS_XLSX)
    owner = find_by_id(owners_rows, session["easy_owner_id"])
    if not owner:
        flash("Owner not found. Start again.")
        return redirect(url_for("easy_new_start"))

    if request.method == "POST":
        pid = str(uuid.uuid4())
        row = {h: "" for h in PETS_HEADERS}
        row.update({
            "id": pid,
            "pet_name": (request.form.get("pet_name") or "").strip(),
            "species": (request.form.get("species") or "").strip(),
            "breed": (request.form.get("breed") or "").strip(),
            "sex": (request.form.get("sex") or "").strip(),
            "dob": (request.form.get("dob") or "").strip(),
            "age_years": (request.form.get("age_years") or "").strip(),
            "weight_kg": (request.form.get("weight_kg") or "").strip(),
            "allergies": (request.form.get("allergies") or "").strip(),
            "chronic_conditions": (request.form.get("chronic_conditions") or "").strip(),
            "vaccinations_summary": (request.form.get("vaccinations_summary") or "").strip(),
            "owner_id": session["easy_owner_id"],
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })
        append_row(PETS_XLSX, PETS_HEADERS, row)
        session["easy_pet_id"] = pid
        return redirect(url_for("easy_booking_step"))

        # Backward compatibility: mirror payment_channel into payment_method
        if not (row.get("payment_method") or "").strip():
            row["payment_method"] = (row.get("payment_channel") or "").strip()

    return render_template("easy_pet_step.html",
                           title=f"{APP_TITLE} | Easy Booking - Pet",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Easy Booking - New Customer",
                           subtitle="Step 2: Pet profile",
                           active="home",
                           owner=owner,
                           back_url=url_for("easy_new_owner")
                           )


# -------------------------
# Routes: Old Customer Flow (Phone Search)
# -------------------------
@app.route("/easy/old")
def easy_old_start():
    gate = require_login()
    if gate: return gate
    easy_clear()
    session["easy_mode"] = "old"
    return redirect(url_for("easy_old_search"))


@app.route("/easy/old/search", methods=["GET", "POST"])
def easy_old_search():
    gate = require_login()
    if gate: return gate

    matches = None
    owner_name = ""
    phone = ""
    pet_name = ""

    if request.method == "POST":
        owner_name = (request.form.get("owner_name") or "").strip()
        phone = (request.form.get("phone") or "").strip()
        pet_name = (request.form.get("pet_name") or "").strip()

        if not (owner_name or phone or pet_name):
            flash("Please enter Owner Name, Phone, or Pet Name to search.")
            return redirect(url_for("easy_old_search"))

        owners_rows = read_all(OWNERS_XLSX)
        pets_rows = read_all(PETS_XLSX)

        def _norm_phone(x: str) -> str:
            return re.sub(r"\D+", "", x or "")

        filtered = owners_rows

        if phone:
            p = _norm_phone(phone)
            filtered = [o for o in filtered if p in _norm_phone(str(o.get("phone", "")))]

        if owner_name:
            on = owner_name.lower()
            filtered = [o for o in filtered if on in str(o.get("owner_name", "")).lower()]

        if pet_name:
            pn = pet_name.lower()
            owner_ids = {str(p.get("owner_id", "")) for p in pets_rows if pn in str(p.get("pet_name", "")).lower()}
            filtered = [o for o in filtered if str(o.get("id", "")) in owner_ids]

        # stable sort
        matches = sorted(filtered, key=lambda x: (str(x.get("owner_name", "")).lower(), str(x.get("phone", ""))))

        if len(matches) == 1:
            session["easy_owner_id"] = matches[0]["id"]
            session["easy_customer_sn"] = str(matches[0].get("customer_sn", "") or "").strip()
            return redirect(url_for("easy_old_pets"))

    return render_template("easy_old_search.html",
                           title=f"{APP_TITLE} | Old Customer Search",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Old Customer",
                           subtitle="Search by Owner Name / Phone / Pet Name",
                           active="home",
                           matches=matches,
                           owner_name=owner_name,
                           phone=phone,
                           pet_name=pet_name
                           )


@app.route("/easy/old/select/<owner_id>")
def easy_old_select_owner(owner_id):
    gate = require_login()
    if gate: return gate
    ensure_owner_sn_schema()
    owners_rows = read_all(OWNERS_XLSX)
    o = find_by_id(owners_rows, owner_id)
    if not o:
        flash("Owner not found.")
        return redirect(url_for("easy_old_search"))
    session["easy_owner_id"] = owner_id
    session["easy_customer_sn"] = str(o.get("customer_sn", "") or "").strip()
    return redirect(url_for("easy_old_pets"))


@app.route("/easy/old/pick_owner", methods=["POST"])
def easy_old_pick_owner():
    gate = require_login()
    if gate: return gate

    owner_id = (request.form.get("owner_id") or "").strip()
    if not owner_id:
        flash("Please select an owner.")
        return redirect(url_for("easy_old_search"))

    return redirect(url_for("easy_old_select_owner", owner_id=owner_id))


@app.route("/easy/old/pets", methods=["GET", "POST"])
def easy_old_pets():
    gate = require_login()
    if gate: return gate

    must = easy_require(["easy_owner_id"])
    if must: return must

    ensure_owner_sn_schema()
    owners_rows = read_all(OWNERS_XLSX)
    owner = find_by_id(owners_rows, session["easy_owner_id"])
    if not owner:
        flash("Owner not found. Start again.")
        return redirect(url_for("easy_old_start"))

    pets_rows = read_all(PETS_XLSX)
    pets_for_owner = [p for p in pets_rows if str(p.get("owner_id", "")) == str(owner.get("id", ""))]
    pets_for_owner.sort(key=lambda x: str(x.get("pet_name", "")).lower())

    if request.method == "POST":
        pid = (request.form.get("pet_id") or "").strip()
        p = find_by_id(pets_rows, pid)
        if not p or str(p.get("owner_id", "")) != str(owner.get("id", "")):
            flash("Invalid pet selection.")
            return redirect(url_for("easy_old_pets"))
        session["easy_pet_id"] = pid
        return redirect(url_for("easy_booking_step"))

    return render_template("easy_old_pets.html",
                           title=f"{APP_TITLE} | Old Customer - Pets",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Old Customer",
                           subtitle="Step 1: Select pet",
                           active="home",
                           owner=owner,
                           pets=pets_for_owner
                           )


@app.route("/easy/old/pets/new", methods=["GET", "POST"])
def easy_old_pet_new():
    gate = require_login()
    if gate: return gate

    must = easy_require(["easy_owner_id"])
    if must: return must

    ensure_owner_sn_schema()
    owners_rows = read_all(OWNERS_XLSX)
    owner = find_by_id(owners_rows, session["easy_owner_id"])
    if not owner:
        flash("Owner not found. Start again.")
        return redirect(url_for("easy_old_start"))

    if request.method == "POST":
        pid = str(uuid.uuid4())
        row = {h: "" for h in PETS_HEADERS}
        row.update({
            "id": pid,
            "pet_name": (request.form.get("pet_name") or "").strip(),
            "species": (request.form.get("species") or "").strip(),
            "breed": (request.form.get("breed") or "").strip(),
            "sex": (request.form.get("sex") or "").strip(),
            "dob": (request.form.get("dob") or "").strip(),
            "age_years": (request.form.get("age_years") or "").strip(),
            "weight_kg": (request.form.get("weight_kg") or "").strip(),
            "allergies": (request.form.get("allergies") or "").strip(),
            "chronic_conditions": (request.form.get("chronic_conditions") or "").strip(),
            "vaccinations_summary": (request.form.get("vaccinations_summary") or "").strip(),
            "owner_id": session["easy_owner_id"],
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })
        append_row(PETS_XLSX, PETS_HEADERS, row)
        session["easy_pet_id"] = pid
        return redirect(url_for("easy_booking_step"))

    return render_template("easy_pet_step.html",
                           title=f"{APP_TITLE} | Old Customer - New Pet",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Old Customer",
                           subtitle="Add new pet then continue",
                           active="home",
                           owner=owner,
                           back_url=url_for("easy_old_pets")
                           )


# -------------------------
# Route: Booking Step (shared for New + Old)
# -------------------------
def _default_start_str():
    # current date/time (minutes precision)
    dt = datetime.now().replace(second=0, microsecond=0)
    return dt.strftime("%Y-%m-%d %H:%M")


@app.route("/easy/booking", methods=["GET", "POST"])
def easy_booking_step():
    gate = require_login()
    if gate: return gate

    must = easy_require(["easy_owner_id", "easy_pet_id"])
    if must: return must

    ensure_owner_sn_schema()
    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    owner = find_by_id(owners_rows, session["easy_owner_id"])
    pet = find_by_id(pets_rows, session["easy_pet_id"])

    if not owner or not pet:
        flash("Owner/Pet not found. Please start again.")
        return redirect(url_for("easy_landing"))

    if request.method == "POST":
        bid = str(uuid.uuid4())
        start = normalize_dt(request.form.get("appointment_start") or "")
        dur = str((request.form.get("duration_min") or "30").strip())
        start_dt = parse_dt(start) or datetime.now()
        if not start:
            start = start_dt.strftime("%Y-%m-%d %H:%M")
        end_dt = start_dt + timedelta(minutes=int(float(dur)))
        token = uuid.uuid4().hex

        row = {h: "" for h in BOOKINGS_HEADERS}
        row.update({
            "id": bid,
            "appointment_start": start,
            "duration_min": dur,
            "appointment_end": end_dt.strftime("%Y-%m-%d %H:%M"),
            "owner_id": str(owner.get("id", "")),
            "pet_id": str(pet.get("id", "")),
            "visit_weight_kg": (request.form.get("visit_weight_kg") or "").strip(),
            "visit_temp_c": (request.form.get("visit_temp_c") or "").strip(),
            "appointment_type": (request.form.get("appointment_type") or "Consultation").strip(),
            "priority": (request.form.get("priority") or "Normal").strip(),
            "status": (request.form.get("status") or "Scheduled").strip(),
            "channel": (request.form.get("channel") or "Walk-in").strip(),
            "reason": (request.form.get("reason") or "").strip(),
            "symptoms": (request.form.get("symptoms") or "").strip(),
            "vet_name": (request.form.get("vet_name") or "").strip(),
            "room": (request.form.get("room") or "").strip(),
            "services_json": (request.form.get("services_json") or "").strip(),
            "service_name": (request.form.get("service_name") or "").strip(),
            "service_fee": "",  # will be computed from services_json

            "paid_amount": (request.form.get("paid_amount") or "").strip(),
            "due_amount": (request.form.get("due_amount") or "").strip(),
            "fee_amount": "",  # computed after discount
            "payment_channel": (request.form.get("payment_channel") or "").strip(),
            "payment_status": (request.form.get("payment_status") or "Unpaid").strip(),
            "payment_method": (request.form.get("payment_method") or "").strip(),
            "invoice_no": "",
            "followup_datetime": "",
            "reminder_channel": (request.form.get("reminder_channel") or "WhatsApp").strip(),
            "reminder_sent": "",
            "reminder_last_opened": "",
            "portal_token": token,
            "owner_confirmed": "",
            "owner_update_message": "",
            "owner_update_datetime": "",
            "ai_last_applied_at": "",
            "notes": (request.form.get("notes") or "").strip(),
            "created_at": now_str(),
            "updated_at": now_str()
        })

        # Payment channel is required
        if not (row.get("payment_channel") or "").strip():
            flash("Payment channel is required (Cash / Visa / Instapay).")
            return redirect(url_for("easy_booking_step"))
        # --- Service/Fee calculation (server-side safety) ---
        services = []
        raw_services = (row.get("services_json") or "").strip()
        if not raw_services:
            raw_services = "[]"
            row["services_json"] = raw_services
        try:
            services = json.loads(raw_services)
            if not isinstance(services, list):
                services = []
        except Exception:
            services = []
            row["services_json"] = "[]"

        subtotal = 0.0
        for it in services:
            if not isinstance(it, dict):
                continue
            qty = it.get("qty", 1) or 1
            try:
                qty = max(1, int(float(qty)))
            except Exception:
                qty = 1
            fee = to_float(it.get("fee", 0.0), 0.0)
            line_total = it.get("line_total", None)
            if line_total is None or str(line_total).strip() == "":
                line_total = fee * qty
            line = to_float(line_total, fee * qty)
            subtotal += line

        svc_fee = round(subtotal, 2)

        if (not (row.get("service_name") or "").strip()) and services:
            first = str((services[0] or {}).get("name", "") or "").strip()
            if first:
                row["service_name"] = first if len(services) == 1 else f"{first} +{len(services) - 1}"

        paid = _safe_money(row.get("paid_amount", 0))

        # Discount handling (validated server-side)
        disc_raw = (request.form.get("discount") if "discount" in request.form else row.get("discount", "")) or ""
        discount = _safe_money(disc_raw)
        if discount < 0:
            discount = 0.0
        if discount > svc_fee:
            discount = svc_fee
        net_fee = round(svc_fee - discount, 2)

        vat_calc = round(net_fee * float(VAT_RATE), 2)
        total_calc = round(net_fee + vat_calc, 2)
        due_calc2 = round(total_calc - paid, 2)

        row["service_fee"] = f"{svc_fee:.2f}" if svc_fee else ""
        row["discount"] = f"{discount:.2f}" if discount else ""
        row["fee_amount"] = f"{net_fee:.2f}" if (svc_fee or discount) else ""
        row["paid_amount"] = f"{paid:.2f}" if paid else ""
        row["due_amount"] = f"{max(due_calc2, 0):.2f}" if (svc_fee or paid) else ""

        if total_calc <= 0:
            row["payment_status"] = ""
        elif paid <= 0:
            row["payment_status"] = "Unpaid"
        elif paid + 0.0001 >= total_calc:
            row["payment_status"] = "Paid"
        else:
            row["payment_status"] = "Partial"

        append_row(BOOKINGS_XLSX, BOOKINGS_HEADERS, row)
        session["easy_booking_id"] = bid
        next_step = (request.form.get("next_step") or "reminder").lower().strip()
        if next_step == "invoice":
            session["easy_skip_reminder"] = "1"
            flash("Exam created. Continue to Invoice step.")
            return redirect(url_for("easy_invoice_step"))
        session.pop("easy_skip_reminder", None)
        flash("Exam created. Continue to Reminder step.")
        return redirect(url_for("easy_reminder_step"))

    mode = session.get("easy_mode", "")
    back_url = url_for("easy_new_pet") if mode == "new" else url_for("easy_old_pets")

    return render_template("easy_booking_step.html",
                           title=f"{APP_TITLE} | Easy Booking - Booking",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Easy Booking",
                           subtitle="Step 3: Create booking",
                           active="home",
                           owner=owner,
                           pet=pet,
                           default_start=_default_start_str(),
                           back_url=back_url,
                           statuses=STATUS_FLOW,
                           types=APPOINTMENT_TYPES,
                           priorities=PRIORITIES,
                           channels=CHANNELS,
                           payment_statuses=PAYMENT_STATUSES,
                           payment_methods=PAYMENT_METHODS,
                           reminder_channels=REMINDER_CHANNELS,
                           default_vet=session.get("username", ""),
                           vets=active_vet_names(),
                           rooms=active_room_names(),
                           services=active_services(),
                           vat_rate=float(VAT_RATE)
                           )


# -------------------------
# Route: Reminder Step
# -------------------------
@app.route("/easy/reminder", methods=["GET", "POST"])
def easy_reminder_step():
    gate = require_login()
    if gate:
        return gate

    booking_id = session.get("easy_booking_id")
    if not booking_id:
        flash("No active booking in Easy Flow. Please start again.")
        return redirect(url_for("easy_landing"))

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    ob = {row.get("id"): row for row in owners_rows}
    pb = {row.get("id"): row for row in pets_rows}

    b_raw = find_by_id(bookings_rows, booking_id)
    if not b_raw:
        flash("Booking not found.")
        return redirect(url_for("easy_landing"))

    owner = ob.get(b_raw.get("owner_id"), {})
    pet = pb.get(b_raw.get("pet_id"), {})
    b = decorate_booking(b_raw, ob, pb)

    # Services selected for this booking (source of truth)
    services = parse_services_json(b_raw.get("services_json", ""))
    if not services:
        # fallback for older bookings
        svc_name = str(b_raw.get("service_name", "") or "").strip()
        if svc_name:
            services = [{"name": svc_name, "fee": round(to_float(b_raw.get("service_fee"), 0.0), 2), "qty": 1,
                         "reminder_at": ""}]

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # 1) Master services management (global)
        if action == "master_add":
            name = (request.form.get("service_name") or "").strip()
            fee = round(to_float(request.form.get("service_fee"), 0.0), 2)
            if not name:
                flash("Service name is required.")
                return redirect(url_for("easy_reminder_step"))

            all_services = get_services(include_inactive=True)
            existing = next((s for s in all_services if str(s.get("name", "")).strip().lower() == name.lower()), None)

            if existing:
                update_row_by_id(
                    SERVICES_XLSX, SERVICES_HEADERS, existing.get("id"),
                    {"name": name, "fee": fee, "active": "1", "updated_at": now_str()}
                )
            else:
                append_row(
                    SERVICES_XLSX, SERVICES_HEADERS,
                    {
                        "id": str(uuid.uuid4()),
                        "name": name,
                        "fee": fee,
                        "active": "1",
                        "created_at": now_str(),
                        "updated_at": now_str(),
                    }
                )
            flash("Service added to master list.")
            return redirect(url_for("easy_reminder_step"))

        if action == "master_toggle":
            service_id = (request.form.get("service_id") or "").strip()
            set_active = (request.form.get("set_active") or "0").strip()
            if service_id:
                update_row_by_id(
                    SERVICES_XLSX, SERVICES_HEADERS, service_id,
                    {"active": "1" if set_active in ("1", "true", "True", "yes") else "0", "updated_at": now_str()}
                )
                flash("Service updated.")
            return redirect(url_for("easy_reminder_step"))

        # 2) Add service to current booking list
        if action == "booking_add":
            service_id = (request.form.get("service_id") or "").strip()
            qty_raw = (request.form.get("qty") or "1").strip()
            try:
                qty = int(float(qty_raw))
            except Exception:
                qty = 1
            if qty < 1:
                qty = 1

            all_services = get_services(include_inactive=True)
            srv = next((s for s in all_services if str(s.get("id", "")).strip() == service_id), None)
            if not srv or not _boolish(srv.get("active", "1")):
                flash("Selected service is not available.")
                return redirect(url_for("easy_reminder_step"))

            name = str(srv.get("name", "") or "").strip()
            fee = round(to_float(srv.get("fee", 0.0), 0.0), 2)

            found = next((it for it in services if str(it.get("name", "")).strip().lower() == name.lower()), None)
            if found:
                try:
                    found_qty = int(float(found.get("qty", 1)))
                except Exception:
                    found_qty = 1
                found["qty"] = max(1, found_qty + qty)
            else:
                services.append({"name": name, "fee": fee, "qty": qty, "reminder_at": ""})

            update_booking_services_and_amounts(booking_id, services)
            flash("Service added to this booking.")
            return redirect(url_for("easy_reminder_step"))

        # 3) Update per-service reminders OR remove a service from the booking list
        if action == "booking_update":
            remove_idx = request.form.get("remove_idx", "").strip()
            if remove_idx != "":
                try:
                    idx = int(remove_idx)
                except Exception:
                    idx = -1
                if 0 <= idx < len(services):
                    services.pop(idx)
                    update_booking_services_and_amounts(booking_id, services)
                    # keep reminders file consistent
                    upsert_service_reminders(b_raw, b, services)
                    flash("Service removed from this booking.")
                return redirect(url_for("easy_reminder_step"))

            # Save reminder datetime per service
            for i in range(len(services)):
                services[i]["reminder_at"] = normalize_dt(request.form.get(f"reminder_at_{i}", ""))
            update_booking_services_and_amounts(booking_id, services)
            upsert_service_reminders(b_raw, b, services)
            flash("Reminder dates saved.")
            return redirect(url_for("easy_reminder_step"))

        flash("No action performed.")
        return redirect(url_for("easy_reminder_step"))

    # Master services (global)
    master_services = []
    for s in get_services(include_inactive=True):
        s2 = dict(s)
        s2["active_bool"] = _boolish(s2.get("active", "1"))
        master_services.append(s2)

    # Active services for dropdown
    active_services_list = []
    for s in get_services(include_inactive=False):
        s2 = dict(s)
        s2["active_bool"] = True
        active_services_list.append(s2)

    # Decorate selected services for UI
    selected_services = []
    for it in (services or []):
        name = str(it.get("name", "") or "").strip()
        fee = round(to_float(it.get("fee", 0.0), 0.0), 2)
        try:
            qty = int(float(it.get("qty", 1)))
        except Exception:
            qty = 1
        if qty < 1:
            qty = 1
        reminder_at = normalize_dt(it.get("reminder_at", ""))
        selected_services.append({
            "name": name,
            "fee": fee,
            "qty": qty,
            "subtotal": round(qty * fee, 2),
            "reminder_at": reminder_at,
            "reminder_at_local": dt_to_local(reminder_at)
        })

    # Appointment reminder (WhatsApp) — configurable templates
    appt_msg = ""
    wa_link = ""
    if b.get("appointment_start"):
        svc_items = parse_services_json(b_raw.get("services_json", ""))
        svc_name = str(b_raw.get("service_name", "") or "").strip()
        if not svc_name and svc_items:
            first = str((svc_items[0] or {}).get("name", "") or "").strip()
            if first:
                svc_name = first if len(svc_items) == 1 else f"{first} +{len(svc_items) - 1}"
        details = format_booking_details(b_raw, services=svc_items)
        appt_msg = booking_message_template(
            b.get("owner_name", ""),
            b.get("pet_name", ""),
            b.get("appointment_start", ""),
            portal_link=b.get("portal_link", ""),
            service_name=svc_name,
            booking_details=details,
            booking_type=str(b.get("appointment_type", "Any") or "Any"),
        )
        wa_link = whatsapp_link(b.get("owner_phone", ""), appt_msg)

    return render_template(
        "easy_reminder_step.html",
        b=b,
        owner=owner,
        pet=pet,
        master_services=master_services,
        active_services=active_services_list,
        selected_services=selected_services,
        subtotal=services_subtotal(services),
        vat_rate=VAT_RATE,
        appt_msg=appt_msg,
        wa_link=wa_link
    )


@app.route("/easy/reminder/<booking_id>/mark_sent")
def easy_reminder_mark_sent(booking_id):
    gate = require_login()
    if gate: return gate

    # Create reminder record + update booking.reminder_sent
    bookings_rows = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings_rows, booking_id)
    if not b:
        flash("Exam not found.")
        return redirect(url_for("easy_reminder_step"))

    # update booking reminder_sent
    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {
        "reminder_sent": now_str(),
        "updated_at": now_str()
    })

    # add reminder row
    rid = str(uuid.uuid4())
    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}
    db = decorate_booking(b, ob, pb)

    token = get_or_create_portal_token(str(b.get("id", "")))
    portal = url_for("portal", token=token, _external=True)
    msg = booking_message_template(db.get("owner_name", ""), db.get("pet_name", ""), db.get("appointment_start", ""),
                                   portal_link=portal)

    append_row(REMINDERS_XLSX, REMINDERS_HEADERS, {
        "id": rid,
        "booking_id": str(b.get("id", "")),
        "owner_id": str(b.get("owner_id", "")),
        "pet_id": str(b.get("pet_id", "")),
        "channel": str(b.get("reminder_channel", "WhatsApp") or "WhatsApp"),
        "status": "Sent",
        "scheduled_for": normalize_dt(str(b.get("appointment_start", ""))),
        "opened_at": "",
        "sent_at": now_str(),
        "message": msg,
        "created_at": now_str(),
        "updated_at": now_str()
    })

    flash("Reminder marked as Sent.")
    return redirect(url_for("easy_reminder_step"))


# -------------------------
# Route: Invoice Step + Printable Invoice
# -------------------------
def ensure_booking_invoice_no(booking_id: str) -> str:
    bookings = read_all(BOOKINGS_XLSX)
    b = find_by_id(bookings, booking_id)
    if not b:
        return ""
    inv = str(b.get("invoice_no", "") or "").strip()
    if inv:
        return inv
    dt = parse_dt(str(b.get("appointment_start", ""))) or datetime.now()
    inv = f"INV-{dt.strftime('%Y%m%d')}-{str(booking_id)[:6].upper()}"
    update_row_by_id(BOOKINGS_XLSX, BOOKINGS_HEADERS, booking_id, {"invoice_no": inv, "updated_at": now_str()})
    return inv


@app.route("/easy/invoice")
def easy_invoice_step():
    gate = require_login()
    if gate: return gate

    must = easy_require(["easy_booking_id"])
    if must: return must

    booking_id = session["easy_booking_id"]
    inv = ensure_booking_invoice_no(booking_id)

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    ob = {str(o["id"]): o for o in owners_rows}
    pb = {str(p["id"]): p for p in pets_rows}

    b_raw = find_by_id(bookings_rows, booking_id)
    if not b_raw:
        flash("Exam not found.")
        return redirect(url_for("easy_booking_step"))

    b = decorate_booking(b_raw, ob, pb)
    owner = ob.get(str(b_raw.get("owner_id", "")), {})
    owner_sn = str(owner.get("customer_sn", "") or "").strip()

    subtotal = round(to_float(b_raw.get("service_fee"), 0.0), 2)
    discount_amt = round(to_float(b_raw.get("discount"), 0.0), 2)
    discount_type = normalize_discount_type(b_raw.get("discount_type"))
    discount_value = str(b_raw.get("discount_value", "") or "").strip()
    paid_amt = round(to_float(b_raw.get("paid_amount"), 0.0), 2)
    due_amt = round(to_float(b_raw.get("due_amount"), 0.0), 2)
    net_fee = round(to_float(b_raw.get("fee_amount"), 0.0), 2)
    fee = net_fee
    skip_reminder = str(session.get("easy_skip_reminder", "")).lower() in ("1", "true", "yes")
    back_url = url_for("easy_booking_step") if skip_reminder else url_for("easy_reminder_step")
    return render_template("easy_invoice_step.html",
                           title=f"{APP_TITLE} | Easy Booking - Invoice",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Easy Booking",
                           subtitle="Step 5: Invoice",
                           active="home",
                           b=b,
                           invoice_no=inv,
                           owner_sn=owner_sn,
                           fee=fee,
                           subtotal=subtotal,
                           discount_amt=discount_amt,
                           discount_type=discount_type,
                           discount_value=discount_value,
                           paid_amt=paid_amt,
                           due_amt=due_amt,
                           net_fee=net_fee,
                           payment_channel=str(b_raw.get("payment_channel", "") or ""),
                           back_url=back_url,
                           skip_reminder=skip_reminder
                           )


@app.route("/easy/invoice/print/<booking_id>")
def easy_invoice_print(booking_id):
    gate = require_login()
    if gate:
        return gate

    ensure_owner_sn_schema()
    inv = ensure_booking_invoice_no(booking_id)

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)

    b = find_by_id(bookings_rows, booking_id)
    if not b:
        return Response("Booking not found.", status=404)

    owner = find_by_id(owners_rows, str(b.get("owner_id", "")))
    pet = find_by_id(pets_rows, str(b.get("pet_id", "")))

    # Line items from services_json
    services = parse_services_json(b.get("services_json", ""))
    line_items = []
    if services:
        for it in services:
            if not isinstance(it, dict):
                continue
            name = str(it.get("name") or "").strip()
            if not name:
                continue
            try:
                qty = max(1, int(float(it.get("qty", 1) or 1)))
            except Exception:
                qty = 1
            unit = round(to_float(it.get("fee", 0.0), 2), 2)
            lt = it.get("line_total")
            if lt is None or str(lt).strip() == "":
                lt = unit * qty
            lt = round(to_float(lt, 0.0), 2)
            line_items.append({"desc": name, "qty": qty, "unit_price": unit, "line_total": lt})

    if not line_items:
        # legacy single service fallback
        desc = str(b.get("service_name", "") or b.get("appointment_type", "") or "Service").strip()
        lt = round(to_float(b.get("service_fee", 0.0), 2), 2)
        if not desc:
            desc = "Service"
        if lt == 0:
            lt = round(to_float(b.get("fee_amount", 0.0), 2), 2)
        line_items = [{"desc": desc, "qty": 1, "unit_price": lt, "line_total": lt}]

    subtotal = round(sum(round(to_float(x.get("line_total"), 0.0), 2) for x in line_items), 2)

    discount_type = normalize_discount_type(b.get("discount_type") or "value")
    discount_value = str(b.get("discount_value") or "").strip()
    discount_amt = round(compute_discount_amount(subtotal, discount_type, discount_value, b.get("discount")), 2)

    net_fee = round(subtotal - discount_amt, 2)
    vat = round(net_fee * float(VAT_RATE), 2)
    total = round(net_fee + vat, 2)

    paid_amt = round(to_float(b.get("paid_amount", 0.0), 2), 2)
    due_amt = round(total - paid_amt, 2)
    if due_amt < 0:
        due_amt = 0.0

    return render_template(
        "easy_invoice_print.html",
        app_title=APP_TITLE,
        invoice_no=inv,
        today=date.today().strftime("%Y-%m-%d"),
        owner_name=safe_get(owner or {}, "owner_name"),
        owner_phone=safe_get(owner or {}, "phone"),
        owner_sn=str((owner or {}).get("customer_sn", "") or "").strip(),
        owner_address=safe_get(owner or {}, "address"),
        pet_name=safe_get(pet or {}, "pet_name"),
        appt_start=normalize_dt(str(b.get("appointment_start", ""))),
        appt_type=str(b.get("appointment_type", "") or ""),
        status=str(b.get("status", "") or ""),
        reason=str(b.get("reason", "") or ""),
        notes=str(b.get("notes", "") or ""),
        payment_channel=str(b.get("payment_channel", "") or ""),
        payment_method=str(b.get("payment_method", "") or ""),
        payment_status=str(b.get("payment_status", "") or ""),
        line_items=line_items,
        subtotal=f"{subtotal:.2f}",
        discount_amt=f"{discount_amt:.2f}",
        discount_type=discount_type,
        discount_value=discount_value,
        net_fee=f"{net_fee:.2f}",
        vat=f"{vat:.2f}",
        total=f"{total:.2f}",
        paid=f"{paid_amt:.2f}",
        due=f"{due_amt:.2f}",
        vat_rate=float(VAT_RATE),
    )


# =====================================================================
# END EASY BOOKING WIZARD ADD-ON
# =====================================================================
# 20 services (edit freely)
SERVICE_CATALOG = [
    {"code": "EXAM", "name": "Examination", "price": 200},
    {"code": "CONS", "name": "Consultation", "price": 150},
    {"code": "VAC1", "name": "Core Vaccine", "price": 250},
    {"code": "RAB", "name": "Rabies Vaccine", "price": 200},
    {"code": "DEW", "name": "Deworming", "price": 120},
    {"code": "FT", "name": "Flea & Tick Treatment", "price": 180},
    {"code": "MIC", "name": "Microchipping", "price": 350},
    {"code": "NAIL", "name": "Nail Trimming", "price": 60},
    {"code": "EAR", "name": "Ear Cleaning", "price": 80},
    {"code": "GROOM", "name": "Basic Grooming", "price": 220},
    {"code": "CBC", "name": "Blood Test (CBC)", "price": 220},
    {"code": "BIO", "name": "Biochemistry Panel", "price": 320},
    {"code": "UR", "name": "Urine Analysis", "price": 150},
    {"code": "XR", "name": "X-Ray", "price": 400},
    {"code": "US", "name": "Ultrasound", "price": 600},
    {"code": "WOUND", "name": "Wound Dressing", "price": 120},
    {"code": "DENT", "name": "Dental Cleaning", "price": 800},
    {"code": "MINOR", "name": "Minor Procedure", "price": 500},
    {"code": "HOSP", "name": "Day Hospitalization", "price": 300},
    {"code": "MED", "name": "Medication / Prescription", "price": 100},
]
_SERVICE_BY_CODE = {s["code"]: s for s in SERVICE_CATALOG}

_SERV_SESS_KEY = "services_by_booking"


def get_booking_services(booking_id: str) -> dict:
    return (session.get(_SERV_SESS_KEY, {}) or {}).get(str(booking_id), {"items": [], "notes": ""})


def services_saved_for_booking(booking_id: str) -> bool:
    data = get_booking_services(booking_id)
    return bool(data.get("items"))


# Make helpers available to templates (optional)
@app.context_processor
def _inject_services_window_globals():
    return dict(
        SERVICE_CATALOG=SERVICE_CATALOG,
        get_booking_services=get_booking_services,
        services_saved_for_booking=services_saved_for_booking,
    )


# --- Popup Window Page (GET) ---
@app.route("/booking/<booking_id>/services/window", methods=["GET"])
def booking_services_window(booking_id):
    # Decide where invoice is in YOUR app:
    # Primary (wizard): easy_booking step=5
    try:
        invoice_url = url_for("easy_booking", step=5, booking_id=booking_id)
    except Exception:
        # If you have another endpoint, replace here:
        # invoice_url = url_for("invoice", booking_id=booking_id)
        invoice_url = "/"

    existing = get_booking_services(booking_id)
    existing_codes = {it.get("code") for it in existing.get("items", []) if isinstance(it, dict)}
    existing_qty = {it.get("code"): int(it.get("qty", 1)) for it in existing.get("items", []) if isinstance(it, dict)}

    # Standalone HTML (popup window)
    html = r"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Services & Notes</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial;background:#0b1220;color:#e5e7eb}
    .wrap{max-width:1100px;margin:18px auto;padding:0 14px}
    .card{border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.03);border-radius:16px;box-shadow:0 18px 50px rgba(0,0,0,.45);overflow:hidden}
    .head{display:flex;align-items:center;justify-content:space-between;padding:14px 16px;border-bottom:1px solid rgba(255,255,255,.10)}
    .title{font-weight:800;font-size:16px}
    .sub{color:#94a3b8;font-size:12px;margin-top:3px}
    .body{padding:14px 16px}
    .grid{display:grid;grid-template-columns: 1.6fr .9fr;gap:14px}
    @media (max-width: 900px){.grid{grid-template-columns:1fr}}
    table{width:100%;border-collapse:collapse}
    th,td{padding:10px;border-bottom:1px solid rgba(255,255,255,.08);font-size:13px}
    th{color:#cbd5e1;text-align:left}
    .qty{width:70px;padding:7px;border-radius:10px;border:1px solid rgba(255,255,255,.16);background:rgba(255,255,255,.04);color:#e5e7eb}
    .note{width:100%;min-height:230px;padding:10px;border-radius:12px;border:1px solid rgba(255,255,255,.16);background:rgba(255,255,255,.04);color:#e5e7eb}
    .summary{margin-top:10px;padding:10px;border-radius:12px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.03)}
    .row{display:flex;justify-content:space-between;gap:10px}
    .actions{display:flex;gap:10px;justify-content:flex-end;margin-top:14px}
    .btn{padding:10px 14px;border-radius:12px;border:1px solid rgba(255,255,255,.16);background:rgba(255,255,255,.05);color:#e5e7eb;cursor:pointer}
    .btn.primary{background:#1f6feb}
    .pill{display:inline-block;padding:6px 10px;border-radius:999px;border:1px solid rgba(255,255,255,.14);color:#cbd5e1;font-size:12px}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="head">
        <div>
          <div class="title">Select Services + Notes</div>
          <div class="sub">Booking: <span class="pill">{{ booking_id }}</span></div>
        </div>
        <div class="pill">Save will open Invoice and close this window</div>
      </div>

      <div class="body">
        <form id="svcForm" method="post" action="{{ url_for('booking_services_window_save', booking_id=booking_id) }}">
          <input type="hidden" name="services_json" id="services_json">
          <input type="hidden" name="invoice_url" value="{{ invoice_url }}">

          <div class="grid">
            <div>
              <table id="svcTable">
                <thead>
                  <tr>
                    <th style="width:36px;"></th>
                    <th>Service</th>
                    <th style="width:110px;">Price</th>
                    <th style="width:120px;">Qty</th>
                    <th style="width:130px;">Line Total</th>
                  </tr>
                </thead>
                <tbody>
                  {% for s in services %}
                  <tr data-code="{{ s.code }}" data-price="{{ s.price }}">
                    <td><input type="checkbox" class="chk" {% if s.code in existing_codes %}checked{% endif %}></td>
                    <td>{{ s.name }}</td>
                    <td>{{ "%.2f"|format(s.price) }}</td>
                    <td>
                      <input type="number" class="qty" min="1"
                        value="{{ existing_qty.get(s.code, 1) }}">
                    </td>
                    <td class="line">0.00</td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>

            <div>
              <div style="font-weight:800;color:#cbd5e1;margin-bottom:8px;">Notes</div>
              <textarea class="note" name="notes" placeholder="Write notes here...">{{ existing.notes or "" }}</textarea>

              <div class="summary">
                <div class="row">
                  <div>Total</div>
                  <div id="total">0.00</div>
                </div>
                <div style="margin-top:6px;color:#94a3b8;font-size:12px;">
                  Tip: Select services, set quantities, then Save.
                </div>
              </div>

              <div class="actions">
                <button type="button" class="btn" onclick="window.close()">Cancel</button>
                <button type="submit" class="btn primary">Save → Invoice</button>
              </div>
            </div>
          </div>
        </form>
      </div>
    </div>
  </div>

<script>
(function(){
  const tbl = document.getElementById("svcTable");
  const totalEl = document.getElementById("total");
  const out = document.getElementById("services_json");
  const form = document.getElementById("svcForm");

  function recalc(){
    let total = 0;
    tbl.querySelectorAll("tbody tr").forEach(tr=>{
      const chk = tr.querySelector(".chk");
      const qtyEl = tr.querySelector(".qty");
      const lineEl = tr.querySelector(".line");
      const price = parseFloat(tr.getAttribute("data-price") || "0");
      const qty = Math.max(1, parseInt(qtyEl.value || "1", 10));

      if(chk.checked){
        const line = price * qty;
        lineEl.textContent = line.toFixed(2);
        total += line;
      } else {
        lineEl.textContent = "0.00";
      }
    });
    totalEl.textContent = total.toFixed(2);
  }

  tbl.addEventListener("change", function(e){
    if(e.target.classList.contains("chk") || e.target.classList.contains("qty")) recalc();
  });

  form.addEventListener("submit", function(){
    const items = [];
    tbl.querySelectorAll("tbody tr").forEach(tr=>{
      const chk = tr.querySelector(".chk");
      if(!chk.checked) return;
      const code = tr.getAttribute("data-code");
      const price = parseFloat(tr.getAttribute("data-price") || "0");
      const qty = Math.max(1, parseInt(tr.querySelector(".qty").value || "1", 10));
      items.push({code: code, qty: qty, unit_price: price});
    });
    out.value = JSON.stringify(items);
  });

  // initial calc (supports pre-selected)
  recalc();
})();
</script>

</body>
</html>
"""
    return render_template_string(
        html,
        booking_id=str(booking_id),
        services=SERVICE_CATALOG,
        existing=existing,
        existing_codes=existing_codes,
        existing_qty=existing_qty,
        invoice_url=invoice_url,
    )


# --- Save from Popup (POST) ---
@app.route("/booking/<booking_id>/services/window/save", methods=["POST"])
def booking_services_window_save(booking_id):
    raw = request.form.get("services_json", "[]")
    notes = (request.form.get("notes") or "").strip()
    invoice_url = request.form.get("invoice_url") or "/"

    try:
        items = json.loads(raw) if raw else []
        if not isinstance(items, list):
            items = []
    except Exception:
        items = []

    clean = []
    for it in items:
        if not isinstance(it, dict):
            continue
        code = str(it.get("code", "")).strip().upper()
        if code not in _SERVICE_BY_CODE:
            continue
        try:
            qty = int(it.get("qty", 1))
        except Exception:
            qty = 1
        qty = max(1, qty)

        svc = _SERVICE_BY_CODE[code]
        unit_price = float(svc["price"])
        line_total = round(unit_price * qty, 2)

        clean.append({
            "code": code,
            "name": svc["name"],
            "qty": qty,
            "unit_price": unit_price,
            "line_total": line_total,
        })

    store = session.get(_SERV_SESS_KEY, {}) or {}
    store[str(booking_id)] = {
        "items": clean,
        "notes": notes,
        "saved_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    session[_SERV_SESS_KEY] = store
    session.modified = True

    # Return a tiny page that redirects MAIN window to invoice, then closes popup
    done = r"""
<!doctype html>
<html><head><meta charset="utf-8"><title>Saved</title></head>
<body style="font-family:system-ui;background:#0b1220;color:#e5e7eb;padding:20px;">
  Saved. Redirecting to invoice...
  <script>
    (function(){
      var invoiceUrl = {{ invoice_url|tojson }};
      try{
        if(window.opener && !window.opener.closed){
          window.opener.location.href = invoiceUrl;
          window.close();
          return;
        }
      }catch(e){}
      window.location.href = invoiceUrl;
    })();
  </script>
</body></html>
"""
    return render_template_string(done, invoice_url=invoice_url)


# -------------------------
# Route: Pet History (for History button)
# -------------------------
@app.route("/pet/history/<pet_id>")
def pet_history(pet_id):
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)

    pet = find_by_id(pets_rows, str(pet_id))
    if not pet:
        if request.args.get("partial") == "1":
            return '<div class="flash">Pet not found.</div>'
        flash("Pet not found.")
        return redirect(url_for("home"))

    owner = find_by_id(owners_rows, str(pet.get("owner_id", "")))
    if not owner:
        owner = {"owner_name": "Unknown", "phone": ""}

    # Filter & sort
    items_raw = [b for b in bookings_rows if str(b.get("pet_id", "")) == str(pet_id)]

    def _sort_key(x):
        return parse_dt(str(x.get("appointment_start", ""))) or datetime.min

    items_raw.sort(key=_sort_key, reverse=True)

    items = []
    for r in items_raw:
        services_disp = []
        raw = (r.get("services_json") or "").strip() or "[]"
        try:
            lst = json.loads(raw)
            if not isinstance(lst, list):
                lst = []
        except Exception:
            lst = []
        for it in lst:
            if not isinstance(it, dict):
                continue
            name = str(it.get("name", "") or "").strip()
            qty = it.get("qty", 1) or 1
            try:
                qty = max(1, int(float(qty)))
            except Exception:
                qty = 1
            fee = to_float(it.get("fee", 0.0), 0.0)
            line = to_float(it.get("line_total", fee * qty), fee * qty)
            if name:
                if qty != 1:
                    services_disp.append(f"{name} — {qty} x {fee:.2f} = {line:.2f}")
                else:
                    services_disp.append(f"{name} — {fee:.2f}")
        total = to_float(r.get("fee_amount", 0), 0.0)
        paid = to_float(r.get("paid_amount", 0), 0.0)
        due = to_float(r.get("due_amount", max(0.0, total - paid)), max(0.0, total - paid))
        dt = str(r.get("appointment_start", "") or "").strip()
        items.append({
            "dt": dt,
            "status": str(r.get("status", "") or ""),
            "vet": str(r.get("vet_name", "") or ""),
            "room": str(r.get("room", "") or ""),
            "services": services_disp,
            "reason": str(r.get("reason", "") or ""),
            "total": f"{total:.2f}",
            "paid": f"{paid:.2f}",
            "due": f"{due:.2f}",
            "weight": str(r.get("visit_weight_kg", "") or "").strip(),
            "temp": str(r.get("visit_temp_c", "") or "").strip(),
        })

    back_url = request.args.get("back") or request.referrer or url_for("home")
    if request.args.get("partial") == "1":
        return render_template("pet_history_partial.html", owner=owner, pet=pet, items=items)
    return render_template("pet_history.html",
                           title=f"{APP_TITLE} | Pet History",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Reports",
                           subtitle="Pet History",
                           active="report",
                           owner=owner,
                           pet=pet,
                           items=items,
                           back_url=back_url
                           )


@app.route("/pet/reminders/<pet_id>")
def pet_reminders(pet_id):
    gate = require_login()
    if gate: return gate

    owners_rows = read_all(OWNERS_XLSX)
    pets_rows = read_all(PETS_XLSX)
    bookings_rows = read_all(BOOKINGS_XLSX)
    rem_rows = read_all(REMINDERS_XLSX)

    pet = find_by_id(pets_rows, str(pet_id))
    if not pet:
        if request.args.get("partial") == "1":
            return '<div class="flash">Pet not found.</div>'
        flash("Pet not found.")
        return redirect(url_for("home"))

    owner = find_by_id(owners_rows, str(pet.get("owner_id", "")))
    if not owner:
        owner = {"owner_name": "Unknown", "phone": "", "email": ""}

    # Reminders for this pet
    reminders = []
    for r in rem_rows:
        if str(r.get("pet_id", "")) == str(pet_id):
            msg = str(r.get("message", "") or "")
            reminders.append({
                **r,
                "whatsapp_link": whatsapp_link(owner.get("phone", ""), msg)
            })
    reminders.sort(key=lambda x: (str(x.get("scheduled_for", "")) or str(x.get("created_at", ""))), reverse=True)

    # Upcoming bookings for this pet
    now_dt = datetime.now()
    upcoming = []
    for b in bookings_rows:
        if str(b.get("pet_id", "")) != str(pet_id):
            continue
        dt = parse_dt(str(b.get("appointment_start", "")))
        if dt and dt >= now_dt and str(b.get("status", "")) not in ("Cancelled", "No-Show"):
            upcoming.append(b)
    upcoming.sort(key=lambda x: parse_dt(str(x.get("appointment_start", ""))) or datetime.max)
    upcoming = upcoming[:10]

    back_url = request.args.get("back") or request.referrer or url_for("home")

    if request.args.get("partial") == "1":
        return render_template("pet_reminders_partial.html", owner=owner, pet=pet, reminders=reminders,
                               upcoming=upcoming, show_header=True)

    return render_template("pet_reminders.html",
                           title=f"{APP_TITLE} | Pet Reminders",
                           app_title=APP_TITLE,
                           logged_in=True,
                           header="Reminders",
                           subtitle="Pet Reminders",
                           active="reminders",
                           owner=owner,
                           pet=pet,
                           reminders=reminders,
                           upcoming=upcoming,
                           show_header=False,
                           back_url=back_url
                           )


# =========================
# STARTUP
# =========================
def main():
    write_templates()
    init_storage()
    if "--seed" in sys.argv:
        seed_demo_data(10)
        print("Seeded demo data (only if files were empty).")
    app.run(debug=True)


if __name__ == "__main__":
    main()
